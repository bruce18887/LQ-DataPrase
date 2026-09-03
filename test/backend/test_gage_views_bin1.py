"""Gage view ``only_bin1`` test (defect 9).

The old view used ``pd.to_numeric(df[bin_col], errors='coerce') == 1`` which
turns text bins ('Bin1' / 'BIN 1') into NaN -> False, silently emptying the
whole frame. The fix routes through ``apps.analysis.services.statistics.
filter_bin1_rows`` (imported read-only). This drives the real endpoint with a
mocked parser returning a text-bin frame and reads the generated workbook back.

Red before fix: with the old filter the frame is emptied, the builder finds no
numeric test column, and the global-mean cell O12 is blank instead of 1.5.
"""
import io
import os
from unittest import mock

import pandas as pd
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from apps.datafiles.models import DataFile

User = get_user_model()

SAMPLE = os.path.join(os.path.dirname(__file__), '..', '..', 'Data',
                      'SampleData', 'Gage', 'gage_m_S1.csv')
GAGE_URL = '/api/v1/gage/generate_summary/'


def _response_bytes(resp):
    if hasattr(resp, 'streaming_content'):
        return b''.join(resp.streaming_content)
    return resp.content


class GageOnlyBin1Tests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='gagebin1', password='pw')
        self.client.force_authenticate(self.user)
        self.files = []
        for i in range(2):
            self.files.append(DataFile.objects.create(
                owner=self.user, filename=f'bin1_{i}.csv',
                file_path=SAMPLE,
                file_size=os.path.getsize(SAMPLE) if os.path.exists(SAMPLE) else 1,
                format_type='CTA8290D', status='ready',
            ))

    def test_text_bin_rows_are_preserved(self):
        # pass bins written as text; only Bin2 is a fail row
        df = pd.DataFrame({
            'SW_Bin': ['Bin1', 'BIN 1', 'Bin2'],
            'T1': [1.0, 2.0, 3.0],
        })
        meta = {'format': 'CTA8290D', 'mins': {'T1': '0'},
                'maxs': {'T1': '100'}, 'units': {}}
        with mock.patch('apps.gage.views.get_cached_parsed_file',
                        return_value=(df, meta, 'CTA8290D')):
            resp = self.client.post(
                GAGE_URL,
                {'file_ids': [f.id for f in self.files],
                 'only_bin1': True, 'ignore_no_limit': False},
                format='json',
            )
        self.assertEqual(resp.status_code, 200,
                         getattr(resp, 'data', ''))
        ws = load_workbook(io.BytesIO(_response_bytes(resp)))['Summary']
        # both files keep T1 = [1.0, 2.0] (the two pass rows) -> global mean 1.5
        self.assertEqual(ws['C12'].value, 'T1')
        self.assertAlmostEqual(float(ws['O12'].value), 1.5, delta=1e-4)
