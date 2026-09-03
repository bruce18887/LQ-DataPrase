"""导出统计口径回归（apps/export）。

覆盖缺陷：

* **#1 ddof 口径 + 二次舍入**：``export_xlsx_optimized`` 曾用 pandas 默认
  ``std()``（ddof=1）算 STD/CPK，而屏幕侧 ``computations.compute_range_statistics``
  / ``histogram`` / buyoff / multi_lot 全是 ``ddof=0``（n=10 时 σ 差 5.4%）；
  且把已 ``round(…, 4)`` 的 Avg/STD 再喂给 ``compute_cpk``，窄分布上
  STD 直接归零 → CPK 恒 0。
* **#2 Range 口径**：表格里 Min/Max 各自 round 到 4 位，Range 必须由这两个
  已展示的值相减，否则用户看到 ``Range ≠ Max − Min``。
* **#10 inf 不进单元格**：``excel_builders`` 曾用 ``dropna()``（不滤 inf），
  含 inf 的列 mean=inf / std=nan，``round(nan, 4)`` 被 excelize 写成文本 ``'NaN'``。
* **#11 数值列白名单**：``dtype in ('int64','float64')`` 漏掉 int32/float32/UInt8；
  bool（真实数据的 ``Dut_Pass``）不是测量值，必须显式排除。
* **pp/ppk 护栏**：``compute_cpk`` 的 pp/ppk 字段将被删除，导出侧不得再消费。

运行：``manage.py test test.backend.test_export_stats_consistency``
"""

import io
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import excelize
from django.test import SimpleTestCase

from apps.analysis.services.statistics import (
    compute_cpk,
    compute_range_statistics,
    filter_finite,
)
from apps.export import excelize_helpers
from apps.export.excel_builders import build_sigma_limit_sheet
from apps.export.export_xlsx_optimized import export_to_xlsx_optimized

STATS_ROW = {'Min': 5, 'Avg': 6, 'Max': 7, 'Range': 8, 'STD': 9, 'CPK': 10}
PARAM_COL = 2          # 第 1 列（A）被统计行标签占用，测试参数固定在第 2 列
EMPTY_META = {'units': {}, 'mins': {}, 'maxs': {}, 'format': 'CTA8290D'}


def _stats_ws(df, metadata):
    """导出 xlsx 并返回 Data sheet（统计区在行 5-10）。"""
    buf = export_to_xlsx_optimized(df, metadata)
    return openpyxl.load_workbook(io.BytesIO(buf))['Data']


def _cell(ws, label, col=PARAM_COL):
    return ws.cell(STATS_ROW[label], col).value


class XlsxStdDdofTests(SimpleTestCase):
    """缺陷 #1：STD/CPK 必须用 ddof=0（总体标准差），与屏幕统计卡一致。"""

    VALUES = [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0]   # n=10

    def setUp(self):
        self.df = pd.DataFrame({'SW_Bin': [1] * 10, 'P': self.VALUES})
        self.metadata = {'units': {'P': 'V'}, 'mins': {'P': '0'},
                         'maxs': {'P': '12'}, 'format': 'CTA8290D'}
        self.ws = _stats_ws(self.df, self.metadata)
        self.series = pd.Series(self.VALUES, dtype=float)

    def test_ddof0_and_ddof1_are_distinguishable_at_4_decimals(self):
        """用例前置：n=10 时两种 ddof 的 σ 在 4 位小数上必须不同（2.8723 vs 3.0277）。"""
        self.assertNotEqual(round(float(self.series.std(ddof=0)), 4),
                            round(float(self.series.std(ddof=1)), 4))

    def test_std_cell_uses_population_std(self):
        expected = round(float(self.series.std(ddof=0)), 4)
        self.assertAlmostEqual(_cell(self.ws, 'STD'), expected, places=6,
                               msg='导出 STD 应为 ddof=0 的总体标准差')

    def test_std_cell_matches_screen_statistics(self):
        """导出 STD == 屏幕 compute_range_statistics 的 std（同一文件同一参数）。"""
        stats = compute_range_statistics(filter_finite(self.series), self.metadata, 'P')
        self.assertAlmostEqual(_cell(self.ws, 'STD'), round(stats['std'], 4), places=6)

    def test_cpk_cell_matches_screen_cpk(self):
        """导出 CPK == 屏幕 CPK（未舍入 mean/std + ddof=0）。"""
        stats = compute_range_statistics(filter_finite(self.series), self.metadata, 'P')
        expected = round(compute_cpk(stats['mean'], stats['std'], 0.0, 12.0)['cpk'], 4)
        self.assertAlmostEqual(_cell(self.ws, 'CPK'), expected, places=6)


class XlsxCpkDoubleRoundingTests(SimpleTestCase):
    """缺陷 #1（二次舍入）：CPK 必须由未舍入的 mean/std 计算。"""

    # 极差 ~1e-4 的窄分布：std(ddof=0) ≈ 3e-5 → round(…, 4) == 0.0
    TINY = [0.12340, 0.12345, 0.12350, 0.12342, 0.12348,
            0.12344, 0.12346, 0.12350, 0.12343, 0.12347]

    def setUp(self):
        self.df = pd.DataFrame({'SW_Bin': [1] * len(self.TINY), 'P': self.TINY})
        self.metadata = {'units': {'P': 'V'}, 'mins': {'P': '0.123'},
                         'maxs': {'P': '0.124'}, 'format': 'CTA8290D'}
        self.series = pd.Series(self.TINY, dtype=float)

    def test_precondition_rounded_std_is_zero(self):
        """用例前置：STD 单元格显示 0.0000，旧代码把 0 喂给 compute_cpk → CPK 恒 0。"""
        self.assertEqual(round(float(self.series.std(ddof=0)), 4), 0.0)

    def test_cpk_not_computed_from_rounded_std(self):
        raw_mean = float(self.series.mean())
        raw_std = float(self.series.std(ddof=0))
        expected = round(compute_cpk(raw_mean, raw_std, 0.123, 0.124)['cpk'], 4)
        self.assertGreater(expected, 1.0, '用例前置：真实 CPK 应远大于 0')
        ws = _stats_ws(self.df, self.metadata)
        self.assertAlmostEqual(_cell(ws, 'CPK'), expected, places=6,
                               msg='CPK 应由未舍入 mean/std 计算，而非 round 后的 0')


class XlsxRangeConsistencyTests(SimpleTestCase):
    """缺陷 #2：Range 必须等于表格里展示的 Max − Min。"""

    VALUES = [1.00006, 2.00004, 1.5]

    def setUp(self):
        self.df = pd.DataFrame({'SW_Bin': [1] * 3, 'P': self.VALUES})
        self.ws = _stats_ws(self.df, dict(EMPTY_META))

    def test_precondition_two_range_calibers_differ(self):
        """用例前置：未舍入 max−min → 1.0；已舍入 Max−Min → 0.9999。"""
        raw = round(max(self.VALUES) - min(self.VALUES), 4)
        displayed = round(round(max(self.VALUES), 4) - round(min(self.VALUES), 4), 4)
        self.assertNotEqual(raw, displayed)

    def test_range_equals_displayed_max_minus_min(self):
        min_cell = _cell(self.ws, 'Min')
        max_cell = _cell(self.ws, 'Max')
        range_cell = _cell(self.ws, 'Range')
        self.assertAlmostEqual(range_cell, round(max_cell - min_cell, 4), places=9,
                               msg='Range 必须与同一张表里的 Min/Max 自洽')


class SigmaLimitInfTests(SimpleTestCase):
    """缺陷 #10：inf 数据不得让 nan/inf 落进单元格（excel_builders）。"""

    def _build(self, df, metadata, sigma=3):
        f = excelize.new_file()

        def _close():
            try:
                f.close()      # save_excelize 正常路径已 close，重复 close 忽略
            except Exception:  # noqa: BLE001
                pass

        self.addCleanup(_close)
        build_sigma_limit_sheet(f, df, metadata, sigma, False)
        buf = excelize_helpers.save_excelize(f)
        return openpyxl.load_workbook(io.BytesIO(buf))['TestItem_Limit']

    def test_inf_row_yields_finite_sigma_limits(self):
        df = pd.DataFrame({'SW_Bin': [1] * 4, 'V1': [1.0, 2.0, np.inf, 4.0]})
        metadata = {'units': {}, 'mins': {'V1': '0'}, 'maxs': {'V1': '10'},
                    'format': 'CTA8290D'}
        ws = self._build(df, metadata)
        self.assertEqual(ws.cell(2, 2).value, 'V1')
        finite = pd.Series([1.0, 2.0, 4.0])
        mean = float(finite.mean())
        std = float(finite.std(ddof=0))
        for col, expected in ((5, mean - 3 * std), (6, mean + 3 * std)):
            value = ws.cell(2, col).value
            self.assertIsInstance(value, (int, float),
                                  msg=f'第 {col} 列应为数值，实际 {value!r}（nan 被写成文本）')
            self.assertAlmostEqual(value, round(expected, 4), places=6)

    def test_all_inf_column_is_skipped_not_written_as_nan(self):
        df = pd.DataFrame({'SW_Bin': [1] * 3, 'V1': [np.inf, np.inf, -np.inf]})
        metadata = {'units': {}, 'mins': {'V1': '0'}, 'maxs': {'V1': '10'},
                    'format': 'CTA8290D'}
        ws = self._build(df, metadata)
        for row in range(2, ws.max_row + 1):
            for col in (5, 6):
                self.assertNotEqual(str(ws.cell(row, col).value).lower(), 'nan')


class NumericColumnWhitelistTests(SimpleTestCase):
    """缺陷 #11：数值参数筛选必须覆盖窄 dtype，且显式排除 bool。"""

    def test_narrow_numeric_dtypes_are_measurable(self):
        from apps.export.columns import is_measurable_numeric
        for dtype in ('int64', 'float64', 'int32', 'float32', 'int16', 'UInt8'):
            with self.subTest(dtype=dtype):
                self.assertTrue(is_measurable_numeric(pd.Series([1, 2, 3], dtype=dtype)))

    def test_bool_is_not_a_measurement(self):
        from apps.export.columns import is_measurable_numeric
        self.assertFalse(is_measurable_numeric(pd.Series([True, False, True])))

    def test_str_and_object_are_not_measurements(self):
        from apps.export.columns import is_measurable_numeric
        self.assertFalse(is_measurable_numeric(pd.Series(['a', 'b'], dtype='str')))
        self.assertFalse(is_measurable_numeric(pd.Series(['a', 'b'], dtype=object)))

    def test_measurable_columns_helper_filters_dataframe(self):
        from apps.export.columns import measurable_numeric_columns
        df = pd.DataFrame({
            'F32': np.array([1.0, 2.0], dtype='float32'),
            'I32': np.array([1, 2], dtype='int32'),
            'Flag': [True, False],
            'Text': pd.Series(['a', 'b'], dtype='str'),
        })
        self.assertEqual(measurable_numeric_columns(df), ['F32', 'I32'])

    def test_int32_column_gets_stats_rows(self):
        """窄 dtype 测量列此前被白名单漏掉 → 统计区整列空白。"""
        df = pd.DataFrame({'SW_Bin': [1] * 4,
                           'I32': np.array([1, 2, 3, 4], dtype='int32')})
        ws = _stats_ws(df, dict(EMPTY_META))
        self.assertAlmostEqual(_cell(ws, 'Min'), 1.0, places=6)
        self.assertAlmostEqual(_cell(ws, 'Max'), 4.0, places=6)
        self.assertAlmostEqual(_cell(ws, 'Avg'), 2.5, places=6)

    def test_bool_column_never_gets_stats_rows(self):
        """bool 是 Pass/Fail 标记，不是测量值：即使 is_numeric_dtype 为真也排除。"""
        df = pd.DataFrame({'SW_Bin': [1] * 4, 'MyFlag': [True, False, True, True]})
        self.assertTrue(pd.api.types.is_numeric_dtype(df['MyFlag']))
        ws = _stats_ws(df, dict(EMPTY_META))
        for label in STATS_ROW:
            self.assertIn(_cell(ws, label), (None, ''),
                          msg=f'bool 列不应有 {label} 统计值')


class NoPpPpkInExportTests(SimpleTestCase):
    """pp/ppk 护栏：compute_cpk 的这两个字段将被删除，导出侧不得消费。

    说明：apps/export 历史上就从未消费 pp/ppk（grep 无命中），因此这组用例
    在修复前也是绿的——它们是防止后续（误）引入的护栏，不是红→绿用例。
    """

    def test_export_sources_do_not_consume_pp_ppk(self):
        root = Path(__file__).resolve().parents[2] / 'apps' / 'export'
        pattern = re.compile(r"['\"](?:pp|ppk|pp_level|ppk_level|pp_color|ppk_color)['\"]")
        offenders = []
        for path in sorted(root.glob('*.py')):
            if path.name == 'tests.py':
                continue    # 既有测试文件不在本轮改动范围
            for match in pattern.finditer(path.read_text(encoding='utf-8')):
                offenders.append(f'{path.name}:{match.group(0)}')
        self.assertEqual(offenders, [], 'apps/export 不得消费 compute_cpk 的 pp/ppk')

    def test_batch_charts_summary_headers_have_no_pp_ppk(self):
        from apps.export.export_batch_charts_xlsx import build_batch_charts_xlsx_with_charts
        df = pd.DataFrame({'P': [1.0, 2.0, 3.0, 4.0, 5.0]})
        metadata = {'units': {'P': 'V'}, 'mins': {'P': '0'}, 'maxs': {'P': '6'},
                    'format': 'CTA8290D'}
        buf = build_batch_charts_xlsx_with_charts(df, metadata, ['P'])
        ws = openpyxl.load_workbook(io.BytesIO(buf))['总览']
        headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
        for header in headers:
            self.assertNotIn('ppk', header.lower())
            self.assertNotEqual(header.lower(), 'pp')
