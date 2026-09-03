"""Numeric Gage R&R builder tests (defects 1-7).

These drive the *live* builder ``apps.gage.gage_legacy_builder.build_gage_summary_excel``
with synthetic DataFrames whose mean/variance are known by hand, then read the
generated .xlsx cells back with openpyxl and assert the reported statistics.

Every case is constructed so it FAILS against the pre-fix builder (verified by
running the identical datasets through ``git show HEAD:`` of the module):

* defect 1 - empty file polluted reproducibility via ``np.empty`` (NaN/garbage)
* defect 2 - single-value file -> ``UnboundLocalError`` / std leaked across items
* defect 3 - limits of exactly 0 and 4 treated as "missing" -> CP/CPK zeroed
* defect 4 - Min CPK / Max CPK both wrote overall_cpk
* defect 5 - R&R% hard 0 while Fail Level said Bad1; missing limit must be N/A
* defect 6 - V/W mixed 6σ with variance fractions; Average read cells back
* defect 7 - system / non-numeric columns were fed into the R&R

Pure-function tests -> SimpleTestCase (no DB).
"""
import io

import pandas as pd
from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.gage.gage_legacy_builder import build_gage_summary_excel


def _dataset(filename, cols, mins, maxs, units=None, fmt='CTA8290D'):
    """Build one file_datasets entry from {col: [values]} + string limit maps."""
    df = pd.DataFrame(cols)
    return {
        'filename': filename,
        'df': df,
        'metadata': {
            'format': fmt,
            'mins': mins,
            'maxs': maxs,
            'units': units or {},
            'tester_id': 'T-1',
            'program_name': 'P',
            'start_time': '2026-01-01',
        },
    }


def _summary_ws(datasets, ignore_no_limit=False):
    data = build_gage_summary_excel(datasets, ignore_no_limit)
    wb = load_workbook(io.BytesIO(data))
    return wb['Summary']


def _v(ws, ref):
    return ws[ref].value


def _f(ws, ref):
    val = ws[ref].value
    return None if val in (None, '') else float(val)


def _test_names(ws):
    names = set()
    for row in range(12, ws.max_row + 1):
        c = ws[f'C{row}'].value
        if c:
            names.add(c)
    return names


# ── Hand-computed expectations for the shared 3-file scenario ──
# T1 = [1,2,3] / [4,5,6] / [7,8,9], low=0 high=100 (tolerance=100), 3 files.
#   per-file std(ddof=0) = 0.8164966, means = 2 / 5 / 8
#   global mean = 5, global std(ddof=0) = 2.5819889
#   repeatability = 6*sqrt((3*0.8164966^2)/3) = 4.8989795
#   reproducibility = 6*std([2,5,8]) = 6*2.4494897 = 14.6969385
#   R&R = sqrt(4.8989795^2 + 14.6969385^2) = 15.4919334
#   R&R% = 15.4919334/100 = 0.15491933 -> 15.49% -> Bad2
#   overall_cp = 100/(6*2.5819889) = 6.4549722
#   overall_cpk = min((5-0)/(3*2.5819889), (100-5)/(3*2.5819889)) = 0.6454972
#   per-file cp = 100/(6*0.8164966) = 20.4124145
#   per-file cpk = 2/2.4494897, 5/2.4494897, 8/2.4494897 = 0.8164966 / 2.0412415 / 3.2659863
def _main_datasets():
    return [
        _dataset('A.csv', {'T1': [1.0, 2.0, 3.0]}, {'T1': '0'}, {'T1': '100'}),
        _dataset('B.csv', {'T1': [4.0, 5.0, 6.0]}, {'T1': '0'}, {'T1': '100'}),
        _dataset('C.csv', {'T1': [7.0, 8.0, 9.0]}, {'T1': '0'}, {'T1': '100'}),
    ]


class GageBuilderNumericTests(SimpleTestCase):
    """Read-back assertions on the generated Summary sheet."""

    DELTA = 1e-4

    # ---- baseline: every reported statistic matches the hand computation ----
    def test_main_known_values(self):
        ws = _summary_ws(_main_datasets())
        # per-file row A (row 12)
        self.assertEqual(_v(ws, 'C12'), 'T1')
        self.assertAlmostEqual(_f(ws, 'E12'), 0.0, delta=self.DELTA)      # LowLimit = 0 (legit)
        self.assertAlmostEqual(_f(ws, 'F12'), 100.0, delta=self.DELTA)    # HighLimit
        self.assertAlmostEqual(_f(ws, 'H12'), 2.0, delta=self.DELTA)      # file mean
        self.assertAlmostEqual(_f(ws, 'I12'), 0.8164966, delta=self.DELTA)  # file std
        self.assertAlmostEqual(_f(ws, 'L12'), 20.4124145, delta=self.DELTA)  # file CP
        self.assertAlmostEqual(_f(ws, 'M12'), 0.8164966, delta=self.DELTA)   # file CPK
        # per-file rows B / C CPK
        self.assertAlmostEqual(_f(ws, 'H13'), 5.0, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'M13'), 2.0412415, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'H14'), 8.0, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'M14'), 3.2659863, delta=self.DELTA)
        # group stats on first data row
        self.assertAlmostEqual(_f(ws, 'O12'), 5.0, delta=self.DELTA)         # global mean
        self.assertAlmostEqual(_f(ws, 'P12'), 2.5819889, delta=self.DELTA)   # global std
        self.assertAlmostEqual(_f(ws, 'Q12'), 15.4919334, delta=self.DELTA)  # 6*std
        self.assertAlmostEqual(_f(ws, 'T12'), 6.4549722, delta=self.DELTA)   # Total CP
        self.assertAlmostEqual(_f(ws, 'U12'), 0.6454972, delta=self.DELTA)   # Total CPK
        self.assertAlmostEqual(_f(ws, 'V12'), 4.8989795, delta=self.DELTA)   # EV (6σ repeatability)
        self.assertAlmostEqual(_f(ws, 'W12'), 14.6969385, delta=self.DELTA)  # AV (6σ reproducibility)
        self.assertAlmostEqual(_f(ws, 'X12'), 15.4919334, delta=self.DELTA)  # GRR
        self.assertAlmostEqual(_f(ws, 'Y12'), 0.15491933, delta=self.DELTA)  # R&R%
        self.assertEqual(_v(ws, 'Z12'), 'Bad2')

    # ---- defect 4: Min CPK / Max CPK must differ (min/max of per-file CPK) ----
    def test_defect4_min_max_cpk(self):
        ws = _summary_ws(_main_datasets())
        min_cpk = _f(ws, 'R12')
        max_cpk = _f(ws, 'S12')
        overall_cpk = _f(ws, 'U12')
        self.assertAlmostEqual(min_cpk, 0.8164966, delta=self.DELTA)  # file A
        self.assertAlmostEqual(max_cpk, 3.2659863, delta=self.DELTA)  # file C
        self.assertNotAlmostEqual(min_cpk, max_cpk, delta=self.DELTA)
        # pre-fix both equalled overall_cpk (0.6455)
        self.assertNotAlmostEqual(min_cpk, overall_cpk, delta=self.DELTA)
        self.assertNotAlmostEqual(max_cpk, overall_cpk, delta=self.DELTA)

    # ---- defect 6: V/W hold 6σ only; Average row computed from memory ----
    def test_defect6_vw_six_sigma_and_average(self):
        ws = _summary_ws(_main_datasets())
        self.assertAlmostEqual(_f(ws, 'V12'), 4.8989795, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'W12'), 14.6969385, delta=self.DELTA)
        # second file row must NOT carry variance fractions (pre-fix: 0.1 / 0.9)
        self.assertFalse(_v(ws, 'V13'))
        self.assertFalse(_v(ws, 'W13'))
        # Average row = 12 + num_tests(1)*num_files(3) = 15, from memory arrays
        self.assertEqual(_v(ws, 'A15'), 'Average')
        self.assertAlmostEqual(_f(ws, 'V15'), 4.8989795, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'W15'), 14.6969385, delta=self.DELTA)

    # ---- defect 1: an empty file must not pollute reproducibility ----
    def test_defect1_empty_file_reproducibility(self):
        empty = [
            _dataset('A.csv', {'T1': [1.0, 2.0, 3.0]}, {'T1': '0'}, {'T1': '100'}),
            _dataset('B.csv', {'T1': [4.0, 5.0, 6.0]}, {'T1': '0'}, {'T1': '100'}),
            _dataset('C.csv', {'T1': [None, None, None]}, {'T1': '0'}, {'T1': '100'}),
        ]
        two = [
            _dataset('A.csv', {'T1': [1.0, 2.0, 3.0]}, {'T1': '0'}, {'T1': '100'}),
            _dataset('B.csv', {'T1': [4.0, 5.0, 6.0]}, {'T1': '0'}, {'T1': '100'}),
        ]
        ws3 = _summary_ws(empty)
        ws2 = _summary_ws(two)
        # reproducibility over the 2 data-bearing files = 6*std([2,5]) = 6*1.5 = 9
        self.assertAlmostEqual(_f(ws3, 'W12'), 9.0, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws3, 'W12'), _f(ws2, 'W12'), delta=self.DELTA)
        # pre-fix produced NaN here (uninitialized np.empty slot)
        self.assertIsNotNone(_f(ws3, 'X12'))
        self.assertAlmostEqual(_f(ws3, 'X12'), (4.0 ** 2 + 9.0 ** 2) ** 0.5, delta=self.DELTA)

    # ---- defect 2: single-value file must not raise / must give std 0 ----
    def test_defect2_single_value_no_unbound(self):
        single = [
            _dataset('A.csv', {'Ts': [5.0]}, {'Ts': '0'}, {'Ts': '100'}),
            _dataset('B.csv', {'Ts': [7.0]}, {'Ts': '0'}, {'Ts': '100'}),
        ]
        # pre-fix: UnboundLocalError on 'fs'
        ws = _summary_ws(single)
        self.assertAlmostEqual(_f(ws, 'H12'), 5.0, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'I12'), 0.0, delta=self.DELTA)  # std of one value

    # ---- defect 2b: single-value std must not inherit previous test item ----
    def test_defect2b_no_cross_item_std_leak(self):
        cross = [
            _dataset('A.csv', {'T1': [1.0, 2.0, 3.0], 'T2': [99.0, None, None]},
                     {'T1': '0', 'T2': '0'}, {'T1': '100', 'T2': '200'}),
            _dataset('B.csv', {'T1': [4.0, 5.0, 6.0], 'T2': [88.0, None, None]},
                     {'T1': '0', 'T2': '0'}, {'T1': '100', 'T2': '200'}),
        ]
        ws = _summary_ws(cross)
        # T1 group rows 12-13, T2 group rows 14-15
        self.assertEqual(_v(ws, 'C12'), 'T1')
        self.assertAlmostEqual(_f(ws, 'I12'), 0.8164966, delta=self.DELTA)  # T1 file A std
        self.assertEqual(_v(ws, 'C14'), 'T2')
        # pre-fix: T2 single-value std leaked T1's 0.8164966 instead of 0
        self.assertAlmostEqual(_f(ws, 'I14'), 0.0, delta=self.DELTA)

    # ---- defect 3: limits exactly 0 and 4 are valid, not "missing" ----
    def test_defect3_limits_zero_and_four(self):
        ds = [
            _dataset('A.csv', {'Tv': [1.9, 2.0, 2.1]}, {'Tv': '0'}, {'Tv': '4'}),
            _dataset('B.csv', {'Tv': [1.8, 2.0, 2.2]}, {'Tv': '0'}, {'Tv': '4'}),
        ]
        ws = _summary_ws(ds)
        # file A: mean 2, std(ddof0)=0.0816497, tol=4
        #   cp = 4/(6*0.0816497) = 8.164966 ; cpk = min(2/(3*std), 2/(3*std)) = 8.164966
        self.assertAlmostEqual(_f(ws, 'E12'), 0.0, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'F12'), 4.0, delta=self.DELTA)
        self.assertAlmostEqual(_f(ws, 'L12'), 8.164966, delta=self.DELTA)  # pre-fix: 0
        self.assertAlmostEqual(_f(ws, 'M12'), 8.164966, delta=self.DELTA)  # pre-fix: 0

    # ---- defect 5: missing tolerance -> R&R% and Fail Level both N/A, no Bad1 ----
    def test_defect5_missing_limit_is_na(self):
        ds = [
            _dataset('A.csv', {'Tm': [1.0, 2.0, 3.0]}, {}, {}),
            _dataset('B.csv', {'Tm': [4.0, 5.0, 6.0]}, {}, {}),
        ]
        ws = _summary_ws(ds)
        self.assertEqual(_v(ws, 'E12'), 'N/A')
        self.assertEqual(_v(ws, 'F12'), 'N/A')
        self.assertEqual(_v(ws, 'Y12'), 'N/A')   # R&R%
        self.assertEqual(_v(ws, 'Z12'), 'N/A')   # Fail Level - same source
        # pre-fix: Bad1 with a contradictory numeric R&R%, bad1_count = 1
        self.assertEqual(_v(ws, 'B3'), 0)

    # ---- defect 5b: a genuine >=30% R&R still counts as Bad1 (display==judgement) ----
    def test_defect5b_real_bad1_still_counted(self):
        # tight tolerance forces R&R% >= 30%
        ds = [
            _dataset('A.csv', {'Tb': [1.0, 2.0, 3.0]}, {'Tb': '0'}, {'Tb': '10'}),
            _dataset('B.csv', {'Tb': [4.0, 5.0, 6.0]}, {'Tb': '0'}, {'Tb': '10'}),
        ]
        ws = _summary_ws(ds)
        # R&R = sqrt(4.899^2 + 9^2)=10.247 ; tol=10 -> 102.47% -> Bad1
        self.assertEqual(_v(ws, 'Z12'), 'Bad1')
        self.assertAlmostEqual(_f(ws, 'Y12'), 10.247 / 10, delta=1e-3)
        self.assertEqual(_v(ws, 'B3'), 1)

    # ---- defect 7: system + non-numeric columns excluded regardless of flag ----
    def test_defect7_system_and_nonnumeric_excluded(self):
        cols = {
            'Serial_No': [1, 2, 3],
            'SW_Bin': [1, 1, 1],
            'QR_Code': ['a', 'b', 'c'],
            'Start_T': ['t1', 't2', 't3'],
            'T1': [1.0, 2.0, 3.0],
        }
        mins = {'Serial_No': '', 'SW_Bin': '', 'QR_Code': '', 'Start_T': '', 'T1': '0'}
        maxs = {'Serial_No': '', 'SW_Bin': '', 'QR_Code': '', 'Start_T': '', 'T1': '100'}
        ds = [_dataset('A.csv', cols, mins, maxs), _dataset('B.csv', cols, mins, maxs)]
        for flag in (False, True):
            with self.subTest(ignore_no_limit=flag):
                ws = _summary_ws(ds, ignore_no_limit=flag)
                names = _test_names(ws)
                self.assertIn('T1', names)
                for syscol in ('Serial_No', 'SW_Bin', 'QR_Code', 'Start_T'):
                    self.assertNotIn(syscol, names)

    # ---- defect 6/7 with num_files > 2: Average row still correct ----
    def test_num_files_greater_than_two_average(self):
        # 4 files, 2 test items -> Average row at 12 + 2*4 = 20
        ds = [
            _dataset('A.csv', {'T1': [1.0, 2.0, 3.0], 'T2': [10.0, 11.0, 12.0]},
                     {'T1': '0', 'T2': '0'}, {'T1': '100', 'T2': '100'}),
            _dataset('B.csv', {'T1': [4.0, 5.0, 6.0], 'T2': [13.0, 14.0, 15.0]},
                     {'T1': '0', 'T2': '0'}, {'T1': '100', 'T2': '100'}),
            _dataset('C.csv', {'T1': [7.0, 8.0, 9.0], 'T2': [16.0, 17.0, 18.0]},
                     {'T1': '0', 'T2': '0'}, {'T1': '100', 'T2': '100'}),
            _dataset('D.csv', {'T1': [10.0, 11.0, 12.0], 'T2': [19.0, 20.0, 21.0]},
                     {'T1': '0', 'T2': '0'}, {'T1': '100', 'T2': '100'}),
        ]
        ws = _summary_ws(ds)
        self.assertEqual(_v(ws, 'A20'), 'Average')
        # both tests have identical per-file std 0.8164966 -> repeatability same
        # T1 means [2,5,8,11] std=3.3541020 -> AV=20.1246118
        # T2 means [11,14,17,20] std=3.3541020 -> AV=20.1246118 ; average equal
        self.assertIsNotNone(_f(ws, 'V20'))
        self.assertAlmostEqual(_f(ws, 'W20'), 20.1246118, delta=1e-3)
