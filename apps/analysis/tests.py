"""Regression tests for the histogram service.

Focus: ``site_histograms`` is populated for every file that has a Site
column, including the single-site case. Previously the backend returned
``site_histograms: None`` for files with only one unique site, so the
front-end histogram had to label the lone bar as "数据分布" instead of
"Site1" — confusing the user. The 2026-06-07 fix removes the
``> 1`` guard on ``site_idx.unique()``; these tests lock in the
behaviour so a future refactor cannot silently regress it.
"""
import numpy as np
import pandas as pd
import types
import math
from django.test import SimpleTestCase, TestCase

from apps.analysis.services.data_services import compute_histogram_stats
from apps.analysis.services.statistics import compute_boxplot_stats


def _make_df(n: int, sites: list[int], seed: int = 0) -> pd.DataFrame:
    """Build a tiny DataFrame with a Site column and a numeric param."""
    rng = np.random.default_rng(seed)
    # Repeat the site list to get `n` rows, then take the first n.
    sites_full = (sites * (n // len(sites) + 1))[:n]
    return pd.DataFrame({
        'Site': [str(s) for s in sites_full],
        'Param1': rng.normal(0, 1, n).tolist(),
    })


def _meta(format_type: str = 'CTA8290D') -> dict:
    return {'format': format_type, 'unit': 'mV'}


class SiteHistogramsTests(SimpleTestCase):
    """``site_histograms`` must be present whenever a Site column exists."""

    def test_multi_site_populates_per_site_keys(self):
        df = _make_df(200, [1, 2, 3, 4])
        out = compute_histogram_stats(df, _meta(), 'Param1', 'Site')
        self.assertIsInstance(out['site_histograms'], dict)
        self.assertEqual(
            set(out['site_histograms'].keys()), {'1', '2', '3', '4'}
        )
        # Each site's bin list has the same length as bin_centers.
        bin_count = len(out['bin_centers'])
        for site, hist in out['site_histograms'].items():
            self.assertEqual(
                len(hist), bin_count, f'site {site} bin count mismatch'
            )

    def test_single_site_populates_one_key(self):
        # Regression: previously returned site_histograms=None and the
        # front-end mis-labelled the only site as "数据分布".
        df = _make_df(200, [1])
        out = compute_histogram_stats(df, _meta(), 'Param1', 'Site')
        self.assertIsInstance(out['site_histograms'], dict)
        self.assertEqual(set(out['site_histograms'].keys()), {'1'})
        self.assertEqual(len(out['site_histograms']['1']), len(out['bin_centers']))

    def test_no_site_column_keeps_none(self):
        # When the file has no Site column at all, site_histograms stays
        # None and the front-end falls back to the "数据分布" label — this
        # is the only legitimate "no per-site data" case.
        df = _make_df(200, [1])  # has a Site column
        df = df.drop(columns=['Site'])
        out = compute_histogram_stats(df, _meta(), 'Param1', site_col=None)
        self.assertIsNone(out['site_histograms'])

    def test_bin_grid_symmetric(self):
        """回归 2026-08-14：bin 网格曾锚定 bin_min（LSL），USL 右侧多出 4 个
        细分 bin（LSL 左 1 中心 vs USL 右 5 中心，X 轴视觉偏左）。修复后网格
        以 [bin_min, bin_max] 为中心对称：两侧各 2 细分 bin + catch-all，
        规格限内保持 20 个 bin。"""
        df = _make_df(200, [1, 2, 3, 4])
        meta = {**_meta(), 'mins': {'Param1': '-3'}, 'maxs': {'Param1': '3'}}
        out = compute_histogram_stats(df, meta, 'Param1', 'Site')
        bc = out['bin_centers']
        lo, hi = out['lower_limit'], out['upper_limit']
        left = [c for c in bc if c < lo]
        right = [c for c in bc if c > hi]
        self.assertEqual(
            len(left), len(right),
            f'bin 网格应两侧对称：LSL 左 {len(left)} 个 vs USL 右 {len(right)} 个',
        )
        self.assertEqual(len(left), 3)  # underflow + 2 细分
        inner = [c for c in bc if lo <= c <= hi]
        self.assertEqual(len(inner), 20)  # 规格限内 20 个 bin（gap = range/20）

    def test_single_site_percentages_sum_to_about_100(self):
        # Same denominator (total_count = all sites) as multi-site case.
        df = _make_df(300, [1])
        out = compute_histogram_stats(df, _meta(), 'Param1', 'Site')
        site_sum = sum(out['site_histograms']['1'])
        # Allow a small tolerance for overflow/underflow bins and rounding.
        self.assertGreater(site_sum, 95.0)
        self.assertLess(site_sum, 100.05)


class EmptyParamFilterTests(SimpleTestCase):
    """Regression: empty-string column names from trailing-comma CSV headers
    must not appear in the histogram fast-path response.

    The CTA8280F parser historically produced an unnamed column ``""`` from
    a trailing comma in the header. That column was all-NaN, passed the
    ``dtype in ('int64','float64')`` numeric_cols check, and ended up in
    the params list. Selecting it 400'd the QQ plot and other endpoints
    (``if param not in df.columns``). The 2026-06-07 fix filters blanks
    out of the response so the param selector never offers the phantom.
    """

    def test_histogram_fast_path_filters_empty_column(self):
        from apps.analysis.views import _filter_blank_params

        # Simulate the params list as it would be built right before the
        # ``{p: {} for p in params}`` dict comprehension in
        # ``AnalysisViewSet.histogram``.
        raw = ['Index_No', 'Dut_No', '', 'CON_VIN', '   ']
        filtered = _filter_blank_params(raw)
        self.assertNotIn('', filtered, 'empty-string column name must be filtered out')
        self.assertNotIn('   ', filtered, 'whitespace-only column name must be filtered out')
        self.assertEqual(set(filtered), {'Index_No', 'Dut_No', 'CON_VIN'})

    def test_histogram_fast_path_keeps_real_params(self):
        from apps.analysis.views import _filter_blank_params

        # When there are no blanks, the filter must be a no-op.
        raw = ['CON_VIN', 'CON_VCC', 'BV_HS_100uA_1']
        self.assertEqual(_filter_blank_params(raw), raw)


class QqPlotParamGuardTests(SimpleTestCase):
    """``compute_qqplot`` and the qqplot view must reject empty / whitespace
    params gracefully — never 500 — because the user-facing param dropdown
    filters blanks but the endpoint is still callable directly.
    """

    def test_compute_qqplot_handles_empty_series(self):
        import numpy as np
        from apps.analysis.services.statistics.computations import compute_qqplot

        # Empty / all-NaN series — should return zeros, not raise.
        empty = pd.Series([], dtype=float)
        result = compute_qqplot(empty)
        self.assertEqual(result['n'], 0)
        self.assertFalse(result['is_normal'])
        self.assertEqual(result['theoretical_quantiles'], [])

        all_nan = pd.Series([np.nan, np.nan, np.nan])
        result = compute_qqplot(all_nan)
        self.assertEqual(result['n'], 0)
        self.assertFalse(result['is_normal'])


class MultiFileAnalysisTests(SimpleTestCase):
    """Multi-file (多文件分析) tab service logic.

    Covers the common-test-item intersection (by exact column name), the
    ``ignore_no_limit`` filter (keep only params with valid limits in every
    file), and the per-file limit attached to each lot in the distribution.
    """

    @staticmethod
    def _file(cols: dict, mins: dict, maxs: dict, fid: int = 1, name: str = 'f'):
        df = pd.DataFrame(cols)
        meta = {'format': 'CTA8290D', 'mins': mins, 'maxs': maxs, 'units': {}}
        return (fid, df, meta, name)

    def test_common_params_intersects_by_name(self):
        from apps.analysis.services.data_services import compute_common_params

        f1 = self._file(
            {'A': [1.0, 2.0], 'B': [1.0, 2.0], 'Site': ['1', '2']},
            {'A': '0', 'B': '0'}, {'A': '5', 'B': '5'}, fid=1)
        f2 = self._file(
            {'A': [1.0, 2.0], 'C': [1.0, 2.0], 'Site': ['1', '2']},
            {'A': '0', 'C': '0'}, {'A': '5', 'C': '5'}, fid=2)
        # 'Site' is object dtype → excluded; only 'A' is numeric in both.
        self.assertEqual(compute_common_params([f1, f2]), ['A'])

    def test_common_params_ignore_no_limit_keeps_only_limited(self):
        from apps.analysis.services.data_services import compute_common_params

        # 'A' has limits in both files; 'B' has limits only in f1.
        f1 = self._file(
            {'A': [1.0, 2.0], 'B': [1.0, 2.0]},
            {'A': '0', 'B': '0'}, {'A': '5', 'B': '5'}, fid=1)
        f2 = self._file(
            {'A': [1.0, 2.0], 'B': [1.0, 2.0]},
            {'A': '0'}, {'A': '5'}, fid=2)
        self.assertEqual(compute_common_params([f1, f2]), ['A', 'B'])
        # With ignore_no_limit, 'B' drops because f2 has no limit for it.
        self.assertEqual(
            compute_common_params([f1, f2], ignore_no_limit=True), ['A'])

    def test_distribution_attaches_file_id_and_limits(self):
        from apps.analysis.services.data_services import (
            compute_multi_lot_distribution,
        )

        s1 = pd.Series([1.0, 2.0, 3.0])
        s2 = pd.Series([2.0, 3.0, 4.0])
        df1 = pd.DataFrame({'A': s1})
        df2 = pd.DataFrame({'A': s2})
        datasets = {
            '1': {'df': df1, 'metadata': {'mins': {'A': '0'}, 'maxs': {'A': '5'}},
                  'series': s1, 'name': 'f1', 'file_id': 1},
            '2': {'df': df2, 'metadata': {'mins': {'A': 'MIN'}, 'maxs': {'A': 'MAX'}},
                  'series': s2, 'name': 'f2', 'file_id': 2},
        }
        out = compute_multi_lot_distribution(datasets, [s1, s2], 'A')
        lots = {lot['file_id']: lot for lot in out['lot_data']}
        # File 1 has numeric limits → drawn.
        self.assertEqual(lots[1]['lower_limit'], 0.0)
        self.assertEqual(lots[1]['upper_limit'], 5.0)
        # File 2's "MIN"/"MAX" markers are not valid numeric limits → omitted.
        self.assertIsNone(lots[2]['lower_limit'])
        self.assertIsNone(lots[2]['upper_limit'])
        # eedeceb 回归修复：25 inner edges → 26 bins（1 underflow + 24 normal +
        # 1 overflow），与单文件直方图同构；重构时误写 range(26) 曾产生 27 个
        # bin，multi-file.spec.ts「X 轴固定 24 个坐标」据此失配
        self.assertEqual(len(out['bin_centers']), 26)

    def test_range_type_not_expanded_to_spec_limits(self):
        """回归 2026-08-13：multi_lot 曾无条件把 bin 范围扩展到全局规格限
        （global_lsl/global_usl），带规格限窄分布参数下 5 种 range_type 的
        X 轴完全相同——与单文件 histogram 的 resolve_limits 语义不一致，
        切换范围类型看起来不生效。修复后各 range_type 范围应各不相同。"""
        import numpy as np
        from apps.analysis.services.data_services import (
            compute_multi_lot_distribution,
        )

        rng = np.random.default_rng(42)
        # 窄分布、规格限 0~2 包住数据（σ≈0.05 → ±6σ 仍在规格限内）
        s1 = pd.Series(rng.normal(1.0, 0.05, 200))
        s2 = pd.Series(rng.normal(1.02, 0.05, 200))
        df1 = pd.DataFrame({'A': s1})
        df2 = pd.DataFrame({'A': s2})
        datasets = {
            '1': {'df': df1, 'metadata': {'mins': {'A': '0'}, 'maxs': {'A': '2'}},
                  'series': s1, 'name': 'f1', 'file_id': 1},
            '2': {'df': df2, 'metadata': {'mins': {'A': '0'}, 'maxs': {'A': '2'}},
                  'series': s2, 'name': 'f2', 'file_id': 2},
        }
        spans = {}
        for rt in ('RDL', 'DR', 'S3', 'S4', 'S6'):
            out = compute_multi_lot_distribution(
                datasets, [s1, s2], 'A', range_type=rt)
            bc = out['bin_centers']
            spans[rt] = max(bc) - min(bc)
        # 修复前：全部被规格限扩展吞成同一范围 → 三个断言必失败
        self.assertNotEqual(spans['S3'], spans['S4'])
        self.assertNotEqual(spans['S4'], spans['S6'])
        self.assertNotEqual(spans['S3'], spans['RDL'])
        # S3 是窄分布（±3σ ≈ 0.3 跨度），RDL 是规格限（0~2）——S3 明显更窄
        self.assertLess(spans['S3'], spans['RDL'])
        # DR 是纯数据范围（窄），不应被扩展到规格限
        self.assertLess(spans['DR'], 0.5)


class BoxPlotStatsDtypeToleranceTests(SimpleTestCase):
    """Lock down ``compute_boxplot_stats`` behaviour on non-float dtypes.

    Originally the function did ``data.dropna()`` then
    ``clean_data.apply(lambda x: abs(x) < float('inf'))`` then ``.quantile()``.
    For boolean Series (e.g. ``Dut_Pass``/``PASSFG``):

    * ``pd.to_numeric`` on bool returns the **same** bool Series (no dtype
      change), so ``.quantile(0.25)`` returns ``np.bool_`` and the
      ``q3 - q1`` step raises
      "numpy boolean subtract, the `-` operator, is not supported".
    * For object/string Series, ``abs('foo')`` raises
      "bad operand type for abs(): 'str'".

    The 2026-06-13 fix coerces via ``pd.to_numeric(..., errors='coerce')
    .astype(float)`` at the entry, so these dtypes no longer crash.
    """

    def test_boolean_series_does_not_crash(self):
        s = pd.Series([True, False, True, False, True, True, False] * 10)
        out = compute_boxplot_stats(s)
        # All booleans parse as 0.0/1.0; min=0, max=1, count=70.
        self.assertEqual(out['count'], 70)
        self.assertEqual(out['min'], 0.0)
        self.assertEqual(out['max'], 1.0)
        self.assertEqual(out['q1'], 0.0)
        self.assertEqual(out['q3'], 1.0)

    def test_string_series_does_not_crash(self):
        # Mix of parseable ("1.5") and unparseable ("foo") strings.
        s = pd.Series(['1.5', '2.5', 'foo', '3.5', 'bar', '4.5'] * 10)
        out = compute_boxplot_stats(s)
        # Only the 4 parseable values survive; the rest become NaN and are
        # dropped.
        self.assertEqual(out['count'], 40)
        self.assertEqual(out['min'], 1.5)
        self.assertEqual(out['max'], 4.5)

    def test_pure_string_series_returns_zero_count(self):
        s = pd.Series(['foo', 'bar', 'baz'] * 5)
        out = compute_boxplot_stats(s)
        # No values survive → count=0, defaults returned.
        self.assertEqual(out['count'], 0)
        self.assertEqual(out['min'], 0.0)
        self.assertEqual(out['max'], 0.0)
        self.assertEqual(out['outliers'], [])

    def test_pure_boolean_constant_series_returns_valid_stats(self):
        # Regression: SW_Bin-style "all-same boolean" (pass=1 across the
        # whole file). Pre-fix this raised inside .quantile().
        s = pd.Series([True] * 100)
        out = compute_boxplot_stats(s)
        self.assertEqual(out['count'], 100)
        self.assertEqual(out['min'], 1.0)
        self.assertEqual(out['max'], 1.0)
        self.assertEqual(out['median'], 1.0)
        # iqr == 0 → no outliers
        self.assertEqual(out['outliers'], [])

    def test_numeric_series_still_works(self):
        # Sanity check: the fix must not regress the normal path.
        rng = np.random.default_rng(42)
        s = pd.Series(rng.normal(0, 1, 200))
        out = compute_boxplot_stats(s)
        self.assertEqual(out['count'], 200)
        self.assertLess(out['q1'], out['median'])
        self.assertLess(out['median'], out['q3'])
        self.assertLessEqual(out['min'], out['q1'])
        self.assertLessEqual(out['q3'], out['max'])


class StaleParamAcrossFileSwitchTests(SimpleTestCase):
    """Lock down the "stale selectedParam after file switch" guard.

    Regression: when the user selects ``R_Kelvin_AGND`` on
    ``gage_m_S4.csv`` (file_id=14518) and then switches to
    ``BPD93204_FT1_ETS163550_12252024.csv`` (file_id=14514, an ETS88 file
    with no such column), the persisted Pinia store value carried over
    and the analysis APIs were called with a column that does not exist
    in the new file. ``qqplot`` and ``boxplot`` returned 400, but
    ``histogram`` had no ``param in df.columns`` guard and crashed with
    ``KeyError: 'R_Kelvin_AGND'`` → 500.

    The 2026-06-13 fix:

    1. The frontend ``AnalysisPage.onFileChange`` resets
       ``selectedParam`` and the persisted store value at the start of
       every file change. This is the primary fix.
    2. As defence in depth, the backend ``histogram`` and ``boxplot``
       views now validate that every requested param exists in the
       current DataFrame and return 400 ``no_valid_params`` with a
       structured ``requested``/``missing`` payload instead of 500.

    These tests pin the backend guard so a future refactor that
    re-introduces the ``df[param]`` KeyError fails CI loudly.
    """

    @staticmethod
    def _patched_view(df, datafile=None, metadata=None):
        """Return ``(APIRequestFactory, force_authenticate, restore)`` after
        monkey-patching ``_load_df_from_request`` on the concrete view modules
        so it returns ``df`` without touching the database.

        We don't need a real DataFile / parsed file: the histogram / qqplot
        views branch on df.columns long before the heavy stats run. This
        keeps the test fast and DB-independent.
        """
        import types
        from rest_framework.test import APIRequestFactory, force_authenticate
        from apps.analysis.views import analysis_views, statistics_views

        if datafile is None:
            datafile = types.SimpleNamespace(
                id=1, filename='fake.csv', format_type='CTA8290D',
            )
        if metadata is None:
            metadata = {'format': 'CTA8290D', 'mins': {}, 'maxs': {}, 'units': {}}

        def fake_load(request):
            return df, datafile, metadata, None

        original_analysis = getattr(analysis_views, '_load_df_from_request', None)
        original_statistics = getattr(statistics_views, '_load_df_from_request', None)

        analysis_views._load_df_from_request = fake_load
        statistics_views._load_df_from_request = fake_load

        def restore():
            if original_analysis is not None:
                analysis_views._load_df_from_request = original_analysis
            if original_statistics is not None:
                statistics_views._load_df_from_request = original_statistics

        return APIRequestFactory(), force_authenticate, restore

    @staticmethod
    def _authed_post(factory, force_authenticate, url, payload):
        """POST a JSON payload with an authenticated user."""
        request = factory.post(url, payload, format='json')
        # ``force_authenticate`` short-circuits DRF's auth classes so the
        # ``IsAuthenticated`` permission on the view passes.
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))
        return request

    @staticmethod
    def _authed_get(factory, force_authenticate, url, params):
        """GET with query params and an authenticated user."""
        request = factory.get(url, params)
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))
        return request

    def test_histogram_view_returns_400_for_unknown_param(self):
        """POST /analysis/histogram/ with a param not in the file → 400."""
        from apps.analysis.views import AnalysisViewSet

        df = pd.DataFrame({'Param0': [1.0, 2.0, 3.0], 'Param1': [4.0, 5.0, 6.0]})
        factory, force_authenticate, restore = self._patched_view(df)
        self.addCleanup(restore)

        request = self._authed_post(factory, force_authenticate,
                                    '/api/v1/analysis/histogram/', {
            'file_id': 1,
            'params': ['__bogus__'],
        })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 400, response.content)
        body = response.data
        self.assertEqual(body.get('error'), 'no_valid_params')
        self.assertEqual(body.get('missing'), ['__bogus__'])
        self.assertEqual(body.get('requested'), ['__bogus__'])

    def test_histogram_view_drops_partial_unknown_params(self):
        """When some params exist and others don't, the existing ones still compute.

        Guards against an over-eager fix that would 400 the whole request
        just because one of the params is stale.
        """
        from apps.analysis.views import AnalysisViewSet

        df = pd.DataFrame({'Param0': [1.0, 2.0, 3.0], 'Param1': [4.0, 5.0, 6.0]})
        factory, force_authenticate, restore = self._patched_view(df)
        self.addCleanup(restore)

        request = self._authed_post(factory, force_authenticate,
                                    '/api/v1/analysis/histogram/', {
            'file_id': 1,
            'params': ['Param0', '__bogus__'],
        })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        # The real param must compute; the bogus one is dropped.
        self.assertEqual(response.status_code, 200, response.content)
        body = response.data
        results = body.get('results', {})
        self.assertIn('Param0', results)
        self.assertNotIn('__bogus__', results)

    def test_qqplot_view_returns_param_not_found_for_unknown_param(self):
        """POST /analysis/qqplot/ with a param not in the file → 400 param_not_found."""
        from apps.analysis.views import AnalysisViewSet

        df = pd.DataFrame({'Param0': [1.0, 2.0, 3.0]})
        factory, force_authenticate, restore = self._patched_view(df)
        self.addCleanup(restore)

        request = self._authed_post(factory, force_authenticate,
                                    '/api/v1/analysis/qqplot/', {
            'file_id': 1,
            'param': '__bogus__',
        })
        view = AnalysisViewSet.as_view({'post': 'qqplot'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 400, response.content)
        self.assertEqual(response.data.get('error'), 'param_not_found')

    def test_boxplot_view_returns_400_for_unknown_param(self):
        """GET /statistics/boxplot/ with a param not in the file → 400 no_valid_params."""
        from apps.analysis.views import StatisticsViewSet

        df = pd.DataFrame({'Param0': [1.0, 2.0, 3.0]})
        factory, force_authenticate, restore = self._patched_view(df)
        self.addCleanup(restore)

        request = self._authed_get(factory, force_authenticate,
                                   '/api/v1/statistics/boxplot/', {
            'file_id': 1,
            'params': ['__bogus__'],
            'group_by': 'site',
        })
        view = StatisticsViewSet.as_view({'get': 'boxplot'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 400, response.content)
        body = response.data
        self.assertEqual(body.get('error'), 'no_valid_params')
        self.assertEqual(body.get('missing'), ['__bogus__'])


class CustomLimitCpkApiTests(SimpleTestCase):
    """``POST /analysis/histogram/`` returns ``custom_cpk`` in CL mode.

    CustomLimit acts as a user-supplied spec-limit override: with
    ``range_type='CL'`` and both bounds provided, the response carries an
    extra ``custom_cpk`` (computed against the custom bounds) alongside the
    RDL-anchored ``cpk`` so the front-end can show before/after values.
    Outside CL the three fields stay ``null``.
    """

    @staticmethod
    def _patched_view(df, metadata):
        """Monkey-patch ``_load_df_from_request`` so no DB/DataFile is needed."""
        import types
        from rest_framework.test import APIRequestFactory, force_authenticate
        from apps.analysis.views import analysis_views

        datafile = types.SimpleNamespace(
            id=1, filename='fake.csv', format_type=metadata.get('format'),
        )

        def fake_load(request):
            return df, datafile, metadata, None

        original = getattr(analysis_views, '_load_df_from_request', None)
        analysis_views._load_df_from_request = fake_load

        def restore():
            if original is not None:
                analysis_views._load_df_from_request = original

        return APIRequestFactory(), force_authenticate, restore

    @staticmethod
    def _authed_post(factory, force_authenticate, url, payload):
        import types
        request = factory.post(url, payload, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))
        return request

    @classmethod
    def _df_meta(cls):
        df = pd.DataFrame({'Param0': [10.4480, 10.4495, 10.4510, 10.4519,
                                      10.4530, 10.4545, 10.4553]})
        metadata = {'format': 'CTA8280F', 'mins': {'Param0': '8.0'},
                    'maxs': {'Param0': '14.3'}, 'units': {'Param0': 'uA'}}
        return df, metadata

    def test_cl_mode_returns_custom_cpk(self):
        """CL with custom bounds → custom_cpk present, differs from RDL cpk."""
        from apps.analysis.views import AnalysisViewSet

        df, metadata = self._df_meta()
        factory, force_authenticate, restore = self._patched_view(df, metadata)
        self.addCleanup(restore)

        request = self._authed_post(factory, force_authenticate,
                                    '/api/v1/analysis/histogram/', {
            'file_id': 1, 'params': ['Param0'],
            'range_type': 'CL', 'custom_low': 10.40, 'custom_high': 10.50,
        })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        r = response.data['results']['Param0']
        self.assertIsNotNone(r['custom_cpk'], 'custom_cpk must be computed in CL mode')
        self.assertNotEqual(r['custom_cpk'], r['cpk'])
        self.assertLess(r['custom_cpk'], r['cpk'])
        self.assertIsNotNone(r['custom_cpk_level'])
        self.assertIn(r['custom_cpk_color'],
                      ('green', 'orange', 'darkorange', 'red', 'gray'))
        # Custom bounds echoed back so the chart can draw the CL markLines.
        self.assertEqual(r['custom_low'], 10.40)
        self.assertEqual(r['custom_high'], 10.50)

    def test_non_cl_mode_returns_null_custom_cpk(self):
        """Outside CL the custom_cpk fields stay null."""
        from apps.analysis.views import AnalysisViewSet

        df, metadata = self._df_meta()
        factory, force_authenticate, restore = self._patched_view(df, metadata)
        self.addCleanup(restore)

        request = self._authed_post(factory, force_authenticate,
                                    '/api/v1/analysis/histogram/', {
            'file_id': 1, 'params': ['Param0'],
            'range_type': 'RDL',
        })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        r = response.data['results']['Param0']
        self.assertIsNone(r['custom_cpk'])
        self.assertIsNone(r['custom_cpk_level'])
        self.assertIsNone(r['custom_cpk_color'])
        self.assertIsNone(r['custom_low'])
        self.assertIsNone(r['custom_high'])


def _num_peaks(y):
    """Count strict local maxima (y[i] greater than both neighbours)."""
    return sum(1 for i in range(1, len(y) - 1) if y[i] > y[i - 1] and y[i] > y[i + 1])


class KdeCurveApiTests(SimpleTestCase):
    """POST /analysis/histogram/ must return the non-parametric KDE curve.

    The normal overlay cannot represent bimodal data; ``kde_curve`` (200
    sampled points) lets the front-end draw a density curve that follows the
    actual shape.  Pin both the presence and the bimodal shape semantics.
    """

    @classmethod
    def _df_meta(cls, vals):
        df = pd.DataFrame({'Param0': vals})
        metadata = {'format': 'CTA8280F', 'mins': {'Param0': '8.0'},
                    'maxs': {'Param0': '14.3'}, 'units': {'Param0': 'uA'}}
        return df, metadata

    def test_histogram_api_returns_kde_curve(self):
        from apps.analysis.views import AnalysisViewSet

        rng = np.random.default_rng(42)
        df, metadata = self._df_meta(rng.normal(11.0, 0.8, 500).tolist())
        factory, force_authenticate, restore = StaleParamAcrossFileSwitchTests._patched_view(df, metadata=metadata)
        self.addCleanup(restore)

        request = StaleParamAcrossFileSwitchTests._authed_post(
            factory, force_authenticate, '/api/v1/analysis/histogram/', {
                'file_id': 1, 'params': ['Param0'], 'range_type': 'RDL',
            })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        r = response.data['results']['Param0']
        curve = r['kde_curve']
        self.assertIsNotNone(curve)
        self.assertEqual(len(curve), 200)
        self.assertGreater(max(p[1] for p in curve), 0)

    def test_histogram_api_kde_curve_shows_two_peaks_for_bimodal(self):
        from apps.analysis.views import AnalysisViewSet

        rng = np.random.default_rng(42)
        vals = np.concatenate([
            rng.normal(9.5, 0.35, 300),
            rng.normal(12.0, 0.35, 300),
        ]).tolist()
        df, metadata = self._df_meta(vals)
        factory, force_authenticate, restore = StaleParamAcrossFileSwitchTests._patched_view(df, metadata=metadata)
        self.addCleanup(restore)

        request = StaleParamAcrossFileSwitchTests._authed_post(
            factory, force_authenticate, '/api/v1/analysis/histogram/', {
                'file_id': 1, 'params': ['Param0'], 'range_type': 'RDL',
            })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        curve = response.data['results']['Param0']['kde_curve']
        self.assertIsNotNone(curve)
        self.assertEqual(_num_peaks([p[1] for p in curve]), 2)

    def test_histogram_api_kde_curves_split_by_outlier_inclusion(self):
        """「KDE含超限」数据源对：fail 数据 → kde_curve 全量 + filtered_kde_curve
        剔除；无异常值 → filtered_kde_curve 为 None。"""
        from apps.analysis.views import AnalysisViewSet

        rng = np.random.default_rng(42)
        vals = np.concatenate([
            rng.normal(11.0, 0.8, 500),
            [30.0] * 20,  # out-of-spec fail values
        ]).tolist()
        df, metadata = self._df_meta(vals)
        factory, force_authenticate, restore = StaleParamAcrossFileSwitchTests._patched_view(df, metadata=metadata)
        self.addCleanup(restore)

        request = StaleParamAcrossFileSwitchTests._authed_post(
            factory, force_authenticate, '/api/v1/analysis/histogram/', {
                'file_id': 1, 'params': ['Param0'], 'range_type': 'DR',
            })
        view = AnalysisViewSet.as_view({'post': 'histogram'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        r = response.data['results']['Param0']
        full = r['kde_curve']
        filtered = r['filtered_kde_curve']
        self.assertIsNotNone(full)
        self.assertIsNotNone(filtered, 'fail data must yield a filtered curve')
        tail_y = [y for x, y in full if 29.0 <= x <= 31.0]
        self.assertTrue(tail_y)
        self.assertGreater(max(tail_y), 0.001, 'full curve must show the fail bump')
        self.assertTrue(all(y == 0 for x, y in filtered if 29.0 <= x <= 31.0),
                        'filtered curve must exclude the fail bump')


class ChartConfigFilterTests(SimpleTestCase):
    """Chart-config switches on the analysis endpoints.

    Locks down the four new switches — ignore_no_test_value /
    data_only_bin1 / only_fail_test_item / only_low_cpk — on the
    histogram fast path (param list), the histogram compute path,
    site_stats and serial_distribution. Reuses the fake
    ``_load_df_from_request`` harness from StaleParamAcrossFileSwitchTests
    (no DB); the fake user has no ``settings`` attribute, which also pins
    the ``get_cpk_b_threshold`` fallback path.
    """

    METADATA = {
        'format': 'CTA8290D',
        'mins': {'Param0': '8.0', 'Param1': '0.5', 'FailOnly': '8.0',
                 'ParamBin1Empty': '8.0', 'ParamLowCpk': '8.0',
                 'ParamHighCpk': '8.0', 'ParamMidCpk': '8.0',
                 'ParamOutlierLow': '8.0'},
        'maxs': {'Param0': '14.3', 'Param1': '3.5', 'FailOnly': '14.3',
                 'ParamBin1Empty': '14.3', 'ParamLowCpk': '14.3',
                 'ParamHighCpk': '14.3', 'ParamMidCpk': '14.3',
                 'ParamOutlierLow': '14.3'},
        'units': {'Param0': 'V', 'Param1': 'uA', 'FailOnly': 'V',
                  'ParamBin1Empty': 'V', 'ParamLowCpk': 'V',
                  'ParamHighCpk': 'V', 'ParamMidCpk': 'V',
                  'ParamOutlierLow': 'V'},
    }

    def _frame(self):
        """10 rows: pass-bin rows are 0,2,3,4,6,7,8,9; fail rows 1,5."""
        return pd.DataFrame({
            'serial': list(range(101, 111)),
            'SW_Bin': [1, 2, 1, 1, 1, 2, 1, 1, 1, 1],
            'Site': [1, 2, 1, 2, 1, 2, 1, 2, 1, 2],
            'Param0': [10.1, 10.2, 10.3, 10.4, 10.5, 10.6, 10.7, 10.8, 10.9, 11.0],
            'Param1': [1.9, 2.0, 2.05, 1.95, 2.1, 1.98, 2.02, 1.96, 2.04, 2.0],
            # fail rows (1,5) exceed USL 14.3 → the only fail test item
            'FailOnly': [10.1, 15.0, 10.3, 10.4, 10.5, 15.0, 10.7, 10.8, 10.9, 11.0],
            'ParamAllNan': [np.nan] * 10,
            # values exist only on fail rows → empty inside Bin1
            'ParamBin1Empty': [np.nan, 10.1, np.nan, np.nan, np.nan,
                               10.2, np.nan, np.nan, np.nan, np.nan],
            # cpk ≈ 0.9 → low
            'ParamLowCpk': [8.5, 8.6, 8.7, 8.8, 8.9, 8.4, 8.6, 8.8, 9.0, 9.2],
            # cpk ≫ 1.33 → not low
            'ParamHighCpk': [10.9, 11.0, 11.05, 11.0, 11.02, 11.03, 10.98, 11.01, 10.99, 11.0],
            # cpk ≈ 1.43 → between 1.33 and 2.0, used for the settings test
            'ParamMidCpk': [8.6, 8.66, 8.72, 8.78, 8.84, 8.96, 9.02, 9.08, 9.14, 9.2],
            # 行 0（bin1）有 14.4 超 usl 14.3 的异常值：全量 CPK 低、filtered CPK 健康
            'ParamOutlierLow': [14.4, 11.0, 11.05, 11.0, 11.02,
                                11.0, 10.98, 11.01, 10.99, 11.0],
            'ParamNoLimit': [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0],
        })

    def _post(self, payload, user=None):
        factory, force_authenticate, restore = \
            StaleParamAcrossFileSwitchTests._patched_view(self._frame(), metadata=self.METADATA)
        self.addCleanup(restore)
        if user is None:
            user = types.SimpleNamespace(
                pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
                is_staff=False, is_superuser=False,
            )
        request = factory.post('/api/v1/analysis/histogram/', payload, format='json')
        force_authenticate(request, user=user)
        return request

    # ── fast path (param list) ──────────────────────────────────────

    def test_fast_path_default_list_keeps_sparse_but_skips_empty(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        keys = set(resp.data['results'].keys())
        self.assertIn('ParamBin1Empty', keys)   # has values on fail rows
        self.assertNotIn('ParamAllNan', keys)   # fully empty, always excluded

    def test_fast_path_data_only_bin1_and_ignore_no_test_value(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'data_only_bin1': True,
                              'ignore_no_test_value': True})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        keys = set(resp.data['results'].keys())
        self.assertNotIn('ParamBin1Empty', keys)  # empty inside Bin1 → hidden
        self.assertIn('Param0', keys)

    def test_fast_path_only_fail_test_item(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'only_fail_test_item': True})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(set(resp.data['results'].keys()), {'FailOnly'})

    def test_fast_path_only_low_cpk_default_threshold(self):
        """Fake user has no settings row → fallback 1.33, no crash."""
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'only_low_cpk': True})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(set(resp.data['results'].keys()), {'ParamLowCpk'})
        self.assertNotIn('ParamMidCpk', resp.data['results'])  # 1.43 > 1.33
        # FailOnly 含超限异常值（15.0）→ filtered CPK 健康 → 不列入（显示口径）

    def test_fast_path_only_low_cpk_uses_filtered_cpk(self):
        """低 CPK 判定无条件跟随 CPK 卡显示口径：异常值拉低的参数不列入。

        Regression: 用户反馈「3.1897 (filtered)」这种剔除异常值后 CPK 健康
        的参数也被列入了低 CPK 列表。前端 CPK 卡总是优先显示 filtered_cpk
        （与异常值处理开关无关），因此默认（无 outlier_handling 参数）时
        也不能把 ParamOutlierLow（含 14.4 异常值、filtered 健康）列入。
        """
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'only_low_cpk': True})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertNotIn('ParamOutlierLow', resp.data['results'])
        self.assertIn('ParamLowCpk', resp.data['results'])

    def test_fast_path_only_low_cpk_reads_user_setting(self):
        user = types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
            settings=types.SimpleNamespace(cpk_b_threshold=2.0),
        )
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'only_low_cpk': True}, user=user)
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertIn('ParamMidCpk', resp.data['results'])  # 1.43 < 2.0

    # ── compute path ────────────────────────────────────────────────

    def test_compute_path_data_only_bin1_total_count(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'params': ['Param0'],
                              'data_only_bin1': True})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['results']['Param0']['total_count'], 8)

    def test_compute_path_only_fail_test_item_drops_non_fail(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'params': ['FailOnly', 'Param1'],
                              'only_fail_test_item': True})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(set(resp.data['results'].keys()), {'FailOnly'})

    # ── site_stats / serial_distribution ────────────────────────────

    def test_site_stats_data_only_bin1_drops_bin2_site(self):
        from apps.analysis.views import statistics_views, StatisticsViewSet
        df = pd.DataFrame({
            'serial': list(range(101, 111)),
            'SW_Bin': [1, 1, 1, 1, 1, 2, 2, 2, 2, 2],
            'Site': [1, 1, 1, 1, 1, 2, 2, 2, 2, 2],
            'Param0': [10.1, 10.2, 10.3, 10.4, 10.5, 10.6, 10.7, 10.8, 10.9, 11.0],
        })
        factory, force_authenticate, restore = \
            StaleParamAcrossFileSwitchTests._patched_view(df, metadata=self.METADATA)
        self.addCleanup(restore)

        request = factory.post('/api/v1/statistics/site_stats/', {
            'file_id': 1, 'param': 'Param0', 'data_only_bin1': True,
        }, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False))
        resp = StatisticsViewSet.as_view({'post': 'site_stats'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        site_names = [s['Site'] for s in resp.data['site_data']]
        self.assertEqual(site_names, ['Site1', 'ALL Site'])  # Site2 gone

    def test_site_stats_display_format(self):
        """site_stats 展示格式：Yield=百分比3位小数(总数)、Fail/<Min/>Max=数量(百分比3位)。

        回归：用户指定格式（2026-08-13）——Yield 改为 `100.000%(5865)`，
        Fail/<Min/>Max 改为 `3(0.181%)`；数字字段 FailCountNum/TotalNum 保留供
        前端行样式等逻辑使用。
        """
        from apps.analysis.views import StatisticsViewSet
        df = pd.DataFrame({
            'serial': list(range(101, 111)),
            'SW_Bin': [1] * 10,
            'Site': [1, 1, 1, 1, 1, 2, 2, 2, 2, 2],
            'Param0': [10.1, 10.2, 10.3, 10.4, 10.5, 10.6, 10.7, 10.8, 10.9, 11.0],
        })
        factory, force_authenticate, restore = \
            StaleParamAcrossFileSwitchTests._patched_view(df, metadata=self.METADATA)
        self.addCleanup(restore)
        request = factory.post('/api/v1/statistics/site_stats/', {
            'file_id': 1, 'param': 'Param0',
        }, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False))
        resp = StatisticsViewSet.as_view({'post': 'site_stats'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        import re
        for row in resp.data['site_data']:
            self.assertRegex(row['Yield'], r'^\d+\.\d{3}%\(\d+\)$', row)
            self.assertRegex(row['FailCount'], r'^\d+\(\d+\.\d{3}%\)$', row)
            self.assertRegex(row['ExceedMin'], r'^\d+\(\d+\.\d{3}%\)$', row)
            self.assertRegex(row['ExceedMax'], r'^\d+\(\d+\.\d{3}%\)$', row)
            self.assertIsInstance(row['FailCountNum'], int)
            self.assertIsInstance(row['TotalNum'], int)
        # Param0 范围 [8.0, 14.3]：10 行全部在限内 → yield 100.000%(10)
        self.assertEqual(resp.data['site_data'][0]['Yield'], '100.000%(5)')

    def test_serial_distribution_data_only_bin1(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'Param0',
                              'data_only_bin1': True})
        resp = AnalysisViewSet.as_view({'post': 'serial_distribution'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        total = sum(len(s['data']) for s in resp.data['series_data'])
        self.assertEqual(total, 8)  # 10 rows − 2 fail rows

    def test_serial_distribution_full_frame_keeps_all_points(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'Param0'})
        resp = AnalysisViewSet.as_view({'post': 'serial_distribution'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        total = sum(len(s['data']) for s in resp.data['series_data'])
        self.assertEqual(total, 10)

    # ── qqplot / boxplot row filter ─────────────────────────────────

    def test_qqplot_data_only_bin1(self):
        """data_only_bin1 narrows the QQ plot to pass-bin rows (n = 8)."""
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'Param0',
                              'data_only_bin1': True})
        resp = AnalysisViewSet.as_view({'post': 'qqplot'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['n'], 8)  # 10 rows − 2 fail rows

    def test_qqplot_full_frame_keeps_all_points(self):
        """Without the switch the QQ plot still covers the whole frame."""
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'Param0'})
        resp = AnalysisViewSet.as_view({'post': 'qqplot'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['n'], 10)

    def test_qqplot_returns_fit_params(self):
        """QQ plot response carries the probplot fit line parameters:
        intercept≈data mean, slope≈data std (the front end draws the
        reference line as y = intercept + slope·x instead of y=x, which
        only fits zero-mean/unit-variance data)."""
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'Param0'})
        resp = AnalysisViewSet.as_view({'post': 'qqplot'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        slope, intercept = resp.data['slope'], resp.data['intercept']
        self.assertIsNotNone(slope)
        self.assertIsNotNone(intercept)
        # Param0 = 10.1..11.0 → mean ≈ 10.5, std ≈ 0.3
        self.assertLess(abs(intercept - 10.5), 1)
        self.assertLess(abs(slope - 0.3), 0.3)

    def test_qqplot_constant_data_horizontal_fit_line(self):
        """Constant data yields the horizontal fit line slope=0 /
        intercept=const (the front end draws y = intercept + slope·x,
        which degenerates to y = const — all points sit on it)."""
        from apps.analysis.services.statistics.computations import compute_qqplot
        result = compute_qqplot(pd.Series([5.0, 5.0, 5.0]))
        self.assertEqual(result['slope'], 0.0)
        self.assertEqual(result['intercept'], 5.0)
        self.assertFalse(result['is_normal'])
        # Short data path carries the fields (None — no fit computed)
        short = compute_qqplot(pd.Series([1.0, 2.0]))
        self.assertIsNone(short['slope'])
        self.assertIsNone(short['intercept'])

    def test_qqplot_data_only_bin1_empty_in_bin1_400(self):
        """Param all-NaN inside Bin1 → existing param_no_valid_data 400,
        never a 500 (filtering must happen before series extraction)."""
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'ParamBin1Empty',
                              'data_only_bin1': True})
        resp = AnalysisViewSet.as_view({'post': 'qqplot'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertEqual(resp.data.get('error'), 'param_no_valid_data')

    def test_boxplot_data_only_bin1_overall_count(self):
        """data_only_bin1 narrows the box plot overall stats (count = 8)."""
        from apps.analysis.views import StatisticsViewSet
        request = self._post({'file_id': 1, 'params': ['Param0'],
                              'data_only_bin1': True})
        resp = StatisticsViewSet.as_view({'post': 'boxplot'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['results']['Param0']['overall']['count'], 8)

    def test_boxplot_data_only_bin1_by_bin_only_bin1(self):
        """Grouped by bin, the filtered frame leaves only pass bin "1"
        (fail rows are removed before grouping — consistent with the
        histogram Bin1 mode)."""
        from apps.analysis.views import StatisticsViewSet
        request = self._post({'file_id': 1, 'params': ['Param0'],
                              'data_only_bin1': True, 'group_by': 'bin'})
        resp = StatisticsViewSet.as_view({'post': 'boxplot'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        by_bin = resp.data['results']['Param0']['by_bin']
        self.assertEqual(set(by_bin.keys()), {'1'})
        self.assertEqual(by_bin['1']['count'], 8)


class MultiFileCorrelationFilterApiTests(SimpleTestCase):
    """数据筛选开关（忽略无Limit/无测试值/仅Pass/仅Fail/低CPK）在
    multi_lot / correlation / correlation_matrix 端点的口径（2026-08-20）。

    与 histogram 同口径的关键不变量：fail 集合必须基于全量 df 预计算
    （bin1 过滤前）——fail 行永远不是 Bin1，先过滤再算 fail 集会清空
    fail 集合。fixture 复用 ChartConfigFilterTests 的帧结构
    （SW_Bin 行 1/5 为 fail；FailOnly 是唯一 fail 测试项）。
    """

    METADATA = ChartConfigFilterTests.METADATA

    def _frame(self):
        return ChartConfigFilterTests._frame(self)

    @staticmethod
    def _fake_objects(by_id):
        """DataFile 查询桩：filter(pk=...) → first() 返回对应 fake 对象。"""
        class _FakeQS:
            def __init__(self, obj):
                self._obj = obj
            def first(self):
                return self._obj

        class _FakeManager:
            def filter(self, **kw):
                return _FakeQS(by_id.get(kw.get('pk')))

        return _FakeManager()

    def _multi_lot(self, payload, frames):
        """打桩 multi_lot 的 DataFile 查询与文件加载，在 patch 作用域内调用视图。

        注意：patch 必须覆盖「构建请求 → 视图执行」全程，不能在请求构建完
        就退出（视图内会再查 DataFile.objects）。
        """
        import types
        from unittest.mock import patch
        from rest_framework.test import APIRequestFactory, force_authenticate
        from apps.analysis.views import analysis_views, AnalysisViewSet

        objs = {}
        for fid in frames:
            objs[fid] = types.SimpleNamespace(
                id=fid, filename=f'f{fid}.csv', format_type='CTA8290D',
                file_path='/tmp/__nonexistent__',  # os.stat 失败 → 缓存 key 无 mtime 守卫
            )

        def fake_load(fid, user_id, obj=None):
            df, meta = frames[fid]
            return df, meta, 'CTA8290D'

        with patch.object(analysis_views.DataFile, 'objects', self._fake_objects(objs)), \
             patch.object(analysis_views, 'get_cached_parsed_file', fake_load):
            factory = APIRequestFactory()
            request = factory.post('/api/v1/analysis/multi_lot/', payload, format='json')
            force_authenticate(request, user=types.SimpleNamespace(
                pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
                is_staff=False, is_superuser=False,
            ))
            resp = AnalysisViewSet.as_view({'post': 'multi_lot'})(request)
            resp.render()
            return resp

    def _correlation_request(self, payload, df=None):
        """打桩 correlation/correlation_matrix 的 _load_df_from_request。"""
        import types
        from rest_framework.test import APIRequestFactory, force_authenticate
        factory, force_auth, restore = StaleParamAcrossFileSwitchTests._patched_view(
            df if df is not None else self._frame(), metadata=self.METADATA)
        self.addCleanup(restore)
        request = factory.post('/api/v1/analysis/correlation/', payload, format='json')
        force_auth(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))
        return request

    # ── multi_lot ───────────────────────────────────────────────────

    def test_multi_lot_data_only_bin1_reduces_distribution_count(self):
        """data_only_bin1：首个公共参数分布的 lot count 从 10 收缩到 8。"""
        f1 = self._frame()
        f2 = self._frame()
        resp = self._multi_lot(
            {'file_ids': [1, 2], 'data_only_bin1': True},
            {1: (f1, self.METADATA), 2: (f2, self.METADATA)})
        self.assertEqual(resp.status_code, 200, resp.content)
        lots = {lot['file_id']: lot for lot in resp.data['lot_data']}
        self.assertEqual(lots[1]['count'], 8, 'bin1 过滤后应只剩 8 行')
        self.assertEqual(lots[2]['count'], 8)

    def test_multi_lot_only_fail_test_item_keeps_fail_items_from_full_df(self):
        """仅显示Fail测试项：common_params 只含 FailOnly。

        回归不变量：fail 集合基于全量 df 预计算（bin1 过滤前）——FailOnly 的
        fail 行（1/5）不是 Bin1，若先过滤行再算 fail 集则 fail 恒空 →
        common_params 空（旧实现路径必失败）。
        """
        f1 = self._frame()
        f2 = self._frame()
        resp = self._multi_lot(
            {'file_ids': [1, 2], 'only_fail_test_item': True},
            {1: (f1, self.METADATA), 2: (f2, self.METADATA)})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['common_params'], ['FailOnly'])

    def test_multi_lot_only_fail_test_item_with_bin1_keeps_fail_items(self):
        """仅Fail + 仅Pass 同时开：fail 集仍来自全量 df（组合开关不互相破坏）。"""
        f1 = self._frame()
        f2 = self._frame()
        resp = self._multi_lot(
            {'file_ids': [1, 2], 'only_fail_test_item': True, 'data_only_bin1': True},
            {1: (f1, self.METADATA), 2: (f2, self.METADATA)})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['common_params'], ['FailOnly'])

    def test_multi_lot_ignore_no_test_value_filters_sparse_params(self):
        """忽略无测试值 + 仅Pass：ParamBin1Empty（值只在 fail 行）从 common 中移除。

        单独开 ignore_no_test_value 时 ParamBin1Empty 有 20% 有效值（2/10）
        ≥ 5% 阈值会被保留——与单文件口径一致；组合 data_only_bin1 后
        bin1 行内全空 → 剔除。
        """
        f1 = self._frame()
        f2 = self._frame()
        resp = self._multi_lot(
            {'file_ids': [1, 2], 'ignore_no_test_value': True,
             'data_only_bin1': True},
            {1: (f1, self.METADATA), 2: (f2, self.METADATA)})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertNotIn('ParamBin1Empty', resp.data['common_params'])
        self.assertIn('Param0', resp.data['common_params'])

    # ── correlation ─────────────────────────────────────────────────

    def test_correlation_data_only_bin1_changes_n(self):
        """correlation：bin1 过滤行后 n 从 10 → 8。"""
        from apps.analysis.views import AnalysisViewSet
        request = self._correlation_request(
            {'file_id': 1, 'param_x': 'Param0', 'param_y': 'Param1',
             'data_only_bin1': True})
        resp = AnalysisViewSet.as_view({'post': 'correlation'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['n'], 8)

    def test_correlation_only_fail_test_item_400_when_param_filtered(self):
        """correlation + 仅Fail：param_x=Param0（非 fail 项）→ 400 no_valid_params。"""
        from apps.analysis.views import AnalysisViewSet
        request = self._correlation_request(
            {'file_id': 1, 'param_x': 'Param0', 'param_y': 'FailOnly',
             'only_fail_test_item': True})
        resp = AnalysisViewSet.as_view({'post': 'correlation'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertEqual(resp.data.get('error'), 'no_valid_params')

    # ── correlation_matrix ──────────────────────────────────────────

    def test_correlation_matrix_data_only_bin1_ok(self):
        """correlation_matrix：bin1 过滤行后正常 200（参数列表仍 ≥2）。"""
        from apps.analysis.views import StatisticsViewSet
        request = self._correlation_request(
            {'file_id': 1, 'params': ['Param0', 'Param1', 'FailOnly'],
             'data_only_bin1': True})
        resp = StatisticsViewSet.as_view({'post': 'correlation_matrix'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)

    def test_correlation_matrix_only_fail_test_item_400_when_below_two(self):
        """correlation_matrix + 仅Fail：重放后只剩 FailOnly 一个参数 → 400。"""
        from apps.analysis.views import StatisticsViewSet
        request = self._correlation_request(
            {'file_id': 1, 'params': ['Param0', 'Param1', 'FailOnly'],
             'only_fail_test_item': True})
        resp = StatisticsViewSet.as_view({'post': 'correlation_matrix'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertEqual(resp.data.get('error'), 'need_at_least_2_params')


class SerialColumnPartIdFallbackTests(SimpleTestCase):
    """STS8200 文件（无 Serial 列）的序列分布回退到 PART_ID。

    回归：STS8200 数据列头是 SITE_NUM/PART_ID/...，没有 Serial_No，
    ``get_serial_column`` 只匹配 "serial" 导致 serial_distribution 返回
    400 no_serial_column，序列分布图无法绘制。修复：helpers 层增加
    'part'+'id' 列回退（与 wafer_map 既有逻辑一致），本测试钉住
    PART_ID 作为序列列 + 按 site 分组 + bin 失败判定的完整行为。
    """

    METADATA = {
        'format': 'STS8200',
        'mins': {'CONT_GATE': '-0.65'},
        'maxs': {'CONT_GATE': '-0.47'},
        'units': {'CONT_GATE': 'V'},
    }

    def _frame(self):
        # 8 行：PART_ID 是每 site 内的部件序号（1..4），无任何 serial 列
        return pd.DataFrame({
            'SITE_NUM': [8, 8, 7, 8, 7, 8, 7, 6],
            'PART_ID': [1, 2, 1, 3, 2, 4, 3, 1],
            'PASSFG': ['True', 'False', 'True', 'True', 'True',
                       'False', 'True', 'True'],
            'SOFT_BIN': [1, 5, 1, 1, 1, 2, 1, 1],
            'CONT_GATE': [-0.54, -0.54, -0.55, -0.53, -0.54, -0.52, -0.55, -0.54],
        })

    def _post(self, payload):
        factory, force_authenticate, restore = \
            StaleParamAcrossFileSwitchTests._patched_view(
                self._frame(), metadata=self.METADATA)
        self.addCleanup(restore)
        import types
        request = factory.post('/api/v1/analysis/serial_distribution/',
                               payload, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False))
        return request

    def test_serial_distribution_falls_back_to_part_id(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'CONT_GATE'})
        resp = AnalysisViewSet.as_view({'post': 'serial_distribution'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['serial_col'], 'PART_ID')
        # 按 site 分组：Site 6/7/8 各一个 series，共 8 个点
        names = [s['name'] for s in resp.data['series_data']]
        self.assertEqual(names, ['Site 6', 'Site 7', 'Site 8'])
        total = sum(len(s['data']) for s in resp.data['series_data'])
        self.assertEqual(total, 8)
        # 连续序列 x 轴 = PART_ID 1..4
        self.assertEqual(resp.data['continuous_serials'], [1, 2, 3, 4])
        # bin 判定：SOFT_BIN 5（行1）与 2（行5）为失败
        self.assertEqual(resp.data['pass_count'], 6)
        self.assertEqual(resp.data['fail_count'], 2)

    def test_part_id_rejected_as_data_param(self):
        """PART_ID 是分组键：作为数据参数必须 400（与 Serial_No 同规则）。"""
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'param': 'PART_ID'})
        resp = AnalysisViewSet.as_view({'post': 'serial_distribution'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertEqual(resp.data['error'], 'param_is_metadata')


class SerialColumnCandidatesTests(SimpleTestCase):
    """CTA8280F Dut_No 序列分布适配 + 多候选 serial 列优先级。

    回归：``Site12358-Chip12345_c.csv``（真实机台导出）[Data] 表头
    无 Serial_No/Index_No，只有 Dut_No（100 × 5 site = 500 dies），
    旧 ``get_serial_column`` 只匹配 serial / part+id → 序列分布 400
    no_serial_column。修复：helpers 集中式候选检测（serial > dut >
    part+id 优先级，Dut_Pass 等布尔列不误匹配）+ 端点 ``serial_col``
    覆盖参数 + 响应 ``serial_candidates``。
    """

    METADATA = {
        'format': 'CTA8280F',
        'mins': {'KELVIN_VIN': '0'},
        'maxs': {'KELVIN_VIN': '10'},
        'units': {'KELVIN_VIN': 'ohm'},
    }

    def _frame_c_shape(self, sites=(1, 2, 3, 5, 8), duts=100):
        """_c 形态：无 Serial_No，Dut_No × Site_No 唯一确定一颗 die。"""
        rows = []
        for s in sites:
            for d in range(1, duts + 1):
                rows.append({
                    'Dut_No': d, 'Site_No': s, 'Dut_Pass': True,
                    'SW_Bin': 1, 'KELVIN_VIN': round(0.5 + d * 0.01, 4),
                })
        return pd.DataFrame(rows)

    def _frame_n_shape(self, duts=3):
        """_n 形态：Serial_No 与 Dut_No 并存（Serial_No 列序在后）。"""
        df = self._frame_c_shape(duts=duts)
        df.insert(0, 'Serial_No', range(1, len(df) + 1))
        return df

    def _patched_post(self, df, payload):
        from apps.analysis.views import AnalysisViewSet
        factory, force_authenticate, restore = \
            StaleParamAcrossFileSwitchTests._patched_view(df, metadata=self.METADATA)
        self.addCleanup(restore)
        request = factory.post('/api/v1/analysis/serial_distribution/',
                               {'file_id': 1, **payload}, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False))
        return AnalysisViewSet.as_view({'post': 'serial_distribution'})(request)

    def test_get_serial_column_prefers_serial_over_dut(self):
        from apps.analysis.services.statistics import get_serial_column
        df = self._frame_n_shape()
        # 优先级：Serial_No 存在时必须选它，即使 Dut_No 列序在前
        self.assertEqual(get_serial_column(df), 'Serial_No')
        # 显式覆盖优先
        self.assertEqual(get_serial_column(df, preferred='Dut_No'), 'Dut_No')
        self.assertEqual(get_serial_column(df, preferred='Nope'), 'Serial_No')

    def test_get_serial_column_falls_back_to_dut_no(self):
        from apps.analysis.services.statistics import get_serial_column
        df = self._frame_c_shape()
        self.assertEqual(get_serial_column(df), 'Dut_No')

    def test_get_serial_candidates_order_and_dut_pass_excluded(self):
        from apps.analysis.services.statistics import get_serial_candidates
        df = self._frame_n_shape()
        self.assertEqual(get_serial_candidates(df), ['Serial_No', 'Dut_No'])
        # Dut_Pass 布尔列不得误匹配
        self.assertNotIn('Dut_Pass', get_serial_candidates(df))
        df2 = self._frame_c_shape()
        self.assertEqual(get_serial_candidates(df2), ['Dut_No'])
        # 既无 serial 也无 dut/part 列 → 空候选
        df3 = pd.DataFrame({'Site_No': [1, 1], 'SW_Bin': [1, 1],
                            'KELVIN_VIN': [0.5, 0.6]})
        self.assertEqual(get_serial_candidates(df3), [])

    def test_serial_distribution_dut_no_full_flow(self):
        """端点全链路：500 行 _c 形态 → 200、Dut_No、500 点、按 site 分组。"""
        df = self._frame_c_shape()
        df.loc[len(df) - 1, 'SW_Bin'] = 5  # 最后一颗 die fail
        resp = self._patched_post(df, {'param': 'KELVIN_VIN'})
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['serial_col'], 'Dut_No')
        self.assertEqual(resp.data['serial_candidates'], ['Dut_No'])
        total = sum(len(s['data']) for s in resp.data['series_data'])
        self.assertEqual(total, 500)
        names = [s['name'] for s in resp.data['series_data']]
        self.assertEqual(names, ['Site 1', 'Site 2', 'Site 3', 'Site 5', 'Site 8'])
        self.assertEqual(resp.data['continuous_serials'][0], 1)
        self.assertEqual(resp.data['continuous_serials'][-1], 100)
        self.assertEqual(resp.data['pass_count'], 499)
        self.assertEqual(resp.data['fail_count'], 1)

    def test_serial_col_override_and_candidates(self):
        """serial_col 覆盖：_n 形态显式选 Dut_No → 响应用 Dut_No。"""
        df = self._frame_n_shape()
        resp = self._patched_post(df, {'param': 'KELVIN_VIN',
                                       'serial_col': 'Dut_No'})
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data['serial_col'], 'Dut_No')
        self.assertEqual(resp.data['serial_candidates'],
                         ['Serial_No', 'Dut_No'])

    def test_serial_col_invalid_400(self):
        df = self._frame_c_shape()
        resp = self._patched_post(df, {'param': 'KELVIN_VIN',
                                       'serial_col': 'Nope'})
        resp.render()
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertEqual(resp.data['error'], 'serial_col_not_found')

    def test_serial_distribution_no_candidate_400(self):
        df = pd.DataFrame({'Site_No': [1, 1], 'SW_Bin': [1, 1],
                           'KELVIN_VIN': [0.5, 0.6]})
        resp = self._patched_post(df, {'param': 'KELVIN_VIN'})
        resp.render()
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertEqual(resp.data['error'], 'no_serial_column')

    def test_param_list_excludes_all_serial_candidates(self):
        """_n 形态（Serial_No + Dut_No 并存）：参数列表两者都不能出现。

        回归：旧逻辑只排除自动选中的第一个 serial 列，Dut_No 作为数值列
        泄漏进参数选择器，用户可把它当直方图参数（本身是分组键）。
        """
        from apps.analysis.views import AnalysisViewSet
        df = self._frame_n_shape()
        df['KELVIN_SW'] = df['KELVIN_VIN'] * 2  # 额外数值列
        factory, force_authenticate, restore = \
            StaleParamAcrossFileSwitchTests._patched_view(df, metadata=self.METADATA)
        self.addCleanup(restore)
        request = factory.post('/api/v1/analysis/histogram/',
                               {'file_id': 1}, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False))
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        params = list(resp.data['results'].keys())
        self.assertNotIn('Serial_No', params)
        self.assertNotIn('Dut_No', params)
        self.assertNotIn('Dut_Pass', params)
        self.assertIn('KELVIN_VIN', params)
        self.assertIn('KELVIN_SW', params)



class FileCorrelationCacheIsolationTests(SimpleTestCase):
    """file_correlation 端点不得污染 LRU 解析缓存。

    回归：file_correlation 曾对 ``get_cached_parsed_file`` 返回的缓存
    DataFrame 原地添加 ``__serial__`` 辅助列，违反缓存"只读"不变量，
    导致后续 browse/histogram/export 请求同一文件时出现幽灵列。
    修复：写列前 ``df = df.copy()``。
    """

    def test_shared_cache_df_not_mutated_by_file_correlation(self):
        import types
        from rest_framework.test import APIRequestFactory, force_authenticate
        from apps.analysis.views import analysis_views

        shared_df = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, 2.0, 3.0],
            'ParamB': [4.0, 5.0, 6.0],
        })
        metadata = {'format': 'CTA8290D', 'mins': {}, 'maxs': {}, 'units': {}}

        orig_404 = analysis_views.get_object_or_404
        orig_load = analysis_views.get_cached_parsed_file
        orig_serial = analysis_views.get_serial_column

        def fake_404(model, *args, **kwargs):
            return types.SimpleNamespace(id=1, filename='fake.csv',
                                         format_type='CTA8290D')

        def fake_load(fid, owner_id, datafile=None):
            # 两个文件返回同一个缓存对象 —— 模拟 LRU 缓存真实行为
            return shared_df, metadata, 'CTA8290D'

        analysis_views.get_object_or_404 = fake_404
        analysis_views.get_cached_parsed_file = fake_load
        analysis_views.get_serial_column = lambda df: 'Serial_No'
        self.addCleanup(lambda: setattr(analysis_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(analysis_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(analysis_views, 'get_serial_column', orig_serial))

        factory = APIRequestFactory()
        request = factory.post('/api/v1/analysis/file_correlation/', {
            'file1_id': 1, 'file2_id': 2,
            # mock 的 metadata 无 limits —— 关掉 ignore_no_limit 让参数参与对比
            'ignore_no_limit': False,
        }, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))

        from apps.analysis.views import AnalysisViewSet
        view = AnalysisViewSet.as_view({'post': 'file_correlation'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        # 缓存对象不允许被端点改写
        self.assertNotIn('__serial__', shared_df.columns)
        # 端点自身逻辑仍可用（3 个共同序列号被匹配；序列列不参与参数对比，
        # 只比较 ParamA/ParamB）
        self.assertEqual(response.data['serials'], [1, 2, 3])
        self.assertEqual(response.data['params'], ['ParamA', 'ParamB'])


class FileCorrelationNanJsonTests(SimpleTestCase):
    """file_correlation 对含 NaN 的参数列必须返回合法 JSON（回归 500）。

    回归：偏差计算把 NaN 单元格直接算进 diff_pct，nan 进入 summary 后
    DRF JSON 序列化抛 ``ValueError: Out of range float values are not
    JSON compliant: nan`` → 整个端点 500。修复：任一侧非有限值即跳过
    该序列号；响应再经 clean_data 兜底。
    """

    def _call_endpoint(self, shared_df, serial='Serial_No', extra=None):
        from rest_framework.test import APIRequestFactory, force_authenticate
        from apps.analysis.views import analysis_views, AnalysisViewSet

        metadata = {'format': 'CTA8290D', 'mins': {}, 'maxs': {}, 'units': {}}

        orig_404 = analysis_views.get_object_or_404
        orig_load = analysis_views.get_cached_parsed_file
        orig_serial = analysis_views.get_serial_column

        analysis_views.get_object_or_404 = (
            lambda model, *a, **k: types.SimpleNamespace(
                id=1, filename='fake.csv', format_type='CTA8290D'))
        analysis_views.get_cached_parsed_file = (
            lambda fid, owner_id, datafile=None: (shared_df, metadata, 'CTA8290D'))
        analysis_views.get_serial_column = lambda df: serial
        self.addCleanup(lambda: setattr(analysis_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(analysis_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(analysis_views, 'get_serial_column', orig_serial))

        factory = APIRequestFactory()
        request = factory.post('/api/v1/analysis/file_correlation/', {
            'file1_id': 1, 'file2_id': 2, 'threshold': 3.0,
            # mock 的 metadata 无 limits —— 关掉 ignore_no_limit 让参数参与对比
            'ignore_no_limit': False,
            **(extra or {}),
        }, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))

        view = AnalysisViewSet.as_view({'post': 'file_correlation'})
        response = view(request)
        response.render()  # 触发 JSON 序列化——无异常即证明无 nan/inf 泄漏
        return response

    def test_nan_param_values_do_not_break_json_serialization(self):
        shared_df = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, float('nan'), 3.0],
            'ParamB': [4.0, 5.0, 6.0],
        })
        response = self._call_endpoint(shared_df)
        self.assertEqual(response.status_code, 200, response.content)

        by_param = {r['param']: r for r in response.data['rows']}
        # 序列号 2 的 ParamA 为 NaN → 该对不参与对比；ParamB 全部有效
        self.assertEqual(by_param['ParamA']['compared'], 2)
        self.assertEqual(by_param['ParamB']['compared'], 3)
        for r in response.data['rows']:
            self.assertTrue(math.isfinite(r['max_diff']), r)
            self.assertTrue(math.isfinite(r['pass_rate']), r)
        self.assertEqual(response.data['serials'], [1, 2, 3])

    def test_all_nan_param_values_returns_empty_comparison(self):
        shared_df = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [float('nan'), float('nan'), float('nan')],
        })
        # ignore_no_data=False：全 NaN 参数保留在结果中（compared=0），
        # 而不是被「忽略无数据」过滤掉
        response = self._call_endpoint(shared_df, extra={'ignore_no_data': False})
        self.assertEqual(response.status_code, 200, response.content)
        by_param = {r['param']: r for r in response.data['rows']}
        self.assertEqual(by_param['ParamA']['compared'], 0)
        self.assertEqual(by_param['ParamA']['fail_count'], 0)
        self.assertEqual(by_param['ParamA']['max_diff'], 0)
        self.assertEqual(by_param['ParamA']['pass_rate'], 0)


class FileCorrelationExportTests(TestCase):
    """file_correlation_export 输出双 Sheet xlsx。

    规格（2026-08：Limit 与测试值拆分为两个 Sheet）：
    - workbook 只含两个 Sheet：'Limit对比' + '测试值对比'（excelize 默认
      Sheet1 已删除，顺序即创建顺序）。
    - Limit 对比 Sheet：标题行（跨 A:I 合并）| 单行表头 Parameters | LSL A |
      USL A | LSL B | USL B | LSL Diff | USL Diff | Unit | 判定 | 每测试项
      一行；Diff 按 diff_rule 标红，判定列 PASS/FAIL（仅 limit 差异）。
    - 测试值对比 Sheet：模板布局——标题行（跨全表合并）| 两行表头
      A2/A3 'Parameters' + 'Limit / Unit (Data A)' 组（LSL A/USL A/Unit）
      + 每序列组标题（E2:H2…）+ 末列 Comment；数据行公式 =Bench-ATE /
      =Delta/ATE（Δ/ATE 口径），%Diff 数字格式 0.00%；|%Diff| > threshold
      → Delta/%Diff 标红（静态判定与 JSON 端点一致）；单侧有值只写该侧；
      Comment 只含超差摘要（N 超差 / PASS）。
    - 无公共序列 → limits_only：数据 Sheet 无序列块（仅 Parameters +
      Comment）。
    - 无公共测试项 → 400 no_common_params。
    """

    def setUp(self):
        from django.contrib.auth import get_user_model
        self.user = get_user_model().objects.create_user(
            username='fc_export_test', password='x')

    def _call_export(self, frames, body=None, metas=None):
        from rest_framework.test import APIClient
        from apps.analysis.views import analysis_views

        body = body or {}
        base_meta = {'format': 'CTA8290D',
                     'mins': {'ParamA': '0.5', 'ParamB': '10'},
                     'maxs': {'ParamA': '2.0', 'ParamB': '40'},
                     'units': {'ParamA': 'V', 'ParamB': 'nA'}}
        metas = metas or {}
        orig_404 = analysis_views.get_object_or_404
        orig_load = analysis_views.get_cached_parsed_file
        orig_serial = analysis_views.get_serial_column

        analysis_views.get_object_or_404 = (
            lambda model, *a, **k: types.SimpleNamespace(
                id=k['pk'], filename=f'FILE{k["pk"]}.csv',
                format_type='CTA8290D'))
        analysis_views.get_cached_parsed_file = (
            lambda fid, owner_id, datafile=None: (
                frames[int(fid)], metas.get(int(fid), base_meta), 'CTA8290D'))
        analysis_views.get_serial_column = lambda df: 'Serial_No'
        self.addCleanup(lambda: setattr(analysis_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(analysis_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(analysis_views, 'get_serial_column', orig_serial))

        client = APIClient()
        client.force_authenticate(user=self.user)
        return client.post('/api/v1/analysis/file_correlation_export/', {
            'file1_id': 1, 'file2_id': 2, **body,
        }, format='json')

    @staticmethod
    def _body(resp):
        """FileResponse is streaming — collect bytes for openpyxl / asserts."""
        return b''.join(resp.streaming_content)

    def test_export_two_sheets_template_layout_formulas_and_red_highlight(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, 2.0, 3.0],
            'ParamB': [10.0, 20.0, 30.0],
        })
        df2 = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.1, 2.1, 3.1],
            'ParamB': [10.0, 21.0, 29.0],
        })
        resp = self._call_export({1: df1, 2: df2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        self.assertIn('spreadsheetml', resp['Content-Type'])
        self.assertIn('FILE1_vs_FILE2_correlation.xlsx', resp['Content-Disposition'])

        wb = load_workbook(io.BytesIO(body))
        # 只含两个 Sheet（默认 Sheet1 已删除），顺序 = 创建顺序
        self.assertEqual(wb.sheetnames, ['Limit对比', '测试值对比'])
        ws_l = wb['Limit对比']
        ws_d = wb['测试值对比']

        # ── Limit 对比 Sheet：标题 + 单行表头 + 每测试项一行 ──
        self.assertEqual(ws_l['A1'].value, 'Data A VS Data B（Limit 对比）')
        self.assertIn('A1:I1', {str(r) for r in ws_l.merged_cells.ranges})
        for cell, h in [('A2', 'Parameters'), ('B2', 'LSL A'), ('C2', 'USL A'),
                        ('D2', 'LSL B'), ('E2', 'USL B'), ('F2', 'LSL Diff'),
                        ('G2', 'USL Diff'), ('H2', 'Unit'), ('I2', '判定')]:
            self.assertEqual(ws_l[cell].value, h)
        # 数据行（文件1 列顺序；diff 全 0 → 判定 PASS）
        self.assertEqual(ws_l['A3'].value, 'ParamA')
        self.assertEqual(ws_l['B3'].value, 0.5)
        self.assertEqual(ws_l['C3'].value, 2.0)
        self.assertEqual(ws_l['D3'].value, 0.5)
        self.assertEqual(ws_l['E3'].value, 2.0)
        self.assertEqual(ws_l['F3'].value, 0.0)
        self.assertEqual(ws_l['G3'].value, 0.0)
        self.assertEqual(ws_l['H3'].value, 'V')
        self.assertEqual(ws_l['I3'].value, 'PASS')
        self.assertEqual(ws_l['A4'].value, 'ParamB')
        self.assertEqual(ws_l['I4'].value, 'PASS')
        self.assertEqual(ws_l.max_row, 4)
        self.assertEqual(ws_l.max_column, 9)

        # ── 测试值对比 Sheet：模板布局（A 参数列 + Limit/Unit(Data A) 组，
        #    E 起序列块）── 3 序列 → 块 E..P，末列 Q（Comment）
        self.assertEqual(ws_d['A1'].value, 'Data A VS Data B（测试值对比）')
        merged = {str(r) for r in ws_d.merged_cells.ranges}
        self.assertIn('A1:Q1', merged)
        self.assertIn('A2:A3', merged)
        self.assertIn('B2:D2', merged)   # Limit / Unit (Data A) 组
        self.assertIn('E2:H2', merged)   # 序列 1 组标题
        self.assertIn('I2:L2', merged)   # 序列 2 组标题
        self.assertIn('M2:P2', merged)   # 序列 3 组标题

        self.assertEqual(ws_d['A2'].value, 'Parameters')   # A2:A3 合并（参数列跨两行）
        self.assertEqual(ws_d['B2'].value, 'Limit / Unit (Data A)')
        self.assertEqual(ws_d['E2'].value, 1)
        self.assertEqual(ws_d['I2'].value, 2)
        self.assertEqual(ws_d['M2'].value, 3)
        self.assertEqual(ws_d['Q2'].value, 'Comment')
        for cell, h in [('B3', 'LSL A'), ('C3', 'USL A'), ('D3', 'Unit'),
                        ('E3', 'ATE'), ('F3', 'Bench'), ('G3', 'Delta'), ('H3', '% Diff'),
                        ('I3', 'ATE'), ('J3', 'Bench'), ('K3', 'Delta'), ('L3', '% Diff'),
                        ('M3', 'ATE'), ('N3', 'Bench'), ('O3', 'Delta'), ('P3', '% Diff'),
                        ('Q3', 'Comment')]:
            self.assertEqual(ws_d[cell].value, h)

        # 数据行按文件1列顺序（ignore_no_limit 默认过滤无 limit 的序列列）
        self.assertEqual(ws_d['A4'].value, 'ParamA')
        self.assertEqual(ws_d['A5'].value, 'ParamB')
        self.assertEqual(ws_d.max_row, 5)
        self.assertEqual(ws_d.max_column, 17)   # Q

        # Data A 的 Limit/Unit（默认口径）
        self.assertEqual(ws_d['B4'].value, 0.5)
        self.assertEqual(ws_d['C4'].value, 2.0)
        self.assertEqual(ws_d['D4'].value, 'V')

        # 公式（Δ/ATE 口径 = Delta/ATE）+ 0.00% 数字格式；单侧有值只写该侧
        self.assertEqual(ws_d['E4'].value, 1.0)
        self.assertEqual(ws_d['F4'].value, 1.1)
        self.assertEqual(ws_d['G4'].value, '=F4-E4')
        self.assertEqual(ws_d['H4'].value, '=G4/E4')
        self.assertEqual(ws_d['H4'].number_format, '0.00%')

        # 序列1：ParamA |10%| > 3 → Delta/%Diff 标红；ParamB 0% 不标红
        self.assertEqual(
            ws_d['G4'].fill.start_color.rgb, ws_d['H4'].fill.start_color.rgb)
        self.assertNotEqual(
            ws_d['G4'].fill.start_color.rgb, ws_d['E4'].fill.start_color.rgb)
        self.assertEqual(
            ws_d['G5'].fill.start_color.rgb, ws_d['E5'].fill.start_color.rgb)

        # Comment 判定摘要（只含超差，Limit 差异在 Limit Sheet 判定列）
        self.assertEqual(ws_d['Q4'].value, '3 超差')   # ParamA: 10/5/3.33% 全超
        self.assertEqual(ws_d['Q5'].value, '2 超差')   # ParamB: 5/-3.33% 超

    def test_export_threshold_controls_red_highlight(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({
            'Serial_No': [1, 2],
            'ParamA': [1.0, 3.0],
        })
        df2 = pd.DataFrame({
            'Serial_No': [1, 2],
            'ParamA': [1.1, 3.1],
        })
        # 阈值放大到 100：ParamA 10%/3.33% 偏差不再标红
        resp = self._call_export({1: df1, 2: df2}, body={'threshold': 100.0})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['测试值对比']
        self.assertEqual(ws['A4'].value, 'ParamA')
        self.assertEqual(
            ws['G4'].fill.start_color.rgb, ws['E4'].fill.start_color.rgb)

    def test_export_single_side_value_writes_ate_only(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.1, float('nan')]})
        resp = self._call_export({1: df1, 2: df2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['测试值对比']
        # 序列2（I..L 块）：bench NaN → 只写 ATE，无公式（同模板单侧行）
        self.assertEqual(ws['I4'].value, 2.0)
        self.assertIsNone(ws['J4'].value)
        self.assertIsNone(ws['K4'].value)
        self.assertIsNone(ws['L4'].value)
        # 序列1（E..H 块）：两侧都有 → 公式
        self.assertEqual(ws['G4'].value, '=F4-E4')

    def test_export_diff_rule_zero_marks_different_limits_red(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        df2 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        # 文件 B 的 LSL 更宽（0.4 < 0.5）→ zero 规则下 Diff≠0 标红
        meta_b = {'format': 'CTA8290D',
                  'mins': {'ParamA': '0.4'}, 'maxs': {'ParamA': '2.0'},
                  'units': {'ParamA': 'V'}}
        resp = self._call_export({1: df1, 2: df2}, metas={2: meta_b})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['Limit对比']
        self.assertEqual(ws['F3'].value, -0.1)
        # LSL Diff 标红 + 判定 FAIL；USL Diff 相等不标红
        self.assertNotEqual(
            ws['F3'].fill.start_color.rgb, ws['G3'].fill.start_color.rgb)
        self.assertEqual(
            ws['F3'].fill.start_color.rgb, ws['I3'].fill.start_color.rgb)
        self.assertEqual(
            ws['G3'].fill.start_color.rgb, ws['H3'].fill.start_color.rgb)

    def test_export_diff_rule_wider_only_marks_tighter_limits_red(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        df2 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        # B 更宽（0.4/2.5）→ wider 规则 pass（不标红）
        meta_wider = {'format': 'CTA8290D',
                      'mins': {'ParamA': '0.4'}, 'maxs': {'ParamA': '2.5'},
                      'units': {'ParamA': 'V'}}
        resp = self._call_export({1: df1, 2: df2}, metas={2: meta_wider},
                                 body={'diff_rule': 'wider'})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['Limit对比']
        self.assertEqual(ws['F3'].value, -0.1)
        self.assertEqual(
            ws['F3'].fill.start_color.rgb, ws['H3'].fill.start_color.rgb)
        self.assertEqual(
            ws['G3'].fill.start_color.rgb, ws['H3'].fill.start_color.rgb)

        # B 更紧（LSL 0.6 > 0.5）→ wider 规则标红 + 判定 FAIL
        meta_tight = {'format': 'CTA8290D',
                      'mins': {'ParamA': '0.6'}, 'maxs': {'ParamA': '2.0'},
                      'units': {'ParamA': 'V'}}
        resp2 = self._call_export({1: df1, 2: df2}, metas={2: meta_tight},
                                  body={'diff_rule': 'wider'})
        body2 = self._body(resp2)
        self.assertEqual(resp2.status_code, 200, body2[:500])
        ws2 = load_workbook(io.BytesIO(body2))['Limit对比']
        self.assertNotEqual(
            ws2['F3'].fill.start_color.rgb, ws2['H3'].fill.start_color.rgb)

    def test_export_max_serials_truncates_sequence_blocks(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2, 3, 4], 'ParamA': [1.0, 2.0, 3.0, 4.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2, 3, 4], 'ParamA': [1.0, 2.0, 3.0, 4.0]})
        resp = self._call_export({1: df1, 2: df2}, body={'max_serials': 2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['测试值对比']
        # 4 个公共序列只对比前 2 个 → 只有 2 组序列块；末列 M 为 Comment
        self.assertEqual(ws['E2'].value, 1)
        self.assertEqual(ws['I2'].value, 2)
        self.assertEqual(ws['M2'].value, 'Comment')
        self.assertEqual(ws.max_column, 13)   # M

    def test_export_ignore_no_data_filters_params(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2],
                            'ParamA': [1.0, 2.0],
                            'ParamB': [float('nan'), float('nan')]})
        df2 = pd.DataFrame({'Serial_No': [1, 2],
                            'ParamA': [1.1, 2.1],
                            'ParamB': [float('nan'), float('nan')]})
        # 默认 ignore_no_data=True → ParamB（无配对数据）被过滤
        resp = self._call_export({1: df1, 2: df2})
        ws = load_workbook(io.BytesIO(self._body(resp)))['测试值对比']
        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws['A4'].value, 'ParamA')
        # ignore_no_data=False → ParamB 保留（compared=0）
        resp2 = self._call_export({1: df1, 2: df2}, body={'ignore_no_data': False})
        ws2 = load_workbook(io.BytesIO(self._body(resp2)))['测试值对比']
        self.assertEqual(ws2['A5'].value, 'ParamB')
        self.assertEqual(ws2['M5'].value, 'PASS')

    def test_export_limits_only_when_no_common_serials(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [99], 'ParamA': [5.0]})
        resp = self._call_export({1: df1, 2: df2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        wb = load_workbook(io.BytesIO(body))
        ws_l = wb['Limit对比']
        self.assertEqual(ws_l['B3'].value, 0.5)
        # 无公共序列 → 数据 Sheet 无序列块：A 参数列 + B-D Limit/Unit + E 列 Comment
        ws_d = wb['测试值对比']
        self.assertEqual(ws_d['E2'].value, 'Comment')
        self.assertEqual(ws_d['E3'].value, 'Comment')
        self.assertEqual(ws_d.max_column, 5)
        self.assertEqual(ws_d['A4'].value, 'ParamA')
        self.assertEqual(ws_d['B4'].value, 0.5)
        self.assertEqual(ws_d['E4'].value, 'PASS')

    def test_export_no_common_params_returns_400(self):
        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'Other': [1.0, 2.0]})
        resp = self._call_export({1: df1, 2: df2})
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.json()['error'], 'no_common_params')


class FileCorrelationServiceTests(SimpleTestCase):
    """compute_file_correlation 纯计算单测（无 mock，直接调服务）。"""

    @staticmethod
    def _frames():
        df1 = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, 2.0, 3.0],
            'ParamB': [10.0, 20.0, 30.0],
        })
        df2 = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.1, 2.1, 3.1],
            'ParamB': [10.0, 21.0, 29.0],
        })
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0.5', 'ParamB': '-'},
                'maxs': {'ParamA': '2.0', 'ParamB': '40'},
                'units': {'ParamA': 'V', 'ParamB': 'nA'}}
        return df1, df2, meta

    def test_vectorized_matches_hand_computation(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1, df2, meta = self._frames()
        r = compute_file_correlation(df1, meta, df2, meta, FileCorrelationConfig())
        self.assertEqual(r['serials'], [1, 2, 3])
        self.assertFalse(r['limits_only'])
        self.assertFalse(r['truncated'])
        # 参数按文件A列顺序（序列列不参与）
        self.assertEqual(r['params'], ['ParamA', 'ParamB'])

        by_param = {row['param']: row for row in r['rows']}
        pa = by_param['ParamA']
        # 有符号 Δ/ATE：10% / 5% / 3.33%
        self.assertEqual([c['delta'] for c in pa['cells']], [0.1, 0.1, 0.1])
        self.assertEqual([round(c['diff_pct'], 2) for c in pa['cells']],
                         [10.0, 5.0, 3.33])
        self.assertEqual([c['fail'] for c in pa['cells']], [True, True, True])
        self.assertEqual(pa['compared'], 3)
        self.assertEqual(pa['fail_count'], 3)
        self.assertEqual(pa['max_diff'], 10.0)

        pb = by_param['ParamB']
        self.assertEqual([round(c['diff_pct'], 2) for c in pb['cells']],
                         [0.0, 5.0, -3.33])
        self.assertEqual([c['fail'] for c in pb['cells']], [False, True, True])
        self.assertEqual(pb['pass_rate'], 33.33)

        # totals 汇总
        self.assertEqual(r['totals']['paired_cells'], 6)
        self.assertEqual(r['totals']['fail_cells'], 5)
        self.assertEqual(r['totals']['overall_pass_rate'], 16.67)

    def test_serial_cap_takes_first_n_ascending(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1 = pd.DataFrame({'Serial_No': [5, 3, 1, 4, 2], 'ParamA': [1.0] * 5})
        df2 = pd.DataFrame({'Serial_No': [1, 2, 3, 4, 5], 'ParamA': [1.0] * 5})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0'}, 'maxs': {'ParamA': '5'}, 'units': {}}
        r = compute_file_correlation(df1, meta, df2, meta,
                                     FileCorrelationConfig(max_serials=2))
        self.assertTrue(r['truncated'])
        self.assertEqual(r['serials'], [1, 2])
        self.assertEqual(len(r['rows'][0]['cells']), 2)

    def test_limit_parsing_sentinels(self):
        from apps.analysis.services.file_correlation import _parse_limit

        for raw in ('·', '-', '—', '', '  ', 'n/a', 'N/A', 'min', 'max', None):
            self.assertIsNone(_parse_limit(raw), raw)
        self.assertEqual(_parse_limit('0.5'), 0.5)
        self.assertEqual(_parse_limit('"1.2"'), 1.2)
        self.assertEqual(_parse_limit('abc'), None)

    def test_ignore_no_limit_filters_params_without_limits(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1, df2, meta = self._frames()
        # 加入两侧都无 limit 的 ParamC（'-'/'-'）→ 默认被 ignore_no_limit 过滤
        df1['ParamC'] = [0.1, 0.2, 0.3]
        df2['ParamC'] = [0.1, 0.2, 0.3]
        meta['mins']['ParamC'] = '-'
        meta['maxs']['ParamC'] = '-'
        r = compute_file_correlation(df1, meta, df2, meta, FileCorrelationConfig())
        self.assertEqual(r['params'], ['ParamA', 'ParamB'])
        # 关闭后 ParamC 参与（数据相同 → 无超差）
        r2 = compute_file_correlation(df1, meta, df2, meta,
                                      FileCorrelationConfig(ignore_no_limit=False))
        self.assertEqual(r2['params'], ['ParamA', 'ParamB', 'ParamC'])

    def test_missing_limit_on_one_side_fails_both_rules(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1, df2, _ = self._frames()
        meta_a = {'mins': {'ParamA': '0.5'}, 'maxs': {'ParamA': '2.0'}, 'units': {}}
        meta_b = {'mins': {}, 'maxs': {'ParamA': '2.0'}, 'units': {}}
        for rule in ('zero', 'wider'):
            r = compute_file_correlation(df1, meta_a, df2, meta_b,
                                         FileCorrelationConfig(diff_rule=rule,
                                                              ignore_no_limit=False))
            row = r['rows'][0]
            self.assertTrue(row['lsl_fail'], rule)
            self.assertFalse(row['usl_fail'], rule)
            self.assertIsNone(row['lsl_diff'], rule)

    def test_diff_rule_zero_and_wider(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1, df2, _ = self._frames()
        meta_a = {'mins': {'ParamA': '0.5'}, 'maxs': {'ParamA': '2.0'}, 'units': {}}
        # B 更宽（0.4/2.5）
        meta_wide = {'mins': {'ParamA': '0.4'}, 'maxs': {'ParamA': '2.5'}, 'units': {}}
        # B 更紧（0.6/1.9）
        meta_tight = {'mins': {'ParamA': '0.6'}, 'maxs': {'ParamA': '1.9'}, 'units': {}}

        r_zero = compute_file_correlation(df1, meta_a, df2, meta_wide,
                                          FileCorrelationConfig(diff_rule='zero'))
        self.assertEqual(r_zero['rows'][0]['lsl_diff'], -0.1)
        self.assertTrue(r_zero['rows'][0]['lsl_fail'])   # diff ≠ 0 → fail
        self.assertTrue(r_zero['rows'][0]['usl_fail'])   # 0.5 ≠ 0 → fail

        r_wider = compute_file_correlation(df1, meta_a, df2, meta_wide,
                                           FileCorrelationConfig(diff_rule='wider'))
        self.assertFalse(r_wider['rows'][0]['lsl_fail'])  # B 更宽 → pass
        self.assertFalse(r_wider['rows'][0]['usl_fail'])

        r_tight = compute_file_correlation(df1, meta_a, df2, meta_tight,
                                           FileCorrelationConfig(diff_rule='wider'))
        self.assertTrue(r_tight['rows'][0]['lsl_fail'])
        self.assertTrue(r_tight['rows'][0]['usl_fail'])

    def test_zero_ate_pair_is_uncomputable(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [0.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.05]})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0'}, 'maxs': {'ParamA': '5'}, 'units': {}}
        r = compute_file_correlation(df1, meta, df2, meta, FileCorrelationConfig())
        cells = r['rows'][0]['cells']
        # ATE=0 的对无法计算 %Diff → 不计入对比、不 fail
        self.assertIsNone(cells[0]['diff_pct'])
        self.assertFalse(cells[0]['fail'])
        self.assertEqual(r['rows'][0]['compared'], 1)

    def test_limits_only_mode_when_no_common_serials(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [99], 'ParamA': [5.0]})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0.5'}, 'maxs': {'ParamA': '2.0'}, 'units': {}}
        r = compute_file_correlation(df1, meta, df2, meta, FileCorrelationConfig())
        self.assertTrue(r['limits_only'])
        self.assertEqual(r['serials'], [])
        self.assertEqual(r['rows'][0]['cells'], [])
        self.assertEqual(r['rows'][0]['compared'], 0)
        # limit 列仍完整可对比
        self.assertEqual(r['rows'][0]['lsl_a'], 0.5)
        self.assertFalse(r['rows'][0]['lsl_fail'])

    def test_no_common_params_raises(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig, NoCommonParamsError)

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'Other': [1.0, 2.0]})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        with self.assertRaises(NoCommonParamsError):
            compute_file_correlation(df1, {}, df2, {}, FileCorrelationConfig())

    def test_duplicate_serial_rows_take_first_occurrence(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1 = pd.DataFrame({'Serial_No': [1, 1, 2], 'ParamA': [1.0, 99.0, 3.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.05, 3.0]})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0'}, 'maxs': {'ParamA': '5'}, 'units': {}}
        r = compute_file_correlation(df1, meta, df2, meta, FileCorrelationConfig())
        # 序列 1 在文件A 出现两次 → 取 first（1.0），diff_pct = 5%
        self.assertEqual(r['rows'][0]['cells'][0]['ate'], 1.0)
        self.assertAlmostEqual(r['rows'][0]['cells'][0]['diff_pct'], 5.0)

    def test_explicit_serials_selected_in_request_order(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1 = pd.DataFrame({'Serial_No': [5, 3, 1, 4, 2], 'ParamA': [1.0] * 5})
        df2 = pd.DataFrame({'Serial_No': [1, 2, 3, 4, 5], 'ParamA': [1.0] * 5})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0'}, 'maxs': {'ParamA': '5'}, 'units': {}}
        r = compute_file_correlation(df1, meta, df2, meta,
                                     FileCorrelationConfig(serials=[3, 1]))
        # 显式选择：保持请求顺序；无「截断」语义
        self.assertEqual(r['serials'], [3, 1])
        self.assertFalse(r['truncated'])
        self.assertEqual(r['totals']['serials'], 2)
        self.assertEqual([c['serial'] for c in r['rows'][0]['cells']], [3, 1])

    def test_explicit_serials_filters_invalid_and_dedups(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        meta = {'mins': {'ParamA': '0'}, 'maxs': {'ParamA': '5'}, 'units': {}}
        r = compute_file_correlation(df1, meta, df2, meta,
                                     FileCorrelationConfig(serials=[2, 99, 2, 1]))
        self.assertEqual(r['serials'], [2, 1])

    def test_explicit_serials_empty_limits_only(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1, df2, meta = self._frames()
        r = compute_file_correlation(df1, meta, df2, meta,
                                     FileCorrelationConfig(serials=[]))
        self.assertTrue(r['limits_only'])
        self.assertEqual(r['serials'], [])
        self.assertFalse(r['truncated'])
        self.assertEqual(r['rows'][0]['cells'], [])
        # limit 列仍完整可对比
        self.assertEqual(r['rows'][0]['lsl_a'], 0.5)

    def test_explicit_serials_overrides_max_serial_fallback(self):
        from apps.analysis.services.file_correlation import (
            compute_file_correlation, FileCorrelationConfig)

        df1, df2, meta = self._frames()
        r = compute_file_correlation(df1, meta, df2, meta,
                                     FileCorrelationConfig(serials=[1, 2, 3],
                                                           max_serials=1))
        self.assertEqual(r['serials'], [1, 2, 3])
        self.assertFalse(r['truncated'])
        self.assertEqual(r['totals']['serials'], 3)

    def test_list_common_serials_ascending(self):
        from apps.analysis.services.file_correlation import list_common_serials

        df1 = pd.DataFrame({'Serial_No': [3, 1, 2], 'ParamA': [1.0, 1.0, 1.0]})
        df2 = pd.DataFrame({'Serial_No': [2, 1, 99], 'ParamA': [1.0, 1.0, 1.0]})
        for d in (df1, df2):
            d['__serial__'] = pd.to_numeric(d['Serial_No'], errors='coerce')
        self.assertEqual(list_common_serials(df1, df2), [1, 2])


class FileCorrelationSerialsApiTests(TestCase):
    """file_correlation_serials 端点：公共序列列表（序列勾选器数据源）。"""

    def setUp(self):
        from django.contrib.auth import get_user_model
        self.user = get_user_model().objects.create_user(
            username='fc_serials_test', password='x')

    def _call(self, frames, body=None):
        from rest_framework.test import APIClient
        from apps.analysis.views import analysis_views

        orig_404 = analysis_views.get_object_or_404
        orig_load = analysis_views.get_cached_parsed_file
        orig_serial = analysis_views.get_serial_column

        analysis_views.get_object_or_404 = (
            lambda model, *a, **k: types.SimpleNamespace(
                id=k['pk'], filename=f'FILE{k["pk"]}.csv',
                format_type='CTA8290D'))
        analysis_views.get_cached_parsed_file = (
            lambda fid, owner_id, datafile=None: (
                frames[int(fid)], {'format': 'CTA8290D'}, 'CTA8290D'))
        analysis_views.get_serial_column = lambda df: 'Serial_No'
        self.addCleanup(lambda: setattr(analysis_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(analysis_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(analysis_views, 'get_serial_column', orig_serial))

        client = APIClient()
        client.force_authenticate(user=self.user)
        return client.post('/api/v1/analysis/file_correlation_serials/', {
            'file1_id': 1, 'file2_id': 2, **(body or {}),
        }, format='json')

    def test_returns_ascending_common_serials(self):
        df1 = pd.DataFrame({'Serial_No': [3, 1, 2], 'ParamA': [1.0, 1.0, 1.0]})
        df2 = pd.DataFrame({'Serial_No': [2, 1, 99], 'ParamA': [1.0, 1.0, 1.0]})
        resp = self._call({1: df1, 2: df2})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data, {'serials': [1, 2], 'total': 2})

    def test_no_common_serials_empty_list(self):
        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [99], 'ParamA': [5.0]})
        resp = self._call({1: df1, 2: df2})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.data, {'serials': [], 'total': 0})

    def test_missing_ids_returns_400(self):
        from rest_framework.test import APIClient
        client = APIClient()
        client.force_authenticate(user=self.user)
        resp = client.post('/api/v1/analysis/file_correlation_serials/', {}, format='json')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.json()['error'], 'need_two_files')


class HistogramSigmaFieldTests(ChartConfigFilterTests):
    """histogram 响应必须同时提供全量与裁剪口径的 σ 字段。

    回归：前端曾在本地用 displayMean/displayStd 重算 σ（与图表标记线的
    全量 σ 矛盾）；后端现统一返回 sigma4 + filtered_sigma3/4/6，前端
    卡片与标记线消费同一组值。
    """

    def test_compute_path_returns_raw_and_filtered_sigma_fields(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'params': ['ParamOutlierLow']})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        r = resp.data['results']['ParamOutlierLow']

        # 全量 σ 字段齐全（含新增的 4σ）
        self.assertIn('sigma3_min', r)
        self.assertIn('sigma4_min', r)
        self.assertIn('sigma6_min', r)
        # mean/std 先各自 round(6) 再组合，与响应值有 ~1e-6 舍入差，用 places=4
        self.assertAlmostEqual(r['sigma4_min'], r['mean'] - 4 * r['std'], places=4)
        self.assertAlmostEqual(r['sigma4_max'], r['mean'] + 4 * r['std'], places=4)

        # ParamOutlierLow 含异常值 → filtered 统计存在
        self.assertTrue(r['outlier_info']['has_outliers'])
        self.assertIsNotNone(r['filtered_mean'])
        self.assertIsNotNone(r['filtered_std'])
        self.assertGreater(r['filtered_std'], 0)

        # 裁剪口径 σ = filtered_mean ± k*filtered_std（与 filtered_mean/std 同源）
        self.assertAlmostEqual(r['filtered_sigma3_min'],
                               r['filtered_mean'] - 3 * r['filtered_std'], places=6)
        self.assertAlmostEqual(r['filtered_sigma3_max'],
                               r['filtered_mean'] + 3 * r['filtered_std'], places=6)
        self.assertAlmostEqual(r['filtered_sigma4_min'],
                               r['filtered_mean'] - 4 * r['filtered_std'], places=6)
        self.assertAlmostEqual(r['filtered_sigma6_max'],
                               r['filtered_mean'] + 6 * r['filtered_std'], places=6)

    def test_normal_curves_returned_raw_and_filtered(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'params': ['ParamOutlierLow']})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        r = resp.data['results']['ParamOutlierLow']

        # 全量曲线存在（raw mean/std）
        self.assertIsNotNone(r['normal_curve'])
        self.assertGreater(len(r['normal_curve']), 1)
        # 有异常值 → 裁剪口径曲线存在
        self.assertIsNotNone(r['filtered_normal_curve'])
        self.assertGreater(len(r['filtered_normal_curve']), 1)
        # 裁剪曲线峰值在 filtered_mean 附近（用 filtered mean/std 计算）
        peak = max(r['filtered_normal_curve'], key=lambda p: p[1])
        self.assertAlmostEqual(peak[0], r['filtered_mean'], delta=0.05)
        # 无异常值参数：filtered 曲线为 None
        request2 = self._post({'file_id': 1, 'params': ['Param0']})
        resp2 = AnalysisViewSet.as_view({'post': 'histogram'})(request2)
        resp2.render()
        r2 = resp2.data['results']['Param0']
        self.assertIsNotNone(r2['normal_curve'])
        self.assertIsNone(r2['filtered_normal_curve'])


    def test_filtered_cpk_level_and_color_returned_with_filtered_cpk(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'params': ['ParamOutlierLow']})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        r = resp.data['results']['ParamOutlierLow']
        self.assertIsNotNone(r['filtered_cpk'])
        # 评级标签与颜色与 CPK 卡显示格式一致（此前前端显示 '(filtered)' 丢评级）。
        # ParamOutlierLow：含 14.4 异常值 → 全量 CPK 低（D级/red）；剔除后 ≈34 → A级/green
        self.assertIsNotNone(r['filtered_cpk_level'])
        self.assertIsNotNone(r['filtered_cpk_color'])
        self.assertEqual(r['cpk_color'], 'red')
        self.assertEqual(r['filtered_cpk_color'], 'green')
        self.assertNotEqual(r['filtered_cpk_level'], r['cpk_level'])

    def test_no_outlier_param_returns_null_filtered_sigma(self):
        from apps.analysis.views import AnalysisViewSet
        request = self._post({'file_id': 1, 'params': ['Param0']})
        resp = AnalysisViewSet.as_view({'post': 'histogram'})(request)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        r = resp.data['results']['Param0']
        self.assertFalse(r['outlier_info']['has_outliers'])
        self.assertIsNone(r['filtered_sigma3_min'])
        self.assertIsNone(r['filtered_sigma6_max'])


class ParamTrendPerFileLimitsTests(SimpleTestCase):
    """compute_param_trend 必须按文件独立解析规格限。

    回归：旧实现把第一个文件的 lsl/usl 复用到所有文件——不同批次/
    程序版本的规格不同时，后续文件的 CPK 数学上错误。
    """

    @staticmethod
    def _file(df, fid, mins, maxs):
        return {
            'df': df,
            'metadata': {'mins': mins, 'maxs': maxs, 'units': {}},
            'file_id': fid,
            'filename': f'f{fid}.csv',
            'timestamp': '',
        }

    def test_per_file_limits_drive_per_file_cpk(self):
        from apps.analysis.services.statistics.trends import compute_param_trend
        df = pd.DataFrame({'Param': [10.0, 10.2, 10.4]})
        files = [
            self._file(df.copy(), 1, {'Param': '9.0'}, {'Param': '11.0'}),
            self._file(df.copy(), 2, {'Param': '5.0'}, {'Param': '15.0'}),
        ]
        out = compute_param_trend(files, 'Param')

        self.assertEqual(out['trend_data'][0]['lsl'], 9.0)
        self.assertEqual(out['trend_data'][1]['lsl'], 5.0)
        # 同分布数据：限值越宽 → CPK 越大（文件 2 用自身限值而非文件 1 的）
        self.assertGreater(out['trend_data'][1]['cpk'], out['trend_data'][0]['cpk'])
        self.assertGreater(out['trend_data'][0]['cpk'], 0.0)
        # 响应级 limits 取首个完整对
        self.assertEqual(out['limits']['lsl'], 9.0)

    def test_file_without_limits_zero_cpk_and_later_file_keeps_own(self):
        from apps.analysis.services.statistics.trends import compute_param_trend
        df = pd.DataFrame({'Param': [10.0, 10.2, 10.4]})
        files = [
            self._file(df.copy(), 1, {}, {}),
            self._file(df.copy(), 2, {'Param': '9.0'}, {'Param': '11.0'}),
        ]
        out = compute_param_trend(files, 'Param')

        # 缺失限值（parse 回退 0.0 会算出负 CPK）→ lsl=None、cpk=0
        self.assertIsNone(out['trend_data'][0]['lsl'])
        self.assertEqual(out['trend_data'][0]['cpk'], 0.0)
        self.assertEqual(out['trend_data'][1]['lsl'], 9.0)
        self.assertGreater(out['trend_data'][1]['cpk'], 0.0)
        # 首个完整对来自文件 2
        self.assertEqual(out['limits']['lsl'], 9.0)


class NormalPdfCurveTests(SimpleTestCase):
    """normal_pdf_curve：高斯 PDF 公式单一来源（前端/导出/响应共用）。"""

    def test_formula_values(self):
        import math
        from apps.analysis.services.statistics import normal_pdf_curve
        # 奇数采样点 → 中心点恰为均值 0，可直接断言公式值
        curve = normal_pdf_curve(0.0, 1.0, -3.0, 3.0, n_points=201)
        self.assertEqual(len(curve), 201)
        center = curve[100]
        self.assertAlmostEqual(center[0], 0.0, places=5)
        self.assertAlmostEqual(center[1], 1 / math.sqrt(2 * math.pi), places=5)
        # 对称性
        self.assertAlmostEqual(curve[99][1], curve[101][1], places=6)

    def test_zero_std_returns_none(self):
        from apps.analysis.services.statistics import normal_pdf_curve
        self.assertIsNone(normal_pdf_curve(1.0, 0.0, 0.0, 2.0))
        self.assertIsNone(normal_pdf_curve(1.0, -1.0, 0.0, 2.0))

    def test_sampling_range(self):
        from apps.analysis.services.statistics import normal_pdf_curve
        curve = normal_pdf_curve(5.0, 2.0, 1.0, 9.0, n_points=100)
        self.assertEqual(curve[0][0], 1.0)
        self.assertEqual(curve[-1][0], 9.0)
        # 采样点单调
        xs = [c[0] for c in curve]
        self.assertEqual(xs, sorted(xs))


class ComputeCpkSingleSidedTests(SimpleTestCase):
    """compute_cpk 单边规格限：cp/pp 必须为 None（双侧才可定义），cpk/ppk 单侧可算。

    回归：缺失侧以 -inf/+inf 传入时 cp = inf，破坏 JSON 序列化且语义错误。
    """

    def test_lower_only_limit(self):
        from apps.analysis.services.statistics import compute_cpk
        result = compute_cpk(10.0, 1.0, float('-inf'), 12.0)
        self.assertIsNone(result['cp'])
        self.assertIsNone(result['pp'])
        self.assertEqual(result['cp_level'], 'N/A')
        self.assertEqual(result['cp_color'], 'gray')
        # 单侧能力 = (12-10)/(3*1) = 0.667
        self.assertAlmostEqual(result['cpk'], 0.6667, places=3)
        self.assertEqual(result['ppk'], result['cpk'])

    def test_upper_only_limit(self):
        from apps.analysis.services.statistics import compute_cpk
        result = compute_cpk(10.0, 1.0, 8.0, float('inf'))
        self.assertIsNone(result['cp'])
        self.assertAlmostEqual(result['cpk'], 0.6667, places=3)

    def test_both_limits_still_return_cp(self):
        from apps.analysis.services.statistics import compute_cpk
        result = compute_cpk(10.0, 1.0, 8.0, 12.0)
        self.assertAlmostEqual(result['cp'], 0.6667, places=3)
        self.assertIsNotNone(result['pp'])


class FilterFiniteTests(SimpleTestCase):
    """filter_finite：NaN/±inf 一步移除，等价于 dropna + inf 过滤。"""

    def test_removes_nan_and_inf(self):
        from apps.analysis.services.statistics import filter_finite
        series = pd.Series([1.0, float('nan'), 2.0, float('inf'), 3.0, float('-inf')])
        out = filter_finite(series)
        self.assertEqual(out.tolist(), [1.0, 2.0, 3.0])

    def test_index_preserved(self):
        from apps.analysis.services.statistics import filter_finite
        series = pd.Series([1.0, float('nan'), 2.0], index=['a', 'b', 'c'])
        out = filter_finite(series)
        self.assertEqual(out.index.tolist(), ['a', 'c'])

    def test_non_numeric_coerced_to_nan(self):
        from apps.analysis.services.statistics import filter_finite
        series = pd.Series([1.0, 'x', 2.0])
        out = filter_finite(series)
        self.assertEqual(out.tolist(), [1.0, 2.0])
