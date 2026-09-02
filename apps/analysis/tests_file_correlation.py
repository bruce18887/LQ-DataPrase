"""文件相关性端点的缓存隔离、NaN JSON 与导出回归测试。
（自 2473 行的 tests.py 按主题拆出，用例逐字搬迁。）
"""
import pandas as pd
import types
import math
from django.test import SimpleTestCase
from django.test import TestCase


class FileCorrelationCacheIsolationTests(SimpleTestCase):
    """file_correlation 端点不得污染 LRU 解析缓存。

    回归：file_correlation 曾对 ``get_cached_parsed_file`` 返回的缓存
    DataFrame 原地添加 ``__serial__`` 辅助列，违反缓存"只读"不变量，
    导致后续 browse/histogram/export 请求同一文件时出现幽灵列。
    修复：写列前 ``df = df.copy()``。
    """

    def test_shared_cache_df_not_mutated_by_file_correlation(self):
        import types
        from rest_framework.test import APIRequestFactory, force_authenticate
        # 必须 patch 实际消费模块（R6③）：三端点已迁到 file_correlation_views
        from apps.analysis.views import file_correlation_views as fc_views

        shared_df = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, 2.0, 3.0],
            'ParamB': [4.0, 5.0, 6.0],
        })
        metadata = {'format': 'CTA8290D', 'mins': {}, 'maxs': {}, 'units': {}}

        orig_404 = fc_views.get_object_or_404
        orig_load = fc_views.get_cached_parsed_file
        orig_serial = fc_views.get_serial_column

        def fake_404(model, *args, **kwargs):
            return types.SimpleNamespace(id=1, filename='fake.csv',
                                         format_type='CTA8290D')

        def fake_load(fid, owner_id, datafile=None):
            # 两个文件返回同一个缓存对象 —— 模拟 LRU 缓存真实行为
            return shared_df, metadata, 'CTA8290D'

        fc_views.get_object_or_404 = fake_404
        fc_views.get_cached_parsed_file = fake_load
        fc_views.get_serial_column = lambda df: 'Serial_No'
        self.addCleanup(lambda: setattr(fc_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(fc_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(fc_views, 'get_serial_column', orig_serial))

        factory = APIRequestFactory()
        request = factory.post('/api/v1/analysis/file_correlation/', {
            'file1_id': 1, 'file2_id': 2,
            # mock 的 metadata 无 limits —— 关掉 ignore_no_limit 让参数参与对比
            'ignore_no_limit': False,
        }, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))

        from apps.analysis.views import AnalysisViewSet
        view = AnalysisViewSet.as_view({'post': 'file_correlation'})
        response = view(request)
        response.render()

        self.assertEqual(response.status_code, 200, response.content)
        # 缓存对象不允许被端点改写
        self.assertNotIn('__serial__', shared_df.columns)
        # 端点自身逻辑仍可用（3 个共同序列号被匹配；序列列不参与参数对比，
        # 只比较 ParamA/ParamB）
        self.assertEqual(response.data['serials'], [1, 2, 3])
        self.assertEqual(response.data['params'], ['ParamA', 'ParamB'])


class FileCorrelationNanJsonTests(SimpleTestCase):
    """file_correlation 对含 NaN 的参数列必须返回合法 JSON（回归 500）。

    回归：偏差计算把 NaN 单元格直接算进 diff_pct，nan 进入 summary 后
    DRF JSON 序列化抛 ``ValueError: Out of range float values are not
    JSON compliant: nan`` → 整个端点 500。修复：任一侧非有限值即跳过
    该序列号；响应再经 clean_data 兜底。
    """

    def _call_endpoint(self, shared_df, serial='Serial_No', extra=None):
        from rest_framework.test import APIRequestFactory, force_authenticate
        # 必须 patch 实际消费模块（R6③）：三端点已迁到 file_correlation_views
        from apps.analysis.views import AnalysisViewSet
        from apps.analysis.views import file_correlation_views as fc_views

        metadata = {'format': 'CTA8290D', 'mins': {}, 'maxs': {}, 'units': {}}

        orig_404 = fc_views.get_object_or_404
        orig_load = fc_views.get_cached_parsed_file
        orig_serial = fc_views.get_serial_column

        fc_views.get_object_or_404 = (
            lambda model, *a, **k: types.SimpleNamespace(
                id=1, filename='fake.csv', format_type='CTA8290D'))
        fc_views.get_cached_parsed_file = (
            lambda fid, owner_id, datafile=None: (shared_df, metadata, 'CTA8290D'))
        fc_views.get_serial_column = lambda df: serial
        self.addCleanup(lambda: setattr(fc_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(fc_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(fc_views, 'get_serial_column', orig_serial))

        factory = APIRequestFactory()
        request = factory.post('/api/v1/analysis/file_correlation/', {
            'file1_id': 1, 'file2_id': 2, 'threshold': 3.0,
            # mock 的 metadata 无 limits —— 关掉 ignore_no_limit 让参数参与对比
            'ignore_no_limit': False,
            **(extra or {}),
        }, format='json')
        force_authenticate(request, user=types.SimpleNamespace(
            pk=1, is_authenticated=True, is_active=True, is_anonymous=False,
            is_staff=False, is_superuser=False,
        ))

        view = AnalysisViewSet.as_view({'post': 'file_correlation'})
        response = view(request)
        response.render()  # 触发 JSON 序列化——无异常即证明无 nan/inf 泄漏
        return response

    def test_nan_param_values_do_not_break_json_serialization(self):
        shared_df = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, float('nan'), 3.0],
            'ParamB': [4.0, 5.0, 6.0],
        })
        response = self._call_endpoint(shared_df)
        self.assertEqual(response.status_code, 200, response.content)

        by_param = {r['param']: r for r in response.data['rows']}
        # 序列号 2 的 ParamA 为 NaN → 该对不参与对比；ParamB 全部有效
        self.assertEqual(by_param['ParamA']['compared'], 2)
        self.assertEqual(by_param['ParamB']['compared'], 3)
        for r in response.data['rows']:
            self.assertTrue(math.isfinite(r['max_diff']), r)
            self.assertTrue(math.isfinite(r['pass_rate']), r)
        self.assertEqual(response.data['serials'], [1, 2, 3])

    def test_all_nan_param_values_returns_empty_comparison(self):
        shared_df = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [float('nan'), float('nan'), float('nan')],
        })
        # ignore_no_data=False：全 NaN 参数保留在结果中（compared=0），
        # 而不是被「忽略无数据」过滤掉
        response = self._call_endpoint(shared_df, extra={'ignore_no_data': False})
        self.assertEqual(response.status_code, 200, response.content)
        by_param = {r['param']: r for r in response.data['rows']}
        self.assertEqual(by_param['ParamA']['compared'], 0)
        self.assertEqual(by_param['ParamA']['fail_count'], 0)
        self.assertEqual(by_param['ParamA']['max_diff'], 0)
        self.assertEqual(by_param['ParamA']['pass_rate'], 0)


class FileCorrelationExportTests(TestCase):
    """file_correlation_export 输出双 Sheet xlsx。

    规格（2026-08：Limit 与测试值拆分为两个 Sheet）：
    - workbook 只含两个 Sheet：'Limit对比' + '测试值对比'（excelize 默认
      Sheet1 已删除，顺序即创建顺序）。
    - Limit 对比 Sheet：标题行（跨 A:I 合并）| 单行表头 Parameters | LSL A |
      USL A | LSL B | USL B | LSL Diff | USL Diff | Unit | 判定 | 每测试项
      一行；Diff 按 diff_rule 标红，判定列 PASS/FAIL（仅 limit 差异）。
    - 测试值对比 Sheet：模板布局——标题行（跨全表合并）| 两行表头
      A2/A3 'Parameters' + 'Limit / Unit (Data A)' 组（LSL A/USL A/Unit）
      + 每序列组标题（E2:H2…）+ 末列 Comment；数据行公式 =Bench-ATE /
      =Delta/ATE（Δ/ATE 口径），%Diff 数字格式 0.00%；|%Diff| > threshold
      → Delta/%Diff 标红（静态判定与 JSON 端点一致）；单侧有值只写该侧；
      Comment 只含超差摘要（N 超差 / PASS）。
    - 无公共序列 → limits_only：数据 Sheet 无序列块（仅 Parameters +
      Comment）。
    - 无公共测试项 → 400 no_common_params。
    """

    def setUp(self):
        from django.contrib.auth import get_user_model
        self.user = get_user_model().objects.create_user(
            username='fc_export_test', password='x')

    def _call_export(self, frames, body=None, metas=None):
        from rest_framework.test import APIClient
        # 必须 patch 实际消费模块（R6③）：三端点已迁到 file_correlation_views
        from apps.analysis.views import file_correlation_views as fc_views

        body = body or {}
        base_meta = {'format': 'CTA8290D',
                     'mins': {'ParamA': '0.5', 'ParamB': '10'},
                     'maxs': {'ParamA': '2.0', 'ParamB': '40'},
                     'units': {'ParamA': 'V', 'ParamB': 'nA'}}
        metas = metas or {}
        orig_404 = fc_views.get_object_or_404
        orig_load = fc_views.get_cached_parsed_file
        orig_serial = fc_views.get_serial_column

        fc_views.get_object_or_404 = (
            lambda model, *a, **k: types.SimpleNamespace(
                id=k['pk'], filename=f'FILE{k["pk"]}.csv',
                format_type='CTA8290D'))
        fc_views.get_cached_parsed_file = (
            lambda fid, owner_id, datafile=None: (
                frames[int(fid)], metas.get(int(fid), base_meta), 'CTA8290D'))
        fc_views.get_serial_column = lambda df: 'Serial_No'
        self.addCleanup(lambda: setattr(fc_views, 'get_object_or_404', orig_404))
        self.addCleanup(lambda: setattr(fc_views, 'get_cached_parsed_file', orig_load))
        self.addCleanup(lambda: setattr(fc_views, 'get_serial_column', orig_serial))

        client = APIClient()
        client.force_authenticate(user=self.user)
        return client.post('/api/v1/analysis/file_correlation_export/', {
            'file1_id': 1, 'file2_id': 2, **body,
        }, format='json')

    @staticmethod
    def _body(resp):
        """FileResponse is streaming — collect bytes for openpyxl / asserts."""
        return b''.join(resp.streaming_content)

    def test_export_two_sheets_template_layout_formulas_and_red_highlight(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.0, 2.0, 3.0],
            'ParamB': [10.0, 20.0, 30.0],
        })
        df2 = pd.DataFrame({
            'Serial_No': [1, 2, 3],
            'ParamA': [1.1, 2.1, 3.1],
            'ParamB': [10.0, 21.0, 29.0],
        })
        resp = self._call_export({1: df1, 2: df2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        self.assertIn('spreadsheetml', resp['Content-Type'])
        self.assertIn('FILE1_vs_FILE2_correlation.xlsx', resp['Content-Disposition'])

        wb = load_workbook(io.BytesIO(body))
        # 只含两个 Sheet（默认 Sheet1 已删除），顺序 = 创建顺序
        self.assertEqual(wb.sheetnames, ['Limit对比', '测试值对比'])
        ws_l = wb['Limit对比']
        ws_d = wb['测试值对比']

        # ── Limit 对比 Sheet：标题 + 单行表头 + 每测试项一行 ──
        self.assertEqual(ws_l['A1'].value, 'Data A VS Data B（Limit 对比）')
        self.assertIn('A1:I1', {str(r) for r in ws_l.merged_cells.ranges})
        for cell, h in [('A2', 'Parameters'), ('B2', 'LSL A'), ('C2', 'USL A'),
                        ('D2', 'LSL B'), ('E2', 'USL B'), ('F2', 'LSL Diff'),
                        ('G2', 'USL Diff'), ('H2', 'Unit'), ('I2', '判定')]:
            self.assertEqual(ws_l[cell].value, h)
        # 数据行（文件1 列顺序；diff 全 0 → 判定 PASS）
        self.assertEqual(ws_l['A3'].value, 'ParamA')
        self.assertEqual(ws_l['B3'].value, 0.5)
        self.assertEqual(ws_l['C3'].value, 2.0)
        self.assertEqual(ws_l['D3'].value, 0.5)
        self.assertEqual(ws_l['E3'].value, 2.0)
        self.assertEqual(ws_l['F3'].value, 0.0)
        self.assertEqual(ws_l['G3'].value, 0.0)
        self.assertEqual(ws_l['H3'].value, 'V')
        self.assertEqual(ws_l['I3'].value, 'PASS')
        self.assertEqual(ws_l['A4'].value, 'ParamB')
        self.assertEqual(ws_l['I4'].value, 'PASS')
        self.assertEqual(ws_l.max_row, 4)
        self.assertEqual(ws_l.max_column, 9)

        # ── 测试值对比 Sheet：模板布局（A 参数列 + Limit/Unit(Data A) 组，
        #    E 起序列块）── 3 序列 → 块 E..P，末列 Q（Comment）
        self.assertEqual(ws_d['A1'].value, 'Data A VS Data B（测试值对比）')
        merged = {str(r) for r in ws_d.merged_cells.ranges}
        self.assertIn('A1:Q1', merged)
        self.assertIn('A2:A3', merged)
        self.assertIn('B2:D2', merged)   # Limit / Unit (Data A) 组
        self.assertIn('E2:H2', merged)   # 序列 1 组标题
        self.assertIn('I2:L2', merged)   # 序列 2 组标题
        self.assertIn('M2:P2', merged)   # 序列 3 组标题

        self.assertEqual(ws_d['A2'].value, 'Parameters')   # A2:A3 合并（参数列跨两行）
        self.assertEqual(ws_d['B2'].value, 'Limit / Unit (Data A)')
        self.assertEqual(ws_d['E2'].value, 1)
        self.assertEqual(ws_d['I2'].value, 2)
        self.assertEqual(ws_d['M2'].value, 3)
        self.assertEqual(ws_d['Q2'].value, 'Comment')
        for cell, h in [('B3', 'LSL A'), ('C3', 'USL A'), ('D3', 'Unit'),
                        ('E3', 'ATE'), ('F3', 'Bench'), ('G3', 'Delta'), ('H3', '% Diff'),
                        ('I3', 'ATE'), ('J3', 'Bench'), ('K3', 'Delta'), ('L3', '% Diff'),
                        ('M3', 'ATE'), ('N3', 'Bench'), ('O3', 'Delta'), ('P3', '% Diff'),
                        ('Q3', 'Comment')]:
            self.assertEqual(ws_d[cell].value, h)

        # 数据行按文件1列顺序（ignore_no_limit 默认过滤无 limit 的序列列）
        self.assertEqual(ws_d['A4'].value, 'ParamA')
        self.assertEqual(ws_d['A5'].value, 'ParamB')
        self.assertEqual(ws_d.max_row, 5)
        self.assertEqual(ws_d.max_column, 17)   # Q

        # Data A 的 Limit/Unit（默认口径）
        self.assertEqual(ws_d['B4'].value, 0.5)
        self.assertEqual(ws_d['C4'].value, 2.0)
        self.assertEqual(ws_d['D4'].value, 'V')

        # 公式（Δ/ATE 口径 = Delta/ATE）+ 0.00% 数字格式；单侧有值只写该侧
        self.assertEqual(ws_d['E4'].value, 1.0)
        self.assertEqual(ws_d['F4'].value, 1.1)
        self.assertEqual(ws_d['G4'].value, '=F4-E4')
        self.assertEqual(ws_d['H4'].value, '=G4/E4')
        self.assertEqual(ws_d['H4'].number_format, '0.00%')

        # 序列1：ParamA |10%| > 3 → Delta/%Diff 标红；ParamB 0% 不标红
        self.assertEqual(
            ws_d['G4'].fill.start_color.rgb, ws_d['H4'].fill.start_color.rgb)
        self.assertNotEqual(
            ws_d['G4'].fill.start_color.rgb, ws_d['E4'].fill.start_color.rgb)
        self.assertEqual(
            ws_d['G5'].fill.start_color.rgb, ws_d['E5'].fill.start_color.rgb)

        # Comment 判定摘要（只含超差，Limit 差异在 Limit Sheet 判定列）
        self.assertEqual(ws_d['Q4'].value, '3 超差')   # ParamA: 10/5/3.33% 全超
        self.assertEqual(ws_d['Q5'].value, '2 超差')   # ParamB: 5/-3.33% 超

    def test_export_threshold_controls_red_highlight(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({
            'Serial_No': [1, 2],
            'ParamA': [1.0, 3.0],
        })
        df2 = pd.DataFrame({
            'Serial_No': [1, 2],
            'ParamA': [1.1, 3.1],
        })
        # 阈值放大到 100：ParamA 10%/3.33% 偏差不再标红
        resp = self._call_export({1: df1, 2: df2}, body={'threshold': 100.0})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['测试值对比']
        self.assertEqual(ws['A4'].value, 'ParamA')
        self.assertEqual(
            ws['G4'].fill.start_color.rgb, ws['E4'].fill.start_color.rgb)

    def test_export_single_side_value_writes_ate_only(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.1, float('nan')]})
        resp = self._call_export({1: df1, 2: df2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['测试值对比']
        # 序列2（I..L 块）：bench NaN → 只写 ATE，无公式（同模板单侧行）
        self.assertEqual(ws['I4'].value, 2.0)
        self.assertIsNone(ws['J4'].value)
        self.assertIsNone(ws['K4'].value)
        self.assertIsNone(ws['L4'].value)
        # 序列1（E..H 块）：两侧都有 → 公式
        self.assertEqual(ws['G4'].value, '=F4-E4')

    def test_export_diff_rule_zero_marks_different_limits_red(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        df2 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        # 文件 B 的 LSL 更宽（0.4 < 0.5）→ zero 规则下 Diff≠0 标红
        meta_b = {'format': 'CTA8290D',
                  'mins': {'ParamA': '0.4'}, 'maxs': {'ParamA': '2.0'},
                  'units': {'ParamA': 'V'}}
        resp = self._call_export({1: df1, 2: df2}, metas={2: meta_b})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['Limit对比']
        self.assertEqual(ws['F3'].value, -0.1)
        # LSL Diff 标红 + 判定 FAIL；USL Diff 相等不标红
        self.assertNotEqual(
            ws['F3'].fill.start_color.rgb, ws['G3'].fill.start_color.rgb)
        self.assertEqual(
            ws['F3'].fill.start_color.rgb, ws['I3'].fill.start_color.rgb)
        self.assertEqual(
            ws['G3'].fill.start_color.rgb, ws['H3'].fill.start_color.rgb)

    def test_export_diff_rule_wider_only_marks_tighter_limits_red(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        df2 = pd.DataFrame({'Serial_No': [1], 'ParamA': [1.0]})
        # B 更宽（0.4/2.5）→ wider 规则 pass（不标红）
        meta_wider = {'format': 'CTA8290D',
                      'mins': {'ParamA': '0.4'}, 'maxs': {'ParamA': '2.5'},
                      'units': {'ParamA': 'V'}}
        resp = self._call_export({1: df1, 2: df2}, metas={2: meta_wider},
                                 body={'diff_rule': 'wider'})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['Limit对比']
        self.assertEqual(ws['F3'].value, -0.1)
        self.assertEqual(
            ws['F3'].fill.start_color.rgb, ws['H3'].fill.start_color.rgb)
        self.assertEqual(
            ws['G3'].fill.start_color.rgb, ws['H3'].fill.start_color.rgb)

        # B 更紧（LSL 0.6 > 0.5）→ wider 规则标红 + 判定 FAIL
        meta_tight = {'format': 'CTA8290D',
                      'mins': {'ParamA': '0.6'}, 'maxs': {'ParamA': '2.0'},
                      'units': {'ParamA': 'V'}}
        resp2 = self._call_export({1: df1, 2: df2}, metas={2: meta_tight},
                                  body={'diff_rule': 'wider'})
        body2 = self._body(resp2)
        self.assertEqual(resp2.status_code, 200, body2[:500])
        ws2 = load_workbook(io.BytesIO(body2))['Limit对比']
        self.assertNotEqual(
            ws2['F3'].fill.start_color.rgb, ws2['H3'].fill.start_color.rgb)

    def test_export_max_serials_truncates_sequence_blocks(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2, 3, 4], 'ParamA': [1.0, 2.0, 3.0, 4.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2, 3, 4], 'ParamA': [1.0, 2.0, 3.0, 4.0]})
        resp = self._call_export({1: df1, 2: df2}, body={'max_serials': 2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        ws = load_workbook(io.BytesIO(body))['测试值对比']
        # 4 个公共序列只对比前 2 个 → 只有 2 组序列块；末列 M 为 Comment
        self.assertEqual(ws['E2'].value, 1)
        self.assertEqual(ws['I2'].value, 2)
        self.assertEqual(ws['M2'].value, 'Comment')
        self.assertEqual(ws.max_column, 13)   # M

    def test_export_ignore_no_data_filters_params(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2],
                            'ParamA': [1.0, 2.0],
                            'ParamB': [float('nan'), float('nan')]})
        df2 = pd.DataFrame({'Serial_No': [1, 2],
                            'ParamA': [1.1, 2.1],
                            'ParamB': [float('nan'), float('nan')]})
        # 默认 ignore_no_data=True → ParamB（无配对数据）被过滤
        resp = self._call_export({1: df1, 2: df2})
        ws = load_workbook(io.BytesIO(self._body(resp)))['测试值对比']
        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws['A4'].value, 'ParamA')
        # ignore_no_data=False → ParamB 保留（compared=0）
        resp2 = self._call_export({1: df1, 2: df2}, body={'ignore_no_data': False})
        ws2 = load_workbook(io.BytesIO(self._body(resp2)))['测试值对比']
        self.assertEqual(ws2['A5'].value, 'ParamB')
        self.assertEqual(ws2['M5'].value, 'PASS')

    def test_export_limits_only_when_no_common_serials(self):
        import io
        from openpyxl import load_workbook

        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [99], 'ParamA': [5.0]})
        resp = self._call_export({1: df1, 2: df2})
        body = self._body(resp)
        self.assertEqual(resp.status_code, 200, body[:500])
        wb = load_workbook(io.BytesIO(body))
        ws_l = wb['Limit对比']
        self.assertEqual(ws_l['B3'].value, 0.5)
        # 无公共序列 → 数据 Sheet 无序列块：A 参数列 + B-D Limit/Unit + E 列 Comment
        ws_d = wb['测试值对比']
        self.assertEqual(ws_d['E2'].value, 'Comment')
        self.assertEqual(ws_d['E3'].value, 'Comment')
        self.assertEqual(ws_d.max_column, 5)
        self.assertEqual(ws_d['A4'].value, 'ParamA')
        self.assertEqual(ws_d['B4'].value, 0.5)
        self.assertEqual(ws_d['E4'].value, 'PASS')

    def test_export_no_common_params_returns_400(self):
        df1 = pd.DataFrame({'Serial_No': [1, 2], 'ParamA': [1.0, 2.0]})
        df2 = pd.DataFrame({'Serial_No': [1, 2], 'Other': [1.0, 2.0]})
        resp = self._call_export({1: df1, 2: df2})
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.json()['error'], 'no_common_params')
