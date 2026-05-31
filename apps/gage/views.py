import io, os
import pandas as pd
import numpy as np
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from django.http import FileResponse

from apps.datafiles.models import DataFile
from apps.datafiles.parsers import get_parser
from apps.analysis.services.statistics import (
    compute_range_statistics, compute_cpk, get_1d_from,
    get_columns_with_limits, parse_limit_string,
)

class GageViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def generate_summary(self, request):
        file_ids = request.data.get('file_ids', [])
        only_bin1 = request.data.get('only_bin1', False)
        ignore_no_limit = request.data.get('ignore_no_limit', False)

        if len(file_ids) < 2:
            return Response({'error': 'need_at_least_2_files'}, status=400)

        file_datasets = []
        for fid in file_ids:
            df_obj = get_object_or_404(DataFile, pk=fid, owner=request.user)
            parser = get_parser(df_obj.format_type)
            df, metadata = parser.parse(df_obj.file_path)
            if df is None:
                continue
            if only_bin1:
                from apps.analysis.services.statistics import get_bin_column_name
                bin_col = get_bin_column_name(df_obj.format_type)
                if bin_col in df.columns:
                    bin_numeric = pd.to_numeric(df[bin_col], errors='coerce')
                    df = df[bin_numeric == 1].copy()
            file_datasets.append({
                'filename': df_obj.filename,
                'df': df,
                'metadata': metadata,
            })

        site_labels = []
        for i, ds in enumerate(file_datasets):
            site_labels.append(f'Site{i + 1}')

        all_col_sets = []
        for ds in file_datasets:
            cols = [c for c in ds['df'].columns if ds['df'][c].dtype in ('int64', 'float64')]
            if ignore_no_limit:
                cols = get_columns_with_limits(ds['df'], ds['metadata'])
            all_col_sets.append(set(cols))

        common = all_col_sets[0]
        for s in all_col_sets[1:]:
            common = common & s
        common = sorted(common)

        if not common:
            return Response({'error': 'no_common_items'}, status=400)

        wb = Workbook()

        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        cpk_a_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cpk_b_fill = PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")
        cpk_c_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        cpk_d_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")

        def apply_header_style(cell):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        def apply_cell_style(cell, align=None):
            cell.alignment = align or center_align
            cell.border = thin_border

        def get_cpk_fill(cpk_val):
            if cpk_val >= 1.67:
                return cpk_a_fill
            elif cpk_val >= 1.33:
                return cpk_b_fill
            elif cpk_val >= 1.0:
                return cpk_c_fill
            else:
                return cpk_d_fill

        def get_cpk_font(cpk_val):
            return Font(bold=True, color="FFFFFF" if cpk_val < 1.33 else "000000")

        # ===== Summary Sheet =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_headers = ['测试项']
        for site_label in site_labels:
            summary_headers += [
                f'{site_label} Mean', f'{site_label} STD', f'{site_label} Cpk',
            ]

        for c, h in enumerate(summary_headers, 1):
            apply_header_style(ws_summary.cell(row=1, column=c, value=h))

        for r, param in enumerate(common, 2):
            apply_cell_style(ws_summary.cell(row=r, column=1, value=param), left_align)
            col = 2
            for ds_idx, ds in enumerate(file_datasets):
                data = get_1d_from(ds['df'], param).dropna()
                count = len(data)
                mean_v = round(float(data.mean()), 6) if count > 0 else 0
                std_v = round(float(data.std(ddof=0)), 6) if count > 1 else 0

                try:
                    lower = float(ds['metadata']['mins'][param])
                    upper = float(ds['metadata']['maxs'][param])
                    cpk_val, cpk_level, _ = compute_cpk(mean_v, std_v, lower, upper)
                except Exception:
                    cpk_val = 0

                for i, val in enumerate([mean_v, std_v, cpk_val]):
                    cell = ws_summary.cell(row=r, column=col + i, value=val)
                    apply_cell_style(cell)
                    if i == 2:
                        cell.fill = get_cpk_fill(cpk_val)
                        cell.font = get_cpk_font(cpk_val)
                        cell.number_format = '0.0000'

                col += 3

        for c_idx in range(1, len(summary_headers) + 1):
            max_width = 12
            for row_data in ws_summary.iter_rows(min_col=c_idx, max_col=c_idx):
                for cell in row_data:
                    if cell.value is not None:
                        max_width = max(max_width, len(str(cell.value)) + 4)
            ws_summary.column_dimensions[get_column_letter(c_idx)].width = max_width

        # ===== Per-Site Sheets =====
        for ds_idx, (site_label, ds) in enumerate(zip(site_labels, file_datasets)):
            df = ds['df']
            meta = ds['metadata']

            ws = wb.create_sheet(title=site_label)
            per_site_headers = ['测试项', 'Count', 'Mean', 'STD', 'Cpk', 'Min', 'Max', 'Range']

            for c, h in enumerate(per_site_headers, 1):
                apply_header_style(ws.cell(row=1, column=c, value=h))

            for r, param in enumerate(common, 2):
                apply_cell_style(ws.cell(row=r, column=1, value=param), left_align)

                data = get_1d_from(df, param).dropna()
                count = len(data)
                mean_v = round(float(data.mean()), 6) if count > 0 else 0
                std_v = round(float(data.std(ddof=0)), 6) if count > 1 else 0
                min_v = round(float(data.min()), 6) if count > 0 else 0
                max_v = round(float(data.max()), 6) if count > 0 else 0
                range_v = round(max_v - min_v, 6)

                try:
                    lower = float(meta['mins'][param])
                    upper = float(meta['maxs'][param])
                    cpk_val, cpk_level, _ = compute_cpk(mean_v, std_v, lower, upper)
                except Exception:
                    cpk_val = 0

                row_data = [count, mean_v, std_v, cpk_val, min_v, max_v, range_v]
                for i, val in enumerate(row_data):
                    cell = ws.cell(row=r, column=2 + i, value=val)
                    apply_cell_style(cell)
                    if i == 2:
                        cell.fill = get_cpk_fill(cpk_val)
                        cell.font = get_cpk_font(cpk_val)
                        cell.number_format = '0.0000'

            for c_idx in range(1, len(per_site_headers) + 1):
                max_width = 12
                for row_data in ws.iter_rows(min_col=c_idx, max_col=c_idx):
                    for cell in row_data:
                        if cell.value is not None:
                            max_width = max(max_width, len(str(cell.value)) + 4)
                ws.column_dimensions[get_column_letter(c_idx)].width = max_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Gage_Summary.xlsx',
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
