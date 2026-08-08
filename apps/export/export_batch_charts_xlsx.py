"""Batch charts Excel export with embedded histogram images."""

import io
import os
from concurrent.futures import ProcessPoolExecutor

import pandas as pd

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

from apps.analysis.services.statistics import (
    get_1d_from, compute_range_statistics, compute_cpk, compute_site_stats,
)
from .charts import _create_histogram_chart
from .chart_workers import render_histogram_worker

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def build_batch_charts_xlsx_with_charts(df, metadata, params, site_col=None,
                                         show_limit=True, show_3sigma=False,
                                         show_4sigma=False, show_6sigma=True,
                                         show_normal=False, show_kde=False):
    """Build batch charts Excel with embedded histogram images.

    Ported from old project's export_batch_distribution_chart_excel + _export_charts_to_xlsx.
    Returns bytes of the .xlsx file.
    """
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "总览"

    fail_fill = PatternFill(start_color="F54927", end_color="F54927", fill_type="solid")
    fail_font = Font(color="FFFFFF", bold=True)
    all_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    normal_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    stats_label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    summary_hyperlink_font = Font(name='Calibri', size=11, color="0563C1")
    return_hyperlink_font = Font(name='Calibri', size=11, color="0563C1")

    CPK_COLOR_MAP = {
        'green': ("4CAF50", "FFFFFF"), 'orange': ("FFA726", "FFFFFF"),
        'darkorange': ("FF7043", "FFFFFF"), 'red': ("F44336", "FFFFFF"),
    }

    def _apply_cpk_style(cell, cpk_color_name):
        bg, fg = CPK_COLOR_MAP.get(cpk_color_name, ("9E9E9E", "FFFFFF"))
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(color=fg, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _get_text_width(text):
        s = str(text) if text is not None else ""
        cn_len = len([c for c in s if '一' <= c <= '鿿'])
        return len(s) + cn_len + 2

    # ── Summary sheet headers ──
    summary_headers = ["序号", "参数", "数据点数", "Mean", "STD", "Min", "Max", "CPK", "CPK Level", "ALL Site Yield"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    col_max_widths = {i: _get_text_width(summary_headers[i - 1]) for i in range(1, len(summary_headers) + 1)}

    # ── Process each parameter: compute stats only, defer image creation ──
    processed_params = []
    summary_data_list = []

    param_order = {col: idx for idx, col in enumerate(df.columns)}
    params_sorted = sorted(params, key=lambda x: param_order.get(x, 999999))

    for selected_param in params_sorted:
        if selected_param not in df.columns:
            continue
        data_series = get_1d_from(df, selected_param).dropna()
        data_series = data_series[data_series.apply(lambda x: abs(x) < float('inf'))]
        if len(data_series) == 0:
            continue

        stats = compute_range_statistics(data_series, metadata, selected_param)
        cpk_result = compute_cpk(stats['mean'], stats['std'], stats['rdl'][0], stats['rdl'][1])
        mean_val = stats['mean']
        std_val = stats['std']
        rdl_min = stats['rdl'][0]
        rdl_max = stats['rdl'][1]

        # Site stats
        site_stats = []
        if site_col and site_col in df.columns:
            site_series = get_1d_from(df, selected_param)
            site_idx = get_1d_from(df, site_col)
            site_stats = compute_site_stats(site_series, site_idx, rdl_min, rdl_max, None, None, False) or []

        processed_params.append(selected_param)
        summary_data_list.append({
            'stats_data': {
                'param_name': selected_param, 'mean_val': round(mean_val, 4),
                'std_val': round(std_val, 4), 'data_range': f"{round(float(data_series.min()), 4)} ~ {round(float(data_series.max()), 4)}",
                'count': len(data_series), 'cpk_str': round(cpk_result['cpk'], 4),
                'cpk_color': cpk_result.get('cpk_color', 'gray'),
                'low_limit': round(rdl_min, 4) if rdl_min is not None else 'N/A',
                'high_limit': round(rdl_max, 4) if rdl_max is not None else 'N/A',
                'unit': stats.get('unit', ''),
            },
            'site_stats': site_stats,
            'chart_data': {
                'data_series': data_series,
                'mean_val': mean_val,
                'std_val': std_val,
                'rdl_min': rdl_min,
                'rdl_max': rdl_max,
            },
        })

    # ── Fill summary rows ──
    for param_idx, (selected_param, entry) in enumerate(zip(processed_params, summary_data_list), 1):
        stats_data = entry['stats_data']
        site_stats_list = entry['site_stats']

        param_safe = selected_param.replace("/", "_").replace("\\", "_").replace(" ", "_").replace("-", "_")[:28]
        cell_param = ws_summary.cell(row=param_idx + 1, column=2, value=selected_param)
        cell_param.hyperlink = f"#'{param_safe}'!A1"
        cell_param.font = summary_hyperlink_font

        ws_summary.cell(row=param_idx + 1, column=1, value=param_idx).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=3, value=stats_data['count']).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=4, value=str(stats_data['mean_val'])).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=5, value=str(stats_data['std_val'])).alignment = Alignment(horizontal="center")

        range_parts = str(stats_data['data_range']).split(' ~ ')
        ws_summary.cell(row=param_idx + 1, column=6, value=range_parts[0] if len(range_parts) > 0 else '').alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=7, value=range_parts[1] if len(range_parts) > 1 else '').alignment = Alignment(horizontal="center")

        cell_cpk = ws_summary.cell(row=param_idx + 1, column=8, value=stats_data['cpk_str'])
        _apply_cpk_style(cell_cpk, stats_data.get('cpk_color', 'gray'))
        ws_summary.cell(row=param_idx + 1, column=9, value=str(stats_data['cpk_str'])).alignment = Alignment(horizontal="center")

        all_site_yield = "100.00%"
        for si in site_stats_list:
            if si.get('Site') == 'ALL Site':
                all_site_yield = si.get('Yield', '100.00%')
                break
        cell_yield = ws_summary.cell(row=param_idx + 1, column=10, value=all_site_yield)
        cell_yield.alignment = Alignment(horizontal="center")
        for si in site_stats_list:
            if si.get('Site') == 'ALL Site':
                if si.get('FailCount', 0) > 0:
                    cell_yield.fill = fail_fill
                    cell_yield.font = fail_font
                else:
                    cell_yield.fill = normal_fill
                break

        row_vals = [(1, param_idx), (2, selected_param), (3, stats_data['count']),
                     (4, str(stats_data['mean_val'])), (5, str(stats_data['std_val'])),
                     (6, range_parts[0] if len(range_parts) > 0 else ''),
                     (7, range_parts[1] if len(range_parts) > 1 else ''),
                     (8, str(stats_data['cpk_str'])), (9, str(stats_data['cpk_str'])), (10, all_site_yield)]
        for col_i, val in row_vals:
            w = _get_text_width(val)
            if w > col_max_widths.get(col_i, 0):
                col_max_widths[col_i] = w

    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = min(col_max_widths.get(col, 15), 40)

    # ── Per-parameter detail sheets: parallel chart rendering + sheets ──
    # openpyxl lazily reads image data during wb.save(), so we must keep each
    # buffer alive until after save() and close them all at the end.
    open_buffers = []

    # 预提取 site 数据（可 pickle 的 ndarray/标量，供多进程任务）
    site_series = None
    site_values = None
    if site_col and site_col in df.columns:
        site_series_raw = df[site_col]
        if isinstance(site_series_raw, pd.DataFrame):
            site_series_raw = site_series_raw.iloc[:, 0]
        site_values = sorted(site_series_raw.dropna().unique(), key=lambda x: (isinstance(x, float), x))
        site_series = site_series_raw.to_numpy()

    # 多进程并行渲染全部参数图表（瓶颈在 matplotlib 渲染；worker 见 chart_workers.py）
    png_results = {}
    try:
        workers = min(8, max(2, os.cpu_count() or 4))
        with ProcessPoolExecutor(max_workers=workers) as ex:
            futures = {}
            for title, entry in zip(processed_params, summary_data_list):
                chart_data = entry['chart_data']
                futures[title] = ex.submit(
                    render_histogram_worker,
                    {
                        'param': title,
                        'data_series': chart_data['data_series'].to_numpy(),
                        'site_values': site_values,
                        'site_series': site_series,
                        'mean_val': chart_data['mean_val'],
                        'std_val': chart_data['std_val'],
                        'rdl_min': chart_data['rdl_min'],
                        'rdl_max': chart_data['rdl_max'],
                        'show_limit': show_limit,
                        'show_3sigma': show_3sigma,
                        'show_4sigma': show_4sigma,
                        'show_6sigma': show_6sigma,
                        'show_normal': show_normal,
                        'show_kde': show_kde,
                    },
                )
            for title in processed_params:
                try:
                    png_results[title] = futures[title].result()
                except Exception:
                    png_results[title] = None  # 单参数失败 → 串行兜底
    except Exception:
        png_results = {}  # 并行整体失败 → 全量串行兜底（与现状等价）

    for idx, (title, entry) in enumerate(zip(processed_params, summary_data_list)):
        stats_data = entry.get('stats_data', {})
        site_stats_list = entry.get('site_stats', [])
        chart_data = entry.get('chart_data', {})

        # Chart PNG: parallel result first, fall back to serial render
        png = png_results.get(title)
        if png is None:
            img_buffer = _create_histogram_chart(
                df, metadata, title, chart_data['data_series'],
                chart_data['mean_val'], chart_data['std_val'],
                chart_data['rdl_min'], chart_data['rdl_max'],
                show_limit=show_limit, show_3sigma=show_3sigma,
                show_4sigma=show_4sigma, show_6sigma=show_6sigma,
                show_normal=show_normal, show_kde=show_kde, site_col=site_col
            )
        else:
            img_buffer = io.BytesIO(png)
        if img_buffer.getbuffer().nbytes == 0:
            img_buffer.close()
            continue
        open_buffers.append(img_buffer)

        param_safe = title.replace("/", "_").replace("\\", "_").replace(" ", "_").replace("-", "_")[:28]
        try:
            ws = wb.create_sheet(title=param_safe)
        except ValueError:
            ws = wb.create_sheet(title=f"Chart_{idx}")

        # Stats header
        header_row = ["统计项", "Low Limit", "High Limit", "Unit"]
        for col_idx, h in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        if stats_data:
            ws.cell(row=2, column=1, value=stats_data.get('param_name', ''))
            ws.cell(row=2, column=2, value=stats_data.get('low_limit', 0)).number_format = '0.0000'
            ws.cell(row=2, column=3, value=stats_data.get('high_limit', 0)).number_format = '0.0000'
            ws.cell(row=2, column=4, value=stats_data.get('unit', ''))

            left_labels = [("Mean", "mean_val", 4, 1), ("STD", "std_val", 5, 1),
                           ("Range", "data_range", 6, 1), ("数据点数", "count", 7, 1),
                           ("CPK", "cpk_str", 8, 1)]
            for label, key, row, col in left_labels:
                ws.cell(row=row, column=col, value=label).fill = stats_label_fill
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                val = stats_data.get(key, '')
                cell = ws.cell(row=row, column=col + 1, value=str(val))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if key == "cpk_str":
                    _apply_cpk_style(cell, stats_data.get('cpk_color', 'gray'))
                elif key in ("mean_val", "std_val"):
                    cell.number_format = '0.0000'

        # Site stats table
        if site_stats_list:
            ws.cell(row=1, column=6, value="Site统计").fill = HEADER_FILL
            ws.cell(row=1, column=6).font = HEADER_FONT
            ws.cell(row=1, column=6).alignment = HEADER_ALIGNMENT

            site_headers = ["Site", "Yield", "FailCount", "ExceedMin", "ExceedMax"]
            for col_idx, h in enumerate(site_headers, 1):
                cell = ws.cell(row=2, column=col_idx + 5, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGNMENT

            for row_offset, site_info in enumerate(site_stats_list):
                is_fail = site_info.get('FailCount', 0) > 0
                is_all = site_info.get('Site') == 'ALL Site'
                for col_idx, key in enumerate(['Site', 'Yield', 'FailCount', 'ExceedMin', 'ExceedMax'], 1):
                    cell = ws.cell(row=row_offset + 3, column=col_idx + 5, value=site_info.get(key, ''))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if is_fail:
                        cell.fill = fail_fill
                        cell.font = fail_font
                    elif is_all:
                        cell.fill = all_fill
                    else:
                        cell.fill = normal_fill

        # Column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
        for col in range(6, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Return hyperlink
        return_cell = ws.cell(row=8, column=5, value="← 返回总览")
        return_cell.hyperlink = "#'总览'!A1"
        return_cell.font = return_hyperlink_font

        # Embed chart image
        max_site_rows = 2 + len(site_stats_list) + 1 if site_stats_list else 2
        chart_row = max(9, max_site_rows + 1)

        img_buffer.seek(0)
        img = XlImage(img_buffer)
        img.width = 800
        img.height = 450
        ws.add_image(img, f'A{chart_row}')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    result = output.getvalue()
    output.close()
    for buf in open_buffers:
        buf.close()
    return result
