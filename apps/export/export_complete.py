"""Complete export functions from old version."""

import io
import pandas as pd
import numpy as np
import excelize
import tempfile
import os
from typing import Dict, List, Optional
from apps.analysis.services.statistics import (
    ensure_numeric, get_bin_column_name, detect_fail_data, get_site_column,
    compute_range_statistics, compute_cpk, compute_site_stats, get_1d_from,
)

# ── Batch chart export (openpyxl + matplotlib) ──

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

COLORS_SITE_8 = ['#E53935', '#1E88E5', '#43A047', '#F9A825', '#8E24AA', '#00ACC1', '#F57C00', '#D81B60']
COLOR_LSL = '#C62828'
COLOR_USL = '#C62828'
COLOR_SIGMA_3 = '#1565C0'
COLOR_SIGMA_4 = '#00838F'
COLOR_SIGMA_6 = '#E65100'
COLOR_NORMAL = '#F57F17'

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _get_export_dpi():
    return 150


def _create_histogram_chart(
    df, metadata, selected_param, data_series, mean_val, std_val,
    rdl_min, rdl_max, show_limit=True, show_3sigma=False,
    show_4sigma=False, show_6sigma=False, show_normal=False, site_col=None
):
    """Ported from old project: create matplotlib histogram with site grouping."""
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False

    if len(data_series) == 0:
        return io.BytesIO()

    param_low_limit = rdl_min if rdl_min is not None else float(data_series.min())
    param_high_limit = rdl_max if rdl_max is not None else float(data_series.max())

    data_gap = (param_high_limit - param_low_limit) / 20 if (param_high_limit - param_low_limit) > 0 else 1.0
    x_labels = [param_low_limit + (i - 2) * data_gap for i in range(25)]
    bin_start = param_low_limit - 2.5 * data_gap
    all_bins = np.array([bin_start + j * data_gap for j in range(26)])

    fig, ax = plt.subplots(figsize=(10, 5.5))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#fafafa')

    colors = COLORS_SITE_8
    y_max_plot = 100

    def _add_vline_label(exp_ax, x, label, exp_color):
        label_y = y_max_plot - y_max_plot * 0.02
        exp_ax.annotate(label, xy=(x, label_y), fontsize=8, color=exp_color,
                        fontweight='bold', ha='center',
                        bbox=dict(boxstyle='round,pad=0.15', facecolor='white', edgecolor=exp_color, alpha=0.85),
                        clip_on=False)

    if site_col and site_col in df.columns:
        site_values = sorted(df[site_col].dropna().unique(), key=lambda x: (isinstance(x, float), x))
        n_sites = len(site_values)
        bar_width = data_gap * 0.8 / n_sites
        site_df = df[[site_col, selected_param]].copy()
        site_col_1d = site_df[site_col]
        if isinstance(site_col_1d, pd.DataFrame):
            site_col_1d = site_col_1d.iloc[:, 0]
        site_df[selected_param] = pd.to_numeric(site_df[selected_param], errors='coerce')
        grouped = site_df.groupby(site_col_1d)

        for idx, site in enumerate(site_values):
            if site in grouped.groups:
                sdata = grouped.get_group(site)[selected_param].dropna()
            else:
                sdata = pd.Series(dtype=float)
            total = len(sdata)
            hist, _ = np.histogram(sdata, bins=all_bins)
            hist_percent = [round((count / total) * 100, 2) if total > 0 else 0 for count in hist]
            bar_data = [hist_percent[i] if i < len(hist_percent) else 0 for i in range(25)]
            offset = (idx - n_sites / 2 + 0.5) * bar_width
            bar_x = [x_labels[i] + offset for i in range(25)]

            ax.bar(
                bar_x, bar_data, width=bar_width * 0.9,
                color=colors[idx % len(colors)], alpha=0.7,
                label=f'Site{site}%', edgecolor='white', linewidth=0.5
            )
    else:
        data_clean = data_series
        hist, _ = np.histogram(data_clean, bins=all_bins)
        total = len(data_clean)
        hist_percent = [round((count / total) * 100, 2) if total > 0 else 0 for count in hist]
        bar_data = [hist_percent[i] if i < len(hist_percent) else 0 for i in range(25)]
        ax.bar(x_labels, bar_data, width=data_gap * 0.9, color='#1E88E5', alpha=0.7, label='数据分布', edgecolor='white', linewidth=0.5)

    if show_limit and rdl_min is not None:
        ax.axvline(x=param_low_limit, color=COLOR_LSL, linewidth=2.5, linestyle='--')
    if show_limit and rdl_max is not None:
        ax.axvline(x=param_high_limit, color=COLOR_USL, linewidth=2.5, linestyle='--')

    for sigma, flag, color, label_prefix in [
        (3, show_3sigma, COLOR_SIGMA_3, '3σ'), (4, show_4sigma, COLOR_SIGMA_4, '4σ'), (6, show_6sigma, COLOR_SIGMA_6, '6σ')
    ]:
        if flag and std_val > 0:
            lower = mean_val - sigma * std_val
            upper = mean_val + sigma * std_val
            ax.axvline(x=lower, color=color, linewidth=2, linestyle=':')
            ax.axvline(x=upper, color=color, linewidth=2, linestyle=':')

    if show_normal and std_val > 0:
        x_pdf = np.linspace(x_labels[0], x_labels[-1], 200)
        pdf_values = (1 / (std_val * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x_pdf - mean_val) / std_val) ** 2)
        max_pdf = np.max(pdf_values)
        if max_pdf > 0:
            normal_scaled = pdf_values / max_pdf * y_max_plot
            ax.plot(x_pdf, normal_scaled, color=COLOR_NORMAL, linewidth=3, label='正态分布')

    ax.set_ylabel('百分比 (%)', fontsize=12)
    ax.set_ylim(0, y_max_plot)
    ax.set_xlim(float(x_labels[0]), float(x_labels[-1]))
    ax.set_xticks(x_labels)
    ax.xaxis.set_major_formatter(ticker.FormatStrFormatter('%.4f'))
    ax.tick_params(axis='x', rotation=45, labelsize=6)
    ax.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.0f%%'))
    ax.grid(True, alpha=0.3)

    if show_limit and rdl_min is not None:
        _add_vline_label(ax, param_low_limit, 'LSL', COLOR_LSL)
    if show_limit and rdl_max is not None:
        _add_vline_label(ax, param_high_limit, 'USL', COLOR_USL)

    for sigma, flag, color, label_prefix in [
        (3, show_3sigma, COLOR_SIGMA_3, '3σ'), (4, show_4sigma, COLOR_SIGMA_4, '4σ'), (6, show_6sigma, COLOR_SIGMA_6, '6σ')
    ]:
        if flag and std_val > 0:
            lower = mean_val - sigma * std_val
            upper = mean_val + sigma * std_val
            _add_vline_label(ax, lower, f'{label_prefix}下限', color)
            _add_vline_label(ax, upper, f'{label_prefix}上限', color)

    ax.legend(fontsize=8, loc='upper center', bbox_to_anchor=(0.5, -0.2), ncol=4)
    ax.set_title(selected_param, fontsize=14, fontweight='bold', color='#0066cc', pad=15)
    plt.tight_layout()

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=_get_export_dpi(), bbox_inches='tight')
    plt.close(fig)
    img_buffer.seek(0)
    return img_buffer


def build_batch_charts_xlsx_with_charts(df, metadata, params, site_col=None,
                                         show_limit=True, show_3sigma=False,
                                         show_4sigma=False, show_6sigma=True,
                                         show_normal=False):
    """Build batch charts Excel with embedded histogram images.

    Ported from old project's export_batch_distribution_chart_excel + _export_charts_to_xlsx.
    Returns bytes of the .xlsx file.
    """
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "总览"

    fail_fill = PatternFill(start_color="F54927", end_color="F54927", fill_type="solid")
    fail_font = Font(color="FFFFFF", bold=True)
    all_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    normal_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    stats_label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    summary_hyperlink_font = Font(name='Calibri', size=11, color="0563C1")
    return_hyperlink_font = Font(name='Calibri', size=11, color="0563C1")

    CPK_COLOR_MAP = {
        'green': ("4CAF50", "FFFFFF"), 'orange': ("FFA726", "FFFFFF"),
        'darkorange': ("FF7043", "FFFFFF"), 'red': ("F44336", "FFFFFF"),
    }

    def _apply_cpk_style(cell, cpk_color_name):
        bg, fg = CPK_COLOR_MAP.get(cpk_color_name, ("9E9E9E", "FFFFFF"))
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(color=fg, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _get_text_width(text):
        s = str(text) if text is not None else ""
        cn_len = len([c for c in s if '一' <= c <= '鿿'])
        return len(s) + cn_len + 2

    # ── Summary sheet headers ──
    summary_headers = ["序号", "参数", "数据点数", "Mean", "STD", "Min", "Max", "CPK", "CPK Level", "ALL Site Yield"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    col_max_widths = {i: _get_text_width(summary_headers[i - 1]) for i in range(1, len(summary_headers) + 1)}

    # ── Process each parameter ──
    chart_images = []
    titles = []
    summary_data_list = []

    param_order = {col: idx for idx, col in enumerate(df.columns)}
    params_sorted = sorted(params, key=lambda x: param_order.get(x, 999999))

    for selected_param in params_sorted:
        if selected_param not in df.columns:
            continue
        data_series = get_1d_from(df, selected_param).dropna()
        data_series = data_series[data_series.apply(lambda x: abs(x) < float('inf'))]
        if len(data_series) == 0:
            continue

        stats = compute_range_statistics(data_series, metadata, selected_param)
        cpk_result = compute_cpk(stats['mean'], stats['std'], stats['rdl'][0], stats['rdl'][1])
        mean_val = stats['mean']
        std_val = stats['std']
        rdl_min = stats['rdl'][0]
        rdl_max = stats['rdl'][1]

        # Site stats
        site_stats = []
        if site_col and site_col in df.columns:
            site_series = get_1d_from(df, selected_param)
            site_idx = get_1d_from(df, site_col)
            site_stats = compute_site_stats(site_series, site_idx, rdl_min, rdl_max, None, None, False) or []

        summary_data_list.append({
            'stats_data': {
                'param_name': selected_param, 'mean_val': round(mean_val, 4),
                'std_val': round(std_val, 4), 'data_range': f"{round(float(data_series.min()), 4)} ~ {round(float(data_series.max()), 4)}",
                'count': len(data_series), 'cpk_str': round(cpk_result['cpk'], 4),
                'cpk_color': cpk_result.get('cpk_color', 'gray'),
                'low_limit': round(rdl_min, 4) if rdl_min is not None else 'N/A',
                'high_limit': round(rdl_max, 4) if rdl_max is not None else 'N/A',
                'unit': stats.get('unit', ''),
            },
            'site_stats': site_stats,
        })

        # Create histogram image
        img_buffer = _create_histogram_chart(
            df, metadata, selected_param, data_series, mean_val, std_val,
            rdl_min, rdl_max, show_limit=show_limit, show_3sigma=show_3sigma,
            show_4sigma=show_4sigma, show_6sigma=show_6sigma,
            show_normal=show_normal, site_col=site_col
        )
        chart_images.append(img_buffer)
        titles.append(selected_param)

    # ── Fill summary rows ──
    for param_idx, (selected_param, entry) in enumerate(zip(params_sorted, [d for d in summary_data_list]), 1):
        stats_data = entry['stats_data']
        site_stats_list = entry['site_stats']

        param_safe = selected_param.replace("/", "_").replace("\\", "_").replace(" ", "_").replace("-", "_")[:28]
        cell_param = ws_summary.cell(row=param_idx + 1, column=2, value=selected_param)
        cell_param.hyperlink = f"#'{param_safe}'!A1"
        cell_param.font = summary_hyperlink_font

        ws_summary.cell(row=param_idx + 1, column=1, value=param_idx).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=3, value=stats_data['count']).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=4, value=str(stats_data['mean_val'])).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=5, value=str(stats_data['std_val'])).alignment = Alignment(horizontal="center")

        range_parts = str(stats_data['data_range']).split(' ~ ')
        ws_summary.cell(row=param_idx + 1, column=6, value=range_parts[0] if len(range_parts) > 0 else '').alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=7, value=range_parts[1] if len(range_parts) > 1 else '').alignment = Alignment(horizontal="center")

        cell_cpk = ws_summary.cell(row=param_idx + 1, column=8, value=stats_data['cpk_str'])
        _apply_cpk_style(cell_cpk, stats_data.get('cpk_color', 'gray'))
        ws_summary.cell(row=param_idx + 1, column=9, value=str(stats_data['cpk_str'])).alignment = Alignment(horizontal="center")

        all_site_yield = "100.00%"
        for si in site_stats_list:
            if si.get('Site') == 'ALL Site':
                all_site_yield = si.get('Yield', '100.00%')
                break
        cell_yield = ws_summary.cell(row=param_idx + 1, column=10, value=all_site_yield)
        cell_yield.alignment = Alignment(horizontal="center")
        for si in site_stats_list:
            if si.get('Site') == 'ALL Site':
                if si.get('FailCount', 0) > 0:
                    cell_yield.fill = fail_fill
                    cell_yield.font = fail_font
                else:
                    cell_yield.fill = normal_fill
                break

        row_vals = [(1, param_idx), (2, selected_param), (3, stats_data['count']),
                     (4, str(stats_data['mean_val'])), (5, str(stats_data['std_val'])),
                     (6, range_parts[0] if len(range_parts) > 0 else ''),
                     (7, range_parts[1] if len(range_parts) > 1 else ''),
                     (8, str(stats_data['cpk_str'])), (9, str(stats_data['cpk_str'])), (10, all_site_yield)]
        for col_i, val in row_vals:
            w = _get_text_width(val)
            if w > col_max_widths.get(col_i, 0):
                col_max_widths[col_i] = w

    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = min(col_max_widths.get(col, 15), 40)

    # ── Per-parameter detail sheets ──
    for idx, (img_buffer, title) in enumerate(zip(chart_images, titles)):
        if img_buffer.getbuffer().nbytes == 0:
            continue
        param_safe = title.replace("/", "_").replace("\\", "_").replace(" ", "_").replace("-", "_")[:28]
        try:
            ws = wb.create_sheet(title=param_safe)
        except ValueError:
            ws = wb.create_sheet(title=f"Chart_{idx}")

        entry = summary_data_list[idx] if idx < len(summary_data_list) else {}
        stats_data = entry.get('stats_data', {})
        site_stats_list = entry.get('site_stats', [])

        # Stats header
        header_row = ["统计项", "Low Limit", "High Limit", "Unit"]
        for col_idx, h in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        if stats_data:
            ws.cell(row=2, column=1, value=stats_data.get('param_name', ''))
            ws.cell(row=2, column=2, value=stats_data.get('low_limit', 0)).number_format = '0.0000'
            ws.cell(row=2, column=3, value=stats_data.get('high_limit', 0)).number_format = '0.0000'
            ws.cell(row=2, column=4, value=stats_data.get('unit', ''))

            left_labels = [("Mean", "mean_val", 4, 1), ("STD", "std_val", 5, 1),
                           ("Range", "data_range", 6, 1), ("数据点数", "count", 7, 1),
                           ("CPK", "cpk_str", 8, 1)]
            for label, key, row, col in left_labels:
                ws.cell(row=row, column=col, value=label).fill = stats_label_fill
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                val = stats_data.get(key, '')
                cell = ws.cell(row=row, column=col + 1, value=str(val))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if key == "cpk_str":
                    _apply_cpk_style(cell, stats_data.get('cpk_color', 'gray'))
                elif key in ("mean_val", "std_val"):
                    cell.number_format = '0.0000'

        # Site stats table
        if site_stats_list:
            ws.cell(row=1, column=6, value="Site统计").fill = HEADER_FILL
            ws.cell(row=1, column=6).font = HEADER_FONT
            ws.cell(row=1, column=6).alignment = HEADER_ALIGNMENT

            site_headers = ["Site", "Yield", "FailCount", "ExceedMin", "ExceedMax"]
            for col_idx, h in enumerate(site_headers, 1):
                cell = ws.cell(row=2, column=col_idx + 5, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGNMENT

            for row_offset, site_info in enumerate(site_stats_list):
                is_fail = site_info.get('FailCount', 0) > 0
                is_all = site_info.get('Site') == 'ALL Site'
                for col_idx, key in enumerate(['Site', 'Yield', 'FailCount', 'ExceedMin', 'ExceedMax'], 1):
                    cell = ws.cell(row=row_offset + 3, column=col_idx + 5, value=site_info.get(key, ''))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if is_fail:
                        cell.fill = fail_fill
                        cell.font = fail_font
                    elif is_all:
                        cell.fill = all_fill
                    else:
                        cell.fill = normal_fill

        # Column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
        for col in range(6, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Return hyperlink
        return_cell = ws.cell(row=8, column=5, value="← 返回总览")
        return_cell.hyperlink = "#'总览'!A1"
        return_cell.font = return_hyperlink_font

        # Embed chart image
        max_site_rows = 2 + len(site_stats_list) + 1 if site_stats_list else 2
        chart_row = max(9, max_site_rows + 1)

        img_buffer.seek(0)
        img = XlImage(img_buffer)
        img.width = 800
        img.height = 450
        ws.add_image(img, f'A{chart_row}')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ── Helper Functions ──

def _convert_to_native_type(val):
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (ValueError, TypeError):
        pass
    if isinstance(val, (bool,)):
        return val
    if isinstance(val, (int,)):
        return int(val) if hasattr(val, 'item') else val
    if isinstance(val, (float,)):
        if hasattr(val, 'item'):
            native = val.item()
        else:
            native = val
        if pd.isna(native):
            return ""
        return native
    if hasattr(val, 'item'):
        native = val.item()
        try:
            if pd.isna(native):
                return ""
        except (ValueError, TypeError):
            pass
        return native
    return str(val) if val is not None else ""



# 导出DataFrame为CSV，支持Site过滤、Pass/Fail过滤和原始格式匹配
def export_to_csv(df: pd.DataFrame, metadata: Dict, site_filter=None, passfail_filter=None, 
                  keep_header=False, match_original_format=False, raw_lines=None) -> bytes:
    try:
       
        export_df = df.copy()
        
        if site_filter and site_filter != "全部":
            site_col = get_site_column(df)
            if site_col:
                export_df = export_df[export_df[site_col].astype(str) == str(site_filter)]
        
        if passfail_filter and passfail_filter != "全部":

            format_type = metadata.get('format', 'CTA8290D')
            bin_col = get_bin_column_name(format_type)
            
            if bin_col in export_df.columns:
                if passfail_filter == "Pass":
                    export_df = export_df[ensure_numeric(export_df, bin_col) == 1]
                elif passfail_filter == "Fail":
                    export_df = export_df[ensure_numeric(export_df, bin_col) != 1]
        
        if keep_header and raw_lines:

            format_type = metadata.get('format', 'CTA8290D')
            config = DATA_FORMAT_CONFIG.get(format_type)
            
            if config and raw_lines:
                data_marker_line = None
                for i, line in enumerate(raw_lines):
                    if config['marker'] in line:
                        data_marker_line = i
                        break
                
                if data_marker_line is not None:
                    if format_type in ['CTA8290D', 'CTA8280F']:
                        header_end_line = data_marker_line + config['max_offset']
                    elif format_type == 'ETS88':
                        header_end_line = data_marker_line + 3
                    else:
                        header_end_line = data_marker_line
                    
                    header_lines = []
                    for i in range(header_end_line + 1):
                        header_lines.append(raw_lines[i].rstrip('\n').rstrip('\r'))
                    
                    header_content = '\n'.join(header_lines)
                    
                    if len(export_df) > 0:
                        if match_original_format:
                            if format_type == 'ETS88':
                                export_bin_col = get_bin_column_name(format_type)
                                coord_x_col = 'XCoord'
                                coord_y_col = 'YCoord'
                                
                                meta_cols = ['Site #', 'Serial #', export_bin_col, coord_x_col, coord_y_col]
                                existing_meta_cols = [c for c in meta_cols if c in export_df.columns]
                                other_cols = [c for c in export_df.columns if c not in meta_cols]
                                ordered_cols = existing_meta_cols + other_cols
                                export_df = export_df[ordered_cols]
                            
                            data_csv = export_df.to_csv(index=False, header=False)
                            csv_content = header_content + '\n' + data_csv
                        else:
                            data_csv = export_df.to_csv(index=False)
                            csv_content = header_content + '\n' + data_csv
                    else:
                        csv_content = header_content
                else:
                    csv_content = export_df.to_csv(index=False)
            else:
                csv_content = export_df.to_csv(index=False)
        elif match_original_format and raw_lines:

            format_type = metadata.get('format', 'CTA8290D')
            config = DATA_FORMAT_CONFIG.get(format_type)
            
            if config and raw_lines:
                data_start = None
                for i, line in enumerate(raw_lines):
                    if config['marker'] in line:
                        data_start = i
                        break
                
                if data_start is not None:
                    marker_line = raw_lines[data_start].strip()
                    
                    if len(export_df) > 0:
                        if format_type == 'ETS88':
                            export_bin_col = get_bin_column_name(format_type)
                            coord_x_col = 'XCoord'
                            coord_y_col = 'YCoord'
                            
                            meta_cols = ['Site #', 'Serial #', export_bin_col, coord_x_col, coord_y_col]
                            existing_meta_cols = [c for c in meta_cols if c in export_df.columns]
                            other_cols = [c for c in export_df.columns if c not in meta_cols]
                            ordered_cols = existing_meta_cols + other_cols
                            export_df = export_df[ordered_cols]
                        
                        data_csv = export_df.to_csv(index=False, header=False)
                        csv_content = marker_line + '\n' + data_csv
                    else:
                        csv_content = marker_line
                else:
                    csv_content = export_df.to_csv(index=False)
            else:
                csv_content = export_df.to_csv(index=False)
        else:
            csv_content = export_df.to_csv(index=False)
        
        return csv_content.encode('utf-8-sig')
    except ImportError:
        # Fallback when app module is not available
        export_df = df.copy()
        if site_filter and site_filter != "全部":
            # Simplified site filtering without get_site_column
            site_cols = [col for col in df.columns if 'Site' in col or 'site' in col]
            if site_cols:
                export_df = export_df[export_df[site_cols[0]].astype(str) == str(site_filter)]
        if passfail_filter and passfail_filter != "全部":
            # Simplified pass/fail filtering
            bin_cols = [col for col in df.columns if 'Bin' in col or 'bin' in col]
            if bin_cols:
                if passfail_filter == "Pass":
                    export_df = export_df[pd.to_numeric(export_df[bin_cols[0]], errors='coerce') == 1]
                elif passfail_filter == "Fail":
                    export_df = export_df[pd.to_numeric(export_df[bin_cols[0]], errors='coerce') != 1]
        return export_df.to_csv(index=False).encode('utf-8-sig')



def export_to_xlsx_optimized(df: pd.DataFrame, metadata: Dict) -> bytes:
    f = excelize.new_file()
    
    try:
        sheet_name = "Data"
        sheet_index = f.new_sheet(sheet_name)
        f.set_active_sheet(sheet_index)
        
        cols = df.columns.tolist()
        num_cols = len(cols)
        
        # 现代专业配色（与 buyoff 统一）
        COLOR_HEADER_BG = "2C3E50"
        COLOR_HEADER_FONT = "FFFFFF"
        COLOR_DATA_BG = "F8F9FA"
        COLOR_ALT_ROW = "EDF2F7"
        COLOR_BORDER = "BDC3C7"
        COLOR_FONT_DARK = "2C3E50"
        COLOR_RED_BG = "F5B7B1"

        header_style_id = f.new_style(excelize.Style(
            font=excelize.Font(bold=True, size=12, color=COLOR_HEADER_FONT, family="Calibri"),
            fill=excelize.Fill(type="pattern", color=[COLOR_HEADER_BG], pattern=1),
            border=[
                excelize.Border(type="left", color=COLOR_BORDER, style=2),
                excelize.Border(type="top", color=COLOR_BORDER, style=2),
                excelize.Border(type="bottom", color=COLOR_BORDER, style=2),
                excelize.Border(type="right", color=COLOR_BORDER, style=2),
            ],
            alignment=excelize.Alignment(horizontal="center", vertical="center"),
        ))

        data_style_id = f.new_style(excelize.Style(
            font=excelize.Font(size=10, color=COLOR_FONT_DARK, family="Calibri"),
            fill=excelize.Fill(type="pattern", color=[COLOR_DATA_BG], pattern=1),
            border=[
                excelize.Border(type="left", color=COLOR_BORDER, style=1),
                excelize.Border(type="top", color=COLOR_BORDER, style=1),
                excelize.Border(type="bottom", color=COLOR_BORDER, style=1),
                excelize.Border(type="right", color=COLOR_BORDER, style=1),
            ],
            alignment=excelize.Alignment(horizontal="center", vertical="center"),
        ))

        red_style_id = f.new_style(excelize.Style(
            font=excelize.Font(bold=True, size=10, color="FFFFFF", family="Calibri"),
            fill=excelize.Fill(type="pattern", color=[COLOR_RED_BG], pattern=1),
            border=[
                excelize.Border(type="left", color=COLOR_BORDER, style=1),
                excelize.Border(type="top", color=COLOR_BORDER, style=1),
                excelize.Border(type="bottom", color=COLOR_BORDER, style=1),
                excelize.Border(type="right", color=COLOR_BORDER, style=1),
            ],
            alignment=excelize.Alignment(horizontal="center", vertical="center"),
        ))

        red_bin_style_id = f.new_style(excelize.Style(
            font=excelize.Font(bold=True, size=10, color="FFFFFF", family="Calibri"),
            fill=excelize.Fill(type="pattern", color=[COLOR_RED_BG], pattern=1),
            border=[
                excelize.Border(type="left", color=COLOR_BORDER, style=1),
                excelize.Border(type="top", color=COLOR_BORDER, style=1),
                excelize.Border(type="bottom", color=COLOR_BORDER, style=1),
                excelize.Border(type="right", color=COLOR_BORDER, style=1),
            ],
            alignment=excelize.Alignment(horizontal="center", vertical="center"),
        ))
        
        for col_idx, col_name in enumerate(cols):
            cell = excelize.coordinates_to_cell_name(col_idx + 1, 1, False)
            f.set_cell_style(sheet_name, cell, cell, header_style_id)
            f.set_cell_value(sheet_name, cell, col_name)
            
            cell = excelize.coordinates_to_cell_name(col_idx + 1, 2, False)
            f.set_cell_style(sheet_name, cell, cell, data_style_id)
            f.set_cell_value(sheet_name, cell, metadata['units'].get(col_name, ""))
            
            cell = excelize.coordinates_to_cell_name(col_idx + 1, 3, False)
            f.set_cell_style(sheet_name, cell, cell, data_style_id)
            f.set_cell_value(sheet_name, cell, metadata['mins'].get(col_name, ""))
            
            cell = excelize.coordinates_to_cell_name(col_idx + 1, 4, False)
            f.set_cell_style(sheet_name, cell, cell, data_style_id)
            f.set_cell_value(sheet_name, cell, metadata['maxs'].get(col_name, ""))
        
        f.set_cell_value(sheet_name, "A5", "Min")
        f.set_cell_value(sheet_name, "A6", "Avg")
        f.set_cell_value(sheet_name, "A7", "Max")
        f.set_cell_value(sheet_name, "A8", "Range")
        f.set_cell_value(sheet_name, "A9", "STD")
        f.set_cell_value(sheet_name, "A10", "CPK")
        
        numeric_cols = [col for col in cols if df[col].dtype in ['int64', 'float64']]
        
        for col_name in numeric_cols:
            col_idx = cols.index(col_name) + 1
            col_data = ensure_numeric(df, col_name).dropna()
            
            if len(col_data) > 0:
                col_min = round(float(col_data.min()), 6)
                col_avg = round(float(col_data.mean()), 6)
                col_max = round(float(col_data.max()), 6)
                col_range = round(float(col_data.max() - col_data.min()), 6)
                col_std = round(float(col_data.std()), 6)
                
                try:
                    min_val = float(metadata['mins'][col_name])
                    max_val = float(metadata['maxs'][col_name])
                    col_cpk = round(min((max_val - col_avg) / (3 * col_std), (col_avg - min_val) / (3 * col_std)), 6) if col_std > 0 else 0
                except (ValueError, TypeError, KeyError):
                    col_cpk = 0
                
                cell = excelize.coordinates_to_cell_name(col_idx, 5, False)
                f.set_cell_style(sheet_name, cell, cell, data_style_id)
                f.set_cell_value(sheet_name, cell, col_min)
                
                cell = excelize.coordinates_to_cell_name(col_idx, 6, False)
                f.set_cell_style(sheet_name, cell, cell, data_style_id)
                f.set_cell_value(sheet_name, cell, col_avg)
                
                cell = excelize.coordinates_to_cell_name(col_idx, 7, False)
                f.set_cell_style(sheet_name, cell, cell, data_style_id)
                f.set_cell_value(sheet_name, cell, col_max)
                
                cell = excelize.coordinates_to_cell_name(col_idx, 8, False)
                f.set_cell_style(sheet_name, cell, cell, data_style_id)
                f.set_cell_value(sheet_name, cell, col_range)
                
                cell = excelize.coordinates_to_cell_name(col_idx, 9, False)
                f.set_cell_style(sheet_name, cell, cell, data_style_id)
                f.set_cell_value(sheet_name, cell, col_std)
                
                cell = excelize.coordinates_to_cell_name(col_idx, 10, False)
                f.set_cell_style(sheet_name, cell, cell, data_style_id)
                f.set_cell_value(sheet_name, cell, col_cpk)
        
        format_type = metadata.get('format', 'CTA8290D')
        target_bin_col = get_bin_column_name(format_type)
        target_bin_col_idx = cols.index(target_bin_col) + 1 if target_bin_col in cols else 1
        

        fail_indices, fail_columns, fail_cells = detect_fail_data(df, metadata)
        
        fail_row_indices = set()
        fail_col_idx_map = {col_name: idx for idx, col_name in enumerate(cols)}
        fail_cells_by_row_idx = {}
        for idx, col_list in fail_cells.items():
            fail_row_indices.add(idx)
            fail_cells_by_row_idx[idx] = set(fail_col_idx_map.get(c, -1) for c in col_list if c in fail_col_idx_map)
        
        data_start_row = 12
        
        df_values = df.values.tolist()
        data_end_row = data_start_row + len(df_values) - 1
        last_col_name = excelize.column_number_to_name(num_cols)
        
        f.set_cell_style(sheet_name, f"A{data_start_row}", f"{last_col_name}{data_end_row}", data_style_id)
        for row_idx in range(len(df_values)):
            excel_row = data_start_row + row_idx
            row_data = [_convert_to_native_type(v) for v in df_values[row_idx]]
            cell_ref = excelize.coordinates_to_cell_name(1, excel_row, False)
            f.set_sheet_row(sheet_name, cell_ref, row_data)
        
        for row_idx in fail_row_indices:
            excel_row = data_start_row + row_idx
            row_fail_col_indices = fail_cells_by_row_idx.get(row_idx, set())
            
            for col_idx in range(num_cols):
                cell = excelize.coordinates_to_cell_name(col_idx + 1, excel_row, False)
                if cols[col_idx] == target_bin_col:
                    f.set_cell_style(sheet_name, cell, cell, red_bin_style_id)
                elif col_idx in row_fail_col_indices:
                    f.set_cell_style(sheet_name, cell, cell, red_style_id)
        
        last_cell = excelize.coordinates_to_cell_name(num_cols, 11, False)
        bin_col_letter = excelize.column_number_to_name(target_bin_col_idx + 1)
        top_left_cell_ref = f"{bin_col_letter}12"
        f.set_panes(sheet_name, excelize.Panes(
            freeze=True,
            split=False,
            x_split=target_bin_col_idx,
            y_split=11,
            top_left_cell=top_left_cell_ref,
        ))
        
        f.auto_filter(sheet_name, f"A11:{last_cell}", [])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        f.save_as(tmp_path)
        f.close()
        
        with open(tmp_path, 'rb') as tmp_file:
            data = tmp_file.read()
        
        os.unlink(tmp_path)
        
        return data
        
    except Exception as e:
        f.close()
        raise e

