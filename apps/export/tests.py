"""Export 导出测试。

回归目标（2026-08-04）：
- to_excel 万行 × 百列大文件导出耗时（excelize 绑定每值 17μs 转换瓶颈 →
  手写数据区 XML + zip 重打包，实测 35s → 2.1s，断言 30s 留余量）
- fail 行标红 / 数据完整性 / 样式完整性（header/数据区/冻结/筛选）
- 手写 XML 的边界（空 df、纯字符串列、XML 转义、全 fail）
- 向量化值转换与逐值转换逐格等价
- API 端点（to_excel / batch_charts）
"""

import io
import os
import time
import unittest

import pandas as pd
import openpyxl
from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from rest_framework.test import APITestCase

from apps.datafiles.models import DataFile
from apps.datafiles.parsers import get_parser
from apps.analysis.services.statistics import detect_fail_data, get_bin_column_name
from apps.export.export_xlsx_optimized import (
    export_to_xlsx_optimized, _vectorized_native_rows,
)
from apps.export.export_csv import _convert_to_native_type

User = get_user_model()

SAMPLE_DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'Data', 'SampleData')
CTA8280F_PATH = os.path.join(
    SAMPLE_DATA_DIR, 'CTA8280F',
    'DA35_BPC50338_CL08D4.01#AEA3_414A07_2604140567_FT_20260420_164504.csv',
)
GAGE_S1_PATH = os.path.join(SAMPLE_DATA_DIR, 'Gage', 'gage_m_S1.csv')


def _empty_metadata():
    return {'units': {}, 'mins': {}, 'maxs': {}, 'format': 'CTA8290D'}


def _assert_plain_cell(cell, msg=''):
    """默认风格单元格 = 无实底填充（excelize 白底=不写 fill，openpyxl 读为 patternType None）。"""
    assert cell.fill.patternType is None, f'{msg} 应为无填充（白底），实际 {cell.fill.patternType}'


@unittest.skipUnless(
    os.path.exists(CTA8280F_PATH),
    'SampleData/CTA8280F 目录不存在（跳过）',
)
class ToExcelLargeFileTests(TestCase):
    """CTA8280F 大文件（10000 行 × 188 列）导出：耗时 + 样式/完整性。"""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.df, cls.metadata = get_parser('CTA8280F').parse(CTA8280F_PATH)
        cls.fail_indices, _, _ = detect_fail_data(cls.df, cls.metadata)
        # 3 个用例共享一次导出（套件墙钟 3×导出 → 1×导出）
        t0 = time.monotonic()
        cls.buf = export_to_xlsx_optimized(cls.df, cls.metadata)
        cls.export_time = time.monotonic() - t0

    def _load_ws(self):
        return openpyxl.load_workbook(io.BytesIO(self.buf))['Data']

    def test_export_large_file_within_time_budget(self):
        """万行文件导出应在合理时间内完成（原实现 40s+，手写 XML 后 ~2s）。"""
        self.assertGreater(len(self.buf), 0, '导出字节流不应为空')
        self.assertLess(self.export_time, 30, f'大文件导出耗时 {self.export_time:.1f}s，超过 30s 预算')

    def test_export_fail_rows_highlighted(self):
        """fail 行仅 SoftBin 列 + 失败测试项标红；其余单元格（含 Dut_Pass）不红。"""
        ws = self._load_ws()
        data_start_row = 12
        _, _, fail_cells = detect_fail_data(self.df, self.metadata)
        bin_col = get_bin_column_name(self.metadata.get('format', 'CTA8290D'))
        col_letters = {}
        for i, c in enumerate(self.df.columns):
            col_letters[c] = openpyxl.utils.get_column_letter(i + 1)
        for data_idx in list(fail_cells.keys())[:5]:
            row = data_start_row + data_idx
            for col in fail_cells[data_idx]:
                cell = ws[f"{col_letters[col]}{row}"]
                self.assertEqual(cell.fill.patternType, 'solid', f'{col} fail 格应有实底填充')
                self.assertEqual(cell.fill.start_color.rgb.upper(), 'FFFF0000', f'{col} fail 格应为纯红底')
            # 非失败测试项 / 记录列不红：Serial_No（第 1 列）与 Dut_Pass 恒不红
            _assert_plain_cell(ws.cell(row, 1), '记录列 Serial_No 不应标红')
            pass_col = next((c for c in self.df.columns if 'pass' in c.lower()), None)
            if pass_col:
                _assert_plain_cell(
                    ws[f"{col_letters[pass_col]}{row}"], 'Dut_Pass 列不应标红',
                )

    def test_export_data_integrity(self):
        """导出行数与列数、表头正确。"""
        ws = self._load_ws()
        self.assertEqual(ws.max_row, 11 + len(self.df), '表头统计区 11 行 + 全部数据行')
        self.assertEqual(ws.max_column, len(self.df.columns), '列数 = 数据列数（无多余辅助列）')
        self.assertEqual(ws.cell(1, 1).value, self.df.columns[0], '表头应为第一列名')

    def test_export_style_completeness(self):
        """样式完整性（默认风格）：表头/统计区/数据区白底黑字、冻结与筛选。"""
        ws = self._load_ws()
        _assert_plain_cell(ws.cell(1, 1), '表头白底')
        self.assertTrue(ws.cell(1, 1).font.bold, '表头加粗')
        self.assertEqual(ws.cell(1, 1).font.color.rgb.upper(), 'FF000000', '表头黑字')
        _assert_plain_cell(ws.cell(5, 1), '统计区白底')
        # 数据区取第一个非 fail 行（fail 行会被红样式覆盖）
        fail_set = set(self.fail_indices)
        non_fail = next(r for r in range(100) if r not in fail_set)
        _assert_plain_cell(ws.cell(12 + non_fail, 1), '非 fail 数据行应为白底')
        self.assertIsNotNone(ws.freeze_panes, '应存在冻结窗格')
        self.assertIsNotNone(ws.auto_filter.ref, '应存在自动筛选')

    def test_export_column_widths_autofit(self):
        """每列宽度按内容自适应（表头 9 字符的列宽应 > 默认 8.43）。"""
        ws = self._load_ws()
        first_col = self.df.columns[0]
        width = ws.column_dimensions['A'].width
        self.assertIsNotNone(width, '第一列应设置自适应宽度')
        self.assertGreater(width, len(first_col) + 1, '列宽应超过表头文本字符数')
        # 最宽的数据列（数值列宽于表头）也应被撑开
        numeric = [c for c in self.df.columns if self.df[c].dtype in ('int64', 'float64')]
        if numeric:
            idx = list(self.df.columns).index(numeric[0]) + 1
            letter = openpyxl.utils.get_column_letter(idx)
            self.assertGreater(ws.column_dimensions[letter].width or 0, 6)

    def test_export_hidden_columns_preserved(self):
        """hidden_columns 参数：列保留在文件中但设为 Excel 隐藏列。"""
        hidden = [self.df.columns[1]]
        buf = export_to_xlsx_optimized(self.df, self.metadata, hidden_columns=hidden)
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        letter = openpyxl.utils.get_column_letter(2)
        self.assertTrue(ws.column_dimensions[letter].hidden, '指定列应为 Excel 隐藏列')
        self.assertEqual(ws.max_column, len(self.df.columns), '隐藏列数据仍应保留')
        self.assertIsNotNone(ws.cell(12, 2).value, '隐藏列数据仍应存在')

    def test_export_single_sheet_no_default_sheet1(self):
        """导出工作簿只含 Data 一个 sheet（excelize 默认 Sheet1 必须删除）。"""
        wb = openpyxl.load_workbook(io.BytesIO(self.buf))
        self.assertEqual(wb.sheetnames, ['Data'], f'仅应有一个 Data sheet，实际 {wb.sheetnames}')

    def test_export_stats_rows_labels_and_rounding(self):
        """统计行第一列显示 Min/Avg/... 标签；记录列无统计；统计值保留 4 位小数。"""
        from apps.datafiles.parsers.base import SYSTEM_COLUMNS
        ws = self._load_ws()
        labels = [ws.cell(5 + i, 1).value for i in range(6)]
        self.assertEqual(labels, ['Min', 'Avg', 'Max', 'Range', 'STD', 'CPK'],
                         '统计行第一列应为 Min/Avg/Max/Range/STD/CPK 标签')
        # 记录级列（第 1 列 Index_No 等）不参与统计 → 统计区单元格为空
        # （第 1 列为统计行标签 Min/Avg/...，为设计内内容，跳过）
        sys_cols = SYSTEM_COLUMNS.get(self.metadata.get('format', 'CTA8290D'), [])
        for i, col in enumerate(self.df.columns):
            if col in sys_cols and self.df[col].dtype in ('int64', 'float64'):
                if i == 0:
                    continue
                for r in range(5, 11):
                    v = ws.cell(r, i + 1).value
                    self.assertIn(v, (None, ''), f'记录列 {col} 不应有统计值')
        # 测试项统计值保留 4 位小数
        test_numeric = next(
            c for c in self.df.columns
            if self.df[c].dtype in ('int64', 'float64') and c not in sys_cols
        )
        pos = list(self.df.columns).index(test_numeric) + 1
        for r in range(5, 11):
            v = ws.cell(r, pos).value
            if isinstance(v, (int, float)):
                self.assertEqual(
                    round(float(v), 4), float(v),
                    f'{test_numeric} {ws.cell(r, 1).value} 统计值应保留 4 位小数，实际 {v}',
                )


@unittest.skipUnless(
    os.path.exists(GAGE_S1_PATH),
    'SampleData/Gage 目录不存在（跳过）',
)
class ToExcelEdgeCaseTests(TestCase):
    """手写 XML 数据区的边界情况（小数据，快）。"""

    def test_empty_df(self):
        """空 df：不崩，仅 11 行统计区。"""
        df = pd.DataFrame(columns=['a', 'b'])
        buf = export_to_xlsx_optimized(df, _empty_metadata())
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        self.assertEqual(ws.max_row, 11)
        self.assertEqual(ws.max_column, 2)

    def test_single_column_strings_and_escape(self):
        """单列纯字符串 + XML 转义字符读回一致。"""
        df = pd.DataFrame({'text': ['甲', '乙&<c>', '']})
        buf = export_to_xlsx_optimized(df, _empty_metadata())
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        self.assertEqual(ws.cell(12, 1).value, '甲')
        self.assertEqual(ws.cell(13, 1).value, '乙&<c>', 'XML 转义应读回原始文本')
        self.assertIsNone(ws.cell(14, 1).value, '空字符串应导出为空单元格')

    def test_all_fail_rows_red(self):
        """全 fail 小 df：SW_Bin 格标红（无限值测试列 → 仅 bin 格红），数据格白底。"""
        df = pd.DataFrame({'V1': [1.0, 2.0, 3.0], 'SW_Bin': [5, 6, 7]})
        buf = export_to_xlsx_optimized(df, _empty_metadata())
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        for r in (12, 13, 14):
            self.assertEqual(ws.cell(r, 2).fill.start_color.rgb.upper(), 'FFFF0000', 'bin 格应纯红底')
            _assert_plain_cell(ws.cell(r, 1), '无 fail 的数据格保持白底')

    def test_fail_cells_scoped_to_bin_and_test_items(self):
        """标红仅限 SoftBin + 失败测试项；Dut_Pass/其它列不标红（用户截图语义）。"""
        df = pd.DataFrame({
            'Serial_No': [1, 2],
            'Dut_Pass': [True, False],
            'V1': [1.0, 5.0],   # 限 0-2 → 第 2 行越限
            'V2': [1.5, 1.5],   # 正常
            'SW_Bin': [1, 9],
        })
        metadata = {
            'units': {},
            'mins': {'V1': '0', 'V2': '0'},
            'maxs': {'V1': '2', 'V2': '2'},
            'format': 'CTA8290D',
        }
        buf = export_to_xlsx_optimized(df, metadata)
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        # 第 13 行（fail）：SW_Bin(5) 与 V1(3) 红；Serial_No(1)/Dut_Pass(2)/V2(4) 不红
        self.assertEqual(ws.cell(13, 5).fill.start_color.rgb.upper(), 'FFFF0000')
        self.assertEqual(ws.cell(13, 3).fill.start_color.rgb.upper(), 'FFFF0000')
        _assert_plain_cell(ws.cell(13, 1), 'Serial_No 不标红')
        _assert_plain_cell(ws.cell(13, 2), 'Dut_Pass 不标红')
        _assert_plain_cell(ws.cell(13, 4), '未失败测试项不标红')
        # 第 12 行（pass）：无红
        _assert_plain_cell(ws.cell(12, 5))

    def test_orange_on_limit_overlap(self):
        """数据恰好等于上/下限 → 橙底；超限 → 红；其余白底。"""
        df = pd.DataFrame({'V': [1.0, 1.5, 2.0, 2.5], 'SW_Bin': [1, 1, 1, 3]})
        metadata = {'units': {}, 'mins': {'V': '1'}, 'maxs': {'V': '2'}, 'format': 'CTA8290D'}
        buf = export_to_xlsx_optimized(df, metadata)
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        self.assertEqual(ws.cell(12, 1).fill.start_color.rgb.upper(), 'FFFFC000', '== Min → 橙底')
        _assert_plain_cell(ws.cell(13, 1), '区间内 → 白底')
        self.assertEqual(ws.cell(14, 1).fill.start_color.rgb.upper(), 'FFFFC000', '== Max → 橙底')
        self.assertEqual(ws.cell(15, 1).fill.start_color.rgb.upper(), 'FFFF0000', '> Max → 红底')

    def test_vectorized_native_equiv(self):
        """_vectorized_native_rows 与逐值 _convert_to_native_type 逐格等价。"""
        df = pd.DataFrame({
            'i': [1, 2, None],
            'f': [1.5, float('nan'), -0.25],
            'b': [True, False, True],
            's': ['a', '乙&<', None],
            'm': [1, 'x', 3.14],
        })
        rows = _vectorized_native_rows(df)
        expected = [
            [_convert_to_native_type(v) for v in df.iloc[i].tolist()]
            for i in range(len(df))
        ]
        self.assertEqual(rows, expected)


@unittest.skipUnless(
    os.path.exists(GAGE_S1_PATH),
    'SampleData/Gage 目录不存在（跳过）',
)
class ExportApiTests(APITestCase):
    """导出 API 端点测试。"""

    def setUp(self):
        self.user = User.objects.create_user(username='exporter', password='pw')
        self.client.force_authenticate(self.user)
        self.datafile = DataFile.objects.create(
            owner=self.user,
            filename='gage_m_S1.csv',
            file_path=GAGE_S1_PATH,
            file_size=os.path.getsize(GAGE_S1_PATH),
            format_type='CTA8290D',
            status='ready',
        )

    def test_to_excel_api(self):
        """POST /export/to_excel/：200 + xlsx 可解析 + 行数列数正确。"""
        resp = self.client.post('/api/v1/export/to_excel/', {'file_id': self.datafile.id})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', resp['Content-Disposition'])
        df_parsed, _ = get_parser('CTA8290D').parse(GAGE_S1_PATH)
        buf = b''.join(resp.streaming_content)
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Data']
        self.assertEqual(ws.max_row, 11 + len(df_parsed), '数据行数 = 文件行数')
        self.assertEqual(ws.max_column, len(df_parsed.columns))

    def test_batch_charts_api(self):
        """POST /export/batch_charts/：200 + 总览 sheet 行数 = 参数数 + 1。"""
        df_parsed, _ = get_parser('CTA8290D').parse(GAGE_S1_PATH)
        numeric_cols = [c for c in df_parsed.columns if df_parsed[c].dtype in ('int64', 'float64')][:2]
        self.assertGreaterEqual(len(numeric_cols), 2, 'GAGE 数据应至少 2 个数值列')
        resp = self.client.post('/api/v1/export/batch_charts/', {
            'file_id': self.datafile.id,
            'params': numeric_cols,
        }, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        buf = b''.join(resp.streaming_content)
        wb = openpyxl.load_workbook(io.BytesIO(buf))
        self.assertEqual(wb['总览'].max_row, 1 + len(numeric_cols), '总览表头 + 每参数一行')

    def test_batch_charts_site_stats_string_failcount(self):
        """站点统计 FailCount 是展示字符串（`3(0.181%)`）时 batch_charts 不抛 TypeError。

        2026-08-13 站点统计改版后 FailCount 为格式化字符串 + 数字字段 FailCountNum；
        总览 Yield 单元格的 fail 判定必须用 FailCountNum（旧代码 `FailCount > 0` 字符串
        与 int 比较 → 500）。断言 ALL Site 行 Yield 单元格应用 fail 红底（F54927）。
        """
        from apps.export.export_batch_charts_xlsx import build_batch_charts_xlsx_with_charts

        df = pd.DataFrame({
            'Site_No': [1, 1, 2, 2],
            # 限 1.4~1.6：Site1 两行全部越限 → ALL Site FailCountNum = 2 > 0
            'Vth': [1.0, 1.1, 2.0, 2.5],
        })
        # 限值必须放在 metadata['mins']/['maxs']——这才是 compute_range_statistics
        # 读取的键（unit 同理是 'units'）。旧写法用 {'limits': {...}}，而 export 侧
        # 根本不消费该键 → rdl 退化成幻影 (0.0, 0.0)，于是「所有 Vth > 0」全被判
        # fail，红底断言碰巧通过——它验证的是幻影限值 bug 而不是站点统计。
        # 改成真实限值后 Site1 的 1.0/1.1 < LSL 1.4 是**真**越限（Site2 的 2.0/2.5
        # > USL 1.6 也是），ALL Site FailCountNum = 4，红底断言才有意义。
        metadata = {'mins': {'Vth': '1.4'}, 'maxs': {'Vth': '1.6'},
                    'units': {'Vth': ''}}
        buf = build_batch_charts_xlsx_with_charts(
            df, metadata, ['Vth'], site_col='Site_No',
            show_limit=True, show_3sigma=False, show_4sigma=False,
            show_6sigma=False, show_normal=False,
        )
        wb = openpyxl.load_workbook(io.BytesIO(buf))
        # 总览第 2 行（参数 Vth）第 10 列（ALL Site Yield）应为 fail 红底
        cell = wb['总览'].cell(row=2, column=10)
        self.assertEqual(cell.fill.start_color.rgb, '00F54927', '有 fail 的 ALL Site 行应为红色填充')

    def test_batch_charts_data_only_bin1(self):
        """POST /export/batch_charts/ + data_only_bin1：200，图表仍按过滤后数据生成。"""
        df_parsed, _ = get_parser('CTA8290D').parse(GAGE_S1_PATH)
        numeric_cols = [c for c in df_parsed.columns if df_parsed[c].dtype in ('int64', 'float64')][:2]
        resp = self.client.post('/api/v1/export/batch_charts/', {
            'file_id': self.datafile.id,
            'params': numeric_cols,
            'data_only_bin1': True,
        }, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        buf = b''.join(resp.streaming_content)
        wb = openpyxl.load_workbook(io.BytesIO(buf))
        self.assertEqual(wb['总览'].max_row, 1 + len(numeric_cols), '总览表头 + 每参数一行')

    def test_sigma_limit_data_only_bin1(self):
        """POST /export/sigma_limit/ + data_only_bin1：200 xlsx，不因行过滤崩溃。"""
        resp = self.client.post('/api/v1/export/sigma_limit/', {
            'file_id': self.datafile.id,
            'sigma': 3,
            'data_only_bin1': True,
        }, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        buf = b''.join(resp.streaming_content)
        self.assertGreater(len(buf), 0)


def _parse_content_disposition(cd: str) -> str:
    """从 Content-Disposition 提取 RFC 5987 filename*= 或 filename= 的文件名。"""
    import re as _re
    star = _re.search(r"filename\*\s*=\s*(?:UTF-8'')?([^;]+)", cd)
    if star:
        return _re.sub(r'^"|"$', '', star.group(1).strip()).replace('%20', ' ')
    plain = _re.search(r'filename\s*=\s*"?([^";]+)"?', cd)
    return plain.group(1).strip() if plain else ''


@unittest.skipUnless(
    os.path.exists(GAGE_S1_PATH),
    'SampleData/Gage 目录不存在（跳过）',
)
class ExportFilenameTemplateApiTests(APITestCase):
    """导出文件名自定义模板端到端（设置 → 导出 → Content-Disposition）。"""

    def setUp(self):
        self.user = User.objects.create_user(username='tpluser', password='pw')
        self.client.force_authenticate(self.user)
        self.datafile = DataFile.objects.create(
            owner=self.user,
            filename='gage_m_S1.csv',
            file_path=GAGE_S1_PATH,
            file_size=os.path.getsize(GAGE_S1_PATH),
            format_type='CTA8290D',
            status='ready',
        )

    def _set_template(self, key: str, template: str):
        resp = self.client.put('/api/v1/auth/settings/',
                               {'export_filename_templates': {key: template}},
                               format='json')
        self.assertEqual(resp.status_code, 200, resp.data)

    def test_to_excel_custom_template_with_datetime(self):
        self._set_template('to_excel', '{filename}_{datetime}')
        resp = self.client.post('/api/v1/export/to_excel/', {'file_id': self.datafile.id})
        self.assertEqual(resp.status_code, 200)
        name = _parse_content_disposition(resp['Content-Disposition'])
        self.assertRegex(name, r'^gage_m_S1_\d{8}_\d{6}\.xlsx$')

    def test_to_csv_default_template(self):
        resp = self.client.post('/api/v1/export/to_csv/', {'file_id': self.datafile.id})
        self.assertEqual(resp.status_code, 200)
        name = _parse_content_disposition(resp['Content-Disposition'])
        self.assertEqual(name, 'gage_m_S1_data.csv')

    def test_sigma_limit_sigma_variable(self):
        self._set_template('sigma_limit', '{filename}_{sigma}sigma')
        resp = self.client.post('/api/v1/export/sigma_limit/',
                                {'file_id': self.datafile.id, 'sigma': 6}, format='json')
        self.assertEqual(resp.status_code, 200)
        name = _parse_content_disposition(resp['Content-Disposition'])
        self.assertEqual(name, 'gage_m_S1_6sigma.xlsx')

    def test_batch_charts_both_formats(self):
        self._set_template('batch_charts', '{filename}_charts')
        for fmt in ('xlsx', 'pptx'):
            resp = self.client.post('/api/v1/export/batch_charts/',
                                    {'file_id': self.datafile.id, 'params': [],
                                     'format': fmt}, format='json')
            self.assertEqual(resp.status_code, 200, fmt)
            name = _parse_content_disposition(resp['Content-Disposition'])
            self.assertEqual(name, f'gage_m_S1_charts.{fmt}')

    def test_html_report_chinese_filename_encoded(self):
        """中文源文件名 → FileResponse 自动 RFC 5987 编码（回归：手写 header 会
        UnicodeEncodeError，且会带上 .csv 扩展名）。"""
        self.datafile.filename = '测试_文件.csv'
        self.datafile.save()
        self._set_template('html_report', '{filename}_report')
        resp = self.client.post('/api/v1/export/html_report/', {'file_id': self.datafile.id})
        self.assertEqual(resp.status_code, 200)
        cd = resp['Content-Disposition']
        self.assertIn("filename*=", cd)
        star = cd.split("filename*=")[1].split("''", 1)[1].split(';')[0].strip()
        import urllib.parse
        decoded = urllib.parse.unquote(star)
        self.assertEqual(decoded, '测试_文件_report.html')

    def test_sanitize_applied_in_response(self):
        self._set_template('to_excel', '{filename}:bad?')
        resp = self.client.post('/api/v1/export/to_excel/', {'file_id': self.datafile.id})
        self.assertEqual(resp.status_code, 200)
        name = _parse_content_disposition(resp['Content-Disposition'])
        self.assertEqual(name, 'gage_m_S1_bad_.xlsx')


class BuildHistogramBinsTests(SimpleTestCase):
    """build_histogram_bins：导出图分箱必须与屏幕直方图一致（/20）。

    回归：export_ppt 曾用 /25 分箱，PPT 直方图与用户审阅的画面 bin 宽度
    不一致；统一收敛到 charts.build_histogram_bins（与 safe_gap /20 同）。
    """

    def test_26_edges_with_gap_of_range_over_20(self):
        from apps.export.charts import build_histogram_bins
        bins, gap = build_histogram_bins(10.0, 30.0)
        self.assertEqual(len(bins), 26)
        self.assertAlmostEqual(gap, 1.0)              # (30-10)/20
        self.assertAlmostEqual(bins[0], 10.0 - 2.5)   # 10 - 2.5*gap
        self.assertAlmostEqual(bins[-1], 10.0 + 22.5)  # 10 - 2.5*gap + 25*gap

    def test_zero_width_range_falls_back_to_gap_1(self):
        from apps.export.charts import build_histogram_bins
        bins, gap = build_histogram_bins(10.0, 10.0)
        self.assertEqual(gap, 1.0)
        self.assertEqual(len(bins), 26)
        self.assertEqual(bins[-1], bins[0] + 25.0)
