"""Native Excel chart builders for batch_charts xlsx export."""

from typing import List, Optional, Tuple

from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import NumericAxis
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLORS_SITE_8 = [
    "E53935", "1E88E5", "43A047", "F9A825", "8E24AA", "00ACC1", "F57C00", "D81B60",
]
COLOR_LSL = "C62828"
COLOR_USL = "C62828"
COLOR_SIGMA_3 = "1565C0"
COLOR_SIGMA_4 = "00838F"
COLOR_SIGMA_6 = "E65100"
COLOR_NORMAL = "F57F17"
COLOR_KDE = "7B1FA2"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

FAIL_FILL = PatternFill(start_color="F54927", end_color="F54927", fill_type="solid")
FAIL_FONT = Font(color="FFFFFF", bold=True)
ALL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
NORMAL_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
STATS_LABEL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
RETURN_HYPERLINK_FONT = Font(name="Calibri", size=11, color="0563C1")

CPK_COLOR_MAP = {
    "green": ("4CAF50", "FFFFFF"),
    "orange": ("FFA726", "FFFFFF"),
    "darkorange": ("FF7043", "FFFFFF"),
    "red": ("F44336", "FFFFFF"),
}


def _safe_sheet_name(name: str) -> str:
    safe = name.replace("/", "_").replace("\\", "_").replace(" ", "_").replace("-", "_")
    return safe[:31]


def _apply_cpk_style(cell, cpk_color_name: str):
    bg, fg = CPK_COLOR_MAP.get(cpk_color_name, ("9E9E9E", "FFFFFF"))
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_data_sheet(
    ws,
    centers,
    site_labels: List[str],
    site_percents: List[List[float]],
    all_pct: List[float],
    normal_pct: Optional[List[float]],
    kde_pct: Optional[List[float]],
    refs: List[Tuple[str, float, str]],
):
    """Write hidden chart-data sheet. refs = [(name, x_value, color), ...]."""
    headers = ["BinCenter"] + site_labels + ["ALL Site"]
    if normal_pct:
        headers.append("Normal")
    if kde_pct:
        headers.append("KDE")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    n = len(centers)
    for i in range(n):
        row = i + 2
        ws.cell(row=row, column=1, value=float(centers[i]))
        col = 2
        for pct_list in site_percents:
            ws.cell(row=row, column=col, value=pct_list[i])
            col += 1
        ws.cell(row=row, column=col, value=all_pct[i])
        col += 1
        if normal_pct:
            ws.cell(row=row, column=col, value=normal_pct[i])
            col += 1
        if kde_pct:
            ws.cell(row=row, column=col, value=kde_pct[i])
            col += 1

    ref_col = len(headers) + 2
    ws.cell(row=1, column=ref_col, value="RefName")
    ws.cell(row=1, column=ref_col + 1, value="RefX")
    ws.cell(row=1, column=ref_col + 2, value="RefY")
    for i, (name, xval, _color) in enumerate(refs):
        row = i + 2
        ws.cell(row=row, column=ref_col, value=name)
        ws.cell(row=row, column=ref_col + 1, value=xval)
        ws.cell(row=row, column=ref_col + 2, value=0)


def _make_bar_chart(data_ws, n_bins: int, n_site_cols: int) -> BarChart:
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.style = 10
    bar.y_axis.title = "百分比 (%)"
    bar.x_axis.title = "Bin center"
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 100

    first_site_col = 2
    last_site_col = first_site_col + n_site_cols
    data_ref = Reference(
        data_ws,
        min_col=first_site_col,
        max_col=last_site_col,
        min_row=1,
        max_row=n_bins + 1,
    )
    cats = Reference(data_ws, min_col=1, min_row=2, max_row=n_bins + 1)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)

    for i, s in enumerate(bar.series):
        color = COLORS_SITE_8[i % len(COLORS_SITE_8)]
        s.graphicalProperties.solidFill = color

    return bar


def _make_line_chart(
    data_ws, n_bins: int, n_site_cols: int, has_normal: bool, has_kde: bool
) -> Optional[LineChart]:
    if not (has_normal or has_kde):
        return None

    line = LineChart()
    line.style = 10
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 100

    col = n_site_cols + 3
    cols = []
    if has_normal:
        cols.append((col, "Normal", COLOR_NORMAL))
        col += 1
    if has_kde:
        cols.append((col, "KDE", COLOR_KDE))

    cats = Reference(data_ws, min_col=1, min_row=2, max_row=n_bins + 1)
    for col_idx, name, color in cols:
        values = Reference(data_ws, min_col=col_idx, min_row=1, max_row=n_bins + 1)
        s = Series(values, title=name)
        s.marker = Marker(symbol="none")
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 28575
        line.series.append(s)
    line.set_categories(cats)

    return line


def _make_scatter_chart(
    data_ws,
    n_bins: int,
    n_site_cols: int,
    has_normal: bool,
    has_kde: bool,
    refs: List[Tuple[str, float, str]],
    x_min: float,
    x_max: float,
) -> Optional[ScatterChart]:
    if not refs:
        return None

    ref_col = n_site_cols + 1 + int(has_normal) + int(has_kde) + 2

    scatter = ScatterChart()
    scatter.x_axis = NumericAxis()
    scatter.x_axis.axId = 20
    scatter.x_axis.crosses = "max"
    scatter.x_axis.scaling.min = x_min
    scatter.x_axis.scaling.max = x_max
    scatter.x_axis.delete = True
    # scatter 与 bar 图表共享主 Y 轴（bar 的 y_axis 默认 axId=100）
    scatter.y_axis = NumericAxis()
    scatter.y_axis.axId = 100
    scatter.y_axis.scaling.min = 0
    scatter.y_axis.scaling.max = 100

    for i, (name, xval, color) in enumerate(refs):
        row = i + 2
        xvals = Reference(data_ws, min_col=ref_col + 1, min_row=row, max_row=row)
        yvals = Reference(data_ws, min_col=ref_col + 2, min_row=row, max_row=row)
        s = Series(yvals, xvals, title=name)
        s.marker = Marker(symbol="none")
        s.graphicalProperties.line.noFill = True

        line_prop = LineProperties(solidFill=color, w=20000)
        err = ErrorBars(
            errDir="y",
            errBarType="plus",
            errValType="fixedVal",
            val=100,
            noEndCap=True,
            spPr=GraphicalProperties(ln=line_prop),
        )
        s.errBars = err
        scatter.series.append(s)

    return scatter


def _build_ref_lines(
    show_limit: bool,
    show_3sigma: bool,
    show_4sigma: bool,
    show_6sigma: bool,
    rdl_min,
    rdl_max,
    mean_val: float,
    std_val: float,
) -> List[Tuple[str, float, str]]:
    refs: List[Tuple[str, float, str]] = []
    if show_limit and rdl_min is not None:
        refs.append(("LSL", rdl_min, COLOR_LSL))
    if show_limit and rdl_max is not None:
        refs.append(("USL", rdl_max, COLOR_USL))
    if std_val <= 0:
        return refs
    for sigma, flag, color, label in [
        (3, show_3sigma, COLOR_SIGMA_3, "-3σ"),
        (4, show_4sigma, COLOR_SIGMA_4, "-4σ"),
        (6, show_6sigma, COLOR_SIGMA_6, "-6σ"),
    ]:
        if flag:
            refs.append((label, mean_val - sigma * std_val, color))
    for sigma, flag, color, label in [
        (3, show_3sigma, COLOR_SIGMA_3, "+3σ"),
        (4, show_4sigma, COLOR_SIGMA_4, "+4σ"),
        (6, show_6sigma, COLOR_SIGMA_6, "+6σ"),
    ]:
        if flag:
            refs.append((label, mean_val + sigma * std_val, color))
    return refs


def _write_param_sheet(
    wb,
    title: str,
    stats_data: dict,
    site_stats_list: List[dict],
    data_sheet_name: str,
    n_bins: int,
    n_site_cols: int,
    has_normal: bool,
    has_kde: bool,
    refs: List[Tuple[str, float, str]],
    x_min: float,
    x_max: float,
):
    """Create parameter sheet with stats tables and native chart."""
    param_safe = _safe_sheet_name(title)
    ws = wb.create_sheet(title=param_safe)

    header_row = ["统计项", "Low Limit", "High Limit", "Unit"]
    for col_idx, h in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    if stats_data:
        ws.cell(row=2, column=1, value=stats_data.get("param_name", ""))
        low_limit = stats_data.get("low_limit", 0)
        high_limit = stats_data.get("high_limit", 0)
        if low_limit != "N/A":
            ws.cell(row=2, column=2, value=low_limit).number_format = "0.0000"
        else:
            ws.cell(row=2, column=2, value=low_limit)
        if high_limit != "N/A":
            ws.cell(row=2, column=3, value=high_limit).number_format = "0.0000"
        else:
            ws.cell(row=2, column=3, value=high_limit)
        ws.cell(row=2, column=4, value=stats_data.get("unit", ""))

        left_labels = [
            ("Mean", "mean_val", 4, 1),
            ("STD", "std_val", 5, 1),
            ("Range", "data_range", 6, 1),
            ("数据点数", "count", 7, 1),
            ("CPK", "cpk_str", 8, 1),
        ]
        for label, key, row, col in left_labels:
            ws.cell(row=row, column=col, value=label).fill = STATS_LABEL_FILL
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            val = stats_data.get(key, "")
            cell = ws.cell(row=row, column=col + 1, value=str(val))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "cpk_str":
                _apply_cpk_style(cell, stats_data.get("cpk_color", "gray"))
            elif key in ("mean_val", "std_val"):
                cell.number_format = "0.0000"

    if site_stats_list:
        ws.cell(row=1, column=6, value="Site统计").fill = HEADER_FILL
        ws.cell(row=1, column=6).font = HEADER_FONT
        ws.cell(row=1, column=6).alignment = HEADER_ALIGNMENT

        site_headers = ["Site", "Yield", "FailCount", "ExceedMin", "ExceedMax"]
        for col_idx, h in enumerate(site_headers, 1):
            cell = ws.cell(row=2, column=col_idx + 5, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        for row_offset, site_info in enumerate(site_stats_list):
            is_fail = site_info.get("FailCount", 0) > 0
            is_all = site_info.get("Site") == "ALL Site"
            for col_idx, key in enumerate(["Site", "Yield", "FailCount", "ExceedMin", "ExceedMax"], 1):
                cell = ws.cell(row=row_offset + 3, column=col_idx + 5, value=site_info.get(key, ""))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_fail:
                    cell.fill = FAIL_FILL
                    cell.font = FAIL_FONT
                elif is_all:
                    cell.fill = ALL_FILL
                else:
                    cell.fill = NORMAL_FILL

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for col in range(6, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    return_cell = ws.cell(row=8, column=5, value="← 返回总览")
    return_cell.hyperlink = "#'总览'!A1"
    return_cell.font = RETURN_HYPERLINK_FONT

    data_ws = wb[data_sheet_name]
    bar = _make_bar_chart(data_ws, n_bins, n_site_cols)
    line = _make_line_chart(data_ws, n_bins, n_site_cols, has_normal, has_kde)
    scatter = _make_scatter_chart(
        data_ws, n_bins, n_site_cols, has_normal, has_kde, refs, x_min, x_max
    )

    if line:
        bar += line
    if scatter:
        bar += scatter

    bar.title = stats_data.get("param_name", title)
    bar.width = 20
    bar.height = 12
    ws.add_chart(bar, "A10")
