"""Batch charts Excel export with native Excel charts (editable, no PNG)."""

import io
import math
from typing import List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.analysis.services.statistics import (
    compute_cpk,
    compute_range_statistics,
    compute_site_stats,
    filter_finite,
    get_1d_from,
)
from apps.analysis.services.statistics.kde import GaussianKDE

from .native_chart_builders import (
    _build_ref_lines,
    _safe_sheet_name,
    _write_data_sheet,
    _write_param_sheet,
)

N_BINS = 25

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

FAIL_FILL = PatternFill(start_color="F54927", end_color="F54927", fill_type="solid")
FAIL_FONT = Font(color="FFFFFF", bold=True)
NORMAL_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
SUMMARY_HYPERLINK_FONT = Font(name="Calibri", size=11, color="0563C1")

CPK_COLOR_MAP = {
    "green": ("4CAF50", "FFFFFF"),
    "orange": ("FFA726", "FFFFFF"),
    "darkorange": ("FF7043", "FFFFFF"),
    "red": ("F44336", "FFFFFF"),
}


def _get_text_width(text) -> int:
    s = str(text) if text is not None else ""
    cn_len = len([c for c in s if "一" <= c <= "鿿"])
    return len(s) + cn_len + 2


def _apply_cpk_style(cell, cpk_color_name: str):
    bg, fg = CPK_COLOR_MAP.get(cpk_color_name, ("9E9E9E", "FFFFFF"))
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_histogram_bins(low: float, high: float) -> Tuple[np.ndarray, float]:
    """Same binning strategy as charts.build_histogram_bins (avoid matplotlib import)."""
    data_gap = (high - low) / 20 if (high - low) > 0 else 1.0
    bin_start = low - 2.5 * data_gap
    bins = np.array([bin_start + j * data_gap for j in range(26)])
    centers = (bins[:-1] + bins[1:]) / 2
    return centers, data_gap


def _compute_site_histograms(
    data_series: np.ndarray,
    site_series: Optional[np.ndarray],
    site_values: Optional[Sequence],
    bin_edges: np.ndarray,
) -> Tuple[List[str], List[List[float]], List[float]]:
    """Return (site_labels, per_site_percentages[site][bin], all_site_percentages)."""
    site_labels = []
    site_percents = []

    if site_series is not None and site_values:
        param_ser = pd.to_numeric(pd.Series(data_series), errors="coerce")
        grouped = pd.DataFrame({"site": pd.Series(site_series), "val": param_ser}).groupby("site")
        for site in site_values:
            site_labels.append(f"Site{site}")
            if site in grouped.groups:
                sdata = grouped.get_group(site)["val"].dropna()
            else:
                sdata = pd.Series(dtype=float)
            total = len(sdata)
            hist, _ = np.histogram(sdata, bins=bin_edges)
            pct = [round((count / total) * 100, 2) if total > 0 else 0.0 for count in hist]
            site_percents.append(pct)

    clean = pd.to_numeric(pd.Series(data_series), errors="coerce").dropna()
    total_all = len(clean)
    hist_all, _ = np.histogram(clean, bins=bin_edges)
    all_pct = [round((count / total_all) * 100, 2) if total_all > 0 else 0.0 for count in hist_all]

    return site_labels, site_percents, all_pct


def _normal_values_at_centers(mean_val: float, std_val: float, centers: np.ndarray) -> Optional[List[float]]:
    if std_val <= 0:
        return None
    scale = 1.0 / (std_val * math.sqrt(2.0 * math.pi))
    y = scale * np.exp(-0.5 * ((centers - mean_val) / std_val) ** 2)
    max_y = float(np.max(y)) if np.any(y) else 0.0
    if max_y <= 0:
        return None
    return [round(float(v) / max_y * 100, 2) for v in y]


def _kde_values_at_centers(data_series: np.ndarray, centers: np.ndarray) -> Optional[List[float]]:
    try:
        kde_vals = np.asarray(data_series, dtype=float)
        if len(kde_vals) < 3 or np.ptp(kde_vals) <= 0:
            return None
        kde = GaussianKDE(kde_vals, bw_method="silverman")
        kde_values = kde(centers)
        max_kde = float(np.max(kde_values))
        if max_kde <= 0:
            return None
        return [round(float(v) / max_kde * 100, 2) for v in kde_values]
    except Exception:
        return None


def build_batch_charts_xlsx_native(
    df: pd.DataFrame,
    metadata: dict,
    params: List[str],
    site_col: Optional[str] = None,
    show_limit: bool = True,
    show_3sigma: bool = False,
    show_4sigma: bool = False,
    show_6sigma: bool = True,
    show_normal: bool = False,
    show_kde: bool = False,
) -> bytes:
    """Build batch charts Excel with native Excel charts. Returns xlsx bytes."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "总览"

    summary_headers = ["序号", "参数", "数据点数", "Mean", "STD", "Min", "Max", "CPK", "CPK Level", "ALL Site Yield"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    col_max_widths = {i: _get_text_width(summary_headers[i - 1]) for i in range(1, len(summary_headers) + 1)}

    param_order = {col: idx for idx, col in enumerate(df.columns)}
    params_sorted = sorted(params, key=lambda x: param_order.get(x, 999999))

    processed_params = []
    summary_data_list = []

    for selected_param in params_sorted:
        if selected_param not in df.columns:
            continue
        data_series = filter_finite(get_1d_from(df, selected_param))
        if len(data_series) == 0:
            continue

        stats = compute_range_statistics(data_series, metadata, selected_param)
        cpk_result = compute_cpk(stats["mean"], stats["std"], stats["rdl"][0], stats["rdl"][1])
        mean_val = stats["mean"]
        std_val = stats["std"]
        rdl_min = stats["rdl"][0]
        rdl_max = stats["rdl"][1]

        site_stats = []
        if site_col and site_col in df.columns:
            site_series = get_1d_from(df, selected_param)
            site_idx = get_1d_from(df, site_col)
            site_stats = compute_site_stats(site_series, site_idx, rdl_min, rdl_max, None, None, False) or []

        processed_params.append(selected_param)
        summary_data_list.append({
            "stats_data": {
                "param_name": selected_param,
                "mean_val": round(mean_val, 4),
                "std_val": round(std_val, 4),
                "data_range": f"{round(float(data_series.min()), 4)} ~ {round(float(data_series.max()), 4)}",
                "count": len(data_series),
                "cpk_str": round(cpk_result["cpk"], 4),
                "cpk_color": cpk_result.get("cpk_color", "gray"),
                "low_limit": round(rdl_min, 4) if rdl_min is not None else "N/A",
                "high_limit": round(rdl_max, 4) if rdl_max is not None else "N/A",
                "unit": stats.get("unit", ""),
            },
            "site_stats": site_stats,
            "chart_data": {
                "data_series": data_series,
                "mean_val": mean_val,
                "std_val": std_val,
                "rdl_min": rdl_min,
                "rdl_max": rdl_max,
            },
        })

    for param_idx, (selected_param, entry) in enumerate(zip(processed_params, summary_data_list), 1):
        stats_data = entry["stats_data"]
        site_stats_list = entry["site_stats"]

        param_safe = _safe_sheet_name(selected_param)
        cell_param = ws_summary.cell(row=param_idx + 1, column=2, value=selected_param)
        cell_param.hyperlink = f"#'{param_safe}'!A1"
        cell_param.font = SUMMARY_HYPERLINK_FONT

        ws_summary.cell(row=param_idx + 1, column=1, value=param_idx).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=3, value=stats_data["count"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=4, value=str(stats_data["mean_val"])).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=5, value=str(stats_data["std_val"])).alignment = Alignment(horizontal="center")

        range_parts = str(stats_data["data_range"]).split(" ~ ")
        ws_summary.cell(row=param_idx + 1, column=6, value=range_parts[0] if len(range_parts) > 0 else "").alignment = Alignment(horizontal="center")
        ws_summary.cell(row=param_idx + 1, column=7, value=range_parts[1] if len(range_parts) > 1 else "").alignment = Alignment(horizontal="center")

        cell_cpk = ws_summary.cell(row=param_idx + 1, column=8, value=stats_data["cpk_str"])
        _apply_cpk_style(cell_cpk, stats_data.get("cpk_color", "gray"))
        ws_summary.cell(row=param_idx + 1, column=9, value=str(stats_data["cpk_str"])).alignment = Alignment(horizontal="center")

        all_site_yield = "100.00%"
        for si in site_stats_list:
            if si.get("Site") == "ALL Site":
                all_site_yield = si.get("Yield", "100.00%")
                break
        cell_yield = ws_summary.cell(row=param_idx + 1, column=10, value=all_site_yield)
        cell_yield.alignment = Alignment(horizontal="center")
        for si in site_stats_list:
            if si.get("Site") == "ALL Site":
                if si.get("FailCount", 0) > 0:
                    cell_yield.fill = FAIL_FILL
                    cell_yield.font = FAIL_FONT
                else:
                    cell_yield.fill = NORMAL_FILL
                break

        row_vals = [
            (1, param_idx), (2, selected_param), (3, stats_data["count"]),
            (4, str(stats_data["mean_val"])), (5, str(stats_data["std_val"])),
            (6, range_parts[0] if len(range_parts) > 0 else ""),
            (7, range_parts[1] if len(range_parts) > 1 else ""),
            (8, str(stats_data["cpk_str"])), (9, str(stats_data["cpk_str"])), (10, all_site_yield),
        ]
        for col_i, val in row_vals:
            w = _get_text_width(val)
            if w > col_max_widths.get(col_i, 0):
                col_max_widths[col_i] = w

    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = min(col_max_widths.get(col, 15), 40)

    site_series = None
    site_values = None
    if site_col and site_col in df.columns:
        site_series_raw = df[site_col]
        if isinstance(site_series_raw, pd.DataFrame):
            site_series_raw = site_series_raw.iloc[:, 0]
        site_values = sorted(site_series_raw.dropna().unique(), key=lambda x: (isinstance(x, float), x))
        site_series = site_series_raw.to_numpy()

    for title, entry in zip(processed_params, summary_data_list):
        stats_data = entry["stats_data"]
        site_stats_list = entry["site_stats"]
        chart_data = entry["chart_data"]

        data_series = chart_data["data_series"]
        mean_val = chart_data["mean_val"]
        std_val = chart_data["std_val"]
        rdl_min = chart_data["rdl_min"]
        rdl_max = chart_data["rdl_max"]

        param_low_limit = rdl_min if rdl_min is not None else float(data_series.min())
        param_high_limit = rdl_max if rdl_max is not None else float(data_series.max())
        centers, data_gap = _build_histogram_bins(param_low_limit, param_high_limit)
        half_gap = data_gap / 2
        bin_edges = np.array([centers[0] - half_gap + i * data_gap for i in range(N_BINS + 1)])

        site_labels, site_percents, all_pct = _compute_site_histograms(
            data_series.to_numpy(), site_series, site_values, bin_edges
        )

        normal_pct = _normal_values_at_centers(mean_val, std_val, centers) if show_normal else None
        kde_pct = _kde_values_at_centers(data_series.to_numpy(), centers) if show_kde else None

        refs = _build_ref_lines(
            show_limit, show_3sigma, show_4sigma, show_6sigma,
            rdl_min, rdl_max, mean_val, std_val,
        )

        param_safe = _safe_sheet_name(title)
        data_sheet_name = _safe_sheet_name(title + "_data")[:31]
        data_ws = wb.create_sheet(title=data_sheet_name)
        _write_data_sheet(
            data_ws, centers, site_labels, site_percents, all_pct,
            normal_pct, kde_pct, refs,
        )
        data_ws.sheet_state = "hidden"

        _write_param_sheet(
            wb, title, stats_data, site_stats_list, data_sheet_name,
            N_BINS, len(site_labels), bool(normal_pct), bool(kde_pct),
            refs, float(centers[0]), float(centers[-1]),
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    result = output.getvalue()
    output.close()
    return result
