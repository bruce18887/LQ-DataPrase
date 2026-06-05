import io
import re
from collections import defaultdict

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from django.http import FileResponse

from apps.datafiles.models import DataFile
from apps.datafiles.services import get_cached_parsed_file
from apps.analysis.services.statistics import (
    calculate_fail_bin_statistics, get_site_column,
    get_bin_column_name, compute_site_yield_data,
)


def _build_phase_summary(phases):
    """Aggregate phases by base phase name (e.g. CP1, FT1, RT1, QA1).
    CP phases: each wafer is independent, so summary shows CP1 aggregate.
    FT/QA/RT: multiple files per phase, aggregate totals.
    """
    groups = defaultdict(lambda: {'total': 0, 'pass': 0, 'fail': 0, 'files': []})
    for p in phases:
        key = p['phase']  # e.g. CP1, FT1, QA1, RT1
        groups[key]['total'] += p['total']
        groups[key]['pass'] += p['pass_count']
        groups[key]['fail'] += p['fail_count']
        groups[key]['files'].append(p['filename'])

    summary = []
    for phase_name, data in groups.items():
        yield_pct = round(data['pass'] / data['total'] * 100, 2) if data['total'] > 0 else 0
        summary.append({
            'phase': phase_name,
            'file_count': len(data['files']),
            'total': data['total'],
            'pass_count': data['pass'],
            'fail_count': data['fail'],
            'yield_pct': yield_pct,
        })

    # Sort by same phase order
    PHASE_ORDER = {'CP': 0, 'FT': 1, 'BF': 2, 'RT': 3, 'QA': 4, '2D3D': 5, 'UNKNOWN': 99}
    def sort_key(s):
        for prefix, order in PHASE_ORDER.items():
            if s['phase'].startswith(prefix):
                m = re.search(r'(\d+)', s['phase'])
                return (order, int(m.group(1)) if m else 0)
        return (99, 0)
    summary.sort(key=sort_key)
    return summary


class BatchReportViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def list_batches(self, request):
        batches = (
            DataFile.objects.filter(owner=request.user, file_type='batch')
            .exclude(batch_name='')
            .values('batch_name')
            .distinct()
            .order_by('-batch_name')
        )
        result = []
        for b in batches:
            name = b['batch_name']
            count = DataFile.objects.filter(
                owner=request.user, file_type='batch', batch_name=name
            ).count()
            result.append({'batch_name': name, 'count': count})
        return Response({'batches': result})

    @action(detail=False, methods=['get'])
    def batch_yield_data(self, request):
        """Return computed yield data for a batch of imported files."""
        batch_name = request.query_params.get('batch_name', '')
        if not batch_name:
            return Response({'error': 'batch_name required'}, status=400)

        files = DataFile.objects.filter(
            owner=request.user, file_type='batch', batch_name=batch_name
        ).order_by('id')

        if not files.exists():
            return Response({'error': '批次无文件'}, status=404)

        phases = []
        file_data_list = []
        bin_agg = defaultdict(int)
        site_agg = defaultdict(lambda: {'pass': 0, 'total': 0})
        all_sites = set()

        for idx, df_obj in enumerate(files):
            df, metadata, fmt = get_cached_parsed_file(df_obj.id, request.user.pk)
            if df is None:
                continue

            total = df.shape[0]
            bin_stats = calculate_fail_bin_statistics(df, metadata)

            pass_count = 0
            bin_info = []
            for bin_name, binfo in bin_stats.items():
                count = binfo.get('count', 0)
                pct = binfo.get('percentage', 0)
                bn = str(bin_name)
                if bn in ('1', 'Bin1'):
                    pass_count = count
                bin_agg[bn] += count
                bin_info.append({
                    'name': bn,
                    'value': count,
                    'pct': f"{pct:.2f}%",
                    'sites': {},
                })

            fail_count = total - pass_count
            yield_pct = round(pass_count / total * 100, 2) if total > 0 else 0

            # Site yield
            site_col = get_site_column(df)
            site_yields = {}
            site_total_map = {}
            site_pass_map = {}
            if site_col:
                bin_col = get_bin_column_name(df_obj.format_type)
                if bin_col in df.columns:
                    yd = compute_site_yield_data(df, bin_col, site_col)
                    for d in yd.get('yield_data', []):
                        site = d['Site']
                        site_yields[site] = d['Yield']
                        site_total_map[site] = d.get('Total', 0)
                        site_pass_map[site] = d.get('PassCount', 0)
                        site_agg[site]['pass'] += d.get('PassCount', 0)
                        site_agg[site]['total'] += d.get('Total', 0)
                        all_sites.add(site)

                    # Populate per-bin site data
                    site_breakdown = yd.get('site_breakdown', {})
                    for bi in bin_info:
                        for site in all_sites:
                            bi['sites'][site] = site_breakdown.get(site, {}).get(bi['name'], 0)

            # Detect phase type from filename
            # Priority: compound (EQC1-QA1, FT1-RT1, EQC1-2D3D) > standard (_CP1_, _FT1_)
            fname_upper = df_obj.filename.upper()
            phase_type = None
            # 1) Compound: EQC1-QA2-1 → QA2, FT1-RT1-1 → RT1, EQC1-2D3D → 2D3D
            m = re.search(r'(?:EQC\d+|QEC\d+|FT\d+)-(CP\d+|FT\d+|QA\d+|RT\d+|BF\d+|2D3D)', fname_upper)
            if m:
                phase_type = m.group(1)
            else:
                # 2) Standard: _CP1_, _FT1_, _QA1_ — use underscore delimiter to avoid CP0283 (program code)
                m = re.search(r'_(CP|FT|QA|RT|BF)(\d+)_', fname_upper)
                if m:
                    phase_type = f'{m.group(1)}{m.group(2)}'
            if not phase_type:
                phase_type = 'UNKNOWN'

            # Extract WAFER_ID from CP filenames: _05_CP1_ → 05
            wafer_id = ''
            if phase_type.startswith('CP'):
                wm = re.search(r'_(\d{2})_CP', fname_upper)
                if wm:
                    wafer_id = wm.group(1)

            phases.append({
                'filename': df_obj.filename,
                'phase': phase_type,
                'wafer_id': wafer_id,
                'lot_id': metadata.get('lot_id', '') or df_obj.program_name or '-',
                'total': total,
                'pass_count': pass_count,
                'fail_count': fail_count,
                'yield_pct': yield_pct,
                'start_time': metadata.get('start_time', '') or (str(df_obj.created_at)[:19] if hasattr(df_obj, 'created_at') else '-'),
                'end_time': metadata.get('end_time', '') or '-',
                'operator': metadata.get('operator', '') or '-',
                'station': metadata.get('station', '') or '-',
                'device_name': metadata.get('device_name', '') or '-',
                'tester_type': metadata.get('tester_type', '') or '-',
                'test_type': metadata.get('test_type', '') or '-',
                'total_test_time': metadata.get('total_test_time', '') or '-',
                'handler': metadata.get('handler', '') or '-',
                'site_yields': site_yields,
                'site_total': site_total_map,
                'site_pass': site_pass_map,
                'bin_info': bin_info,
                'program_name': df_obj.program_name or '',
            })

            file_data_list.append({
                'df': df,
                'metadata': metadata,
                'file_id': df_obj.id,
                'filename': df_obj.filename,
                'timestamp': str(getattr(df_obj, 'created_at', '')),
            })

        if not phases:
            return Response({'error': '无法解析批次文件'}, status=400)

        # KPI — overall yield (CP+FT input based, 2D3D counts as FT)
        cp_bin1 = sum(p['pass_count'] for p in phases if p['phase'].startswith('CP'))
        ft_bin1 = sum(p['pass_count'] for p in phases if p['phase'].startswith('FT') or p['phase'] == '2D3D')
        rt_bin1 = sum(p['pass_count'] for p in phases if p['phase'].startswith('RT'))
        input_total = sum(p['total'] for p in phases if p['phase'].startswith('CP') or p['phase'].startswith('FT') or p['phase'] == '2D3D')
        if input_total == 0:
            input_total = sum(p['total'] for p in phases)
        overall_yield = round((cp_bin1 + ft_bin1 + rt_bin1) / input_total * 100, 2) if input_total > 0 else 0

        total_all = sum(p['total'] for p in phases)
        pass_all = sum(p['pass_count'] for p in phases)
        fail_all = total_all - pass_all

        # Yield trend via SPC
        from apps.analysis.services.statistics import compute_yield_trend
        trend = compute_yield_trend(file_data_list)

        # Site pass data for chart
        site_pass_data = []
        for site in sorted(all_sites):
            d = site_agg[site]
            site_pass_data.append({
                'site': site,
                'pass': d['pass'],
                'total': d['total'],
                'yield': round(d['pass'] / d['total'] * 100, 2) if d['total'] > 0 else 0,
            })

        # Bin distribution for pie chart
        bin_distribution = []
        for bn in sorted(bin_agg.keys(), key=lambda x: (x not in ('1', 'Bin1'), x)):
            bin_distribution.append({'name': bn, 'value': bin_agg[bn]})

        # Site matrix table data
        sorted_sites = sorted(all_sites)
        site_matrix = []
        for p in phases:
            row = {'phase': p['phase']}
            for site in sorted_sites:
                st_val = p['site_total'].get(site, 0)
                sp_val = p['site_pass'].get(site, 0)
                sy = p['site_yields'].get(site, 0)
                try:
                    sy_val = float(sy)
                    row[f'{site}_yield'] = f"{sy_val:.2f}%" if st_val > 0 else 'N/A'
                except (ValueError, TypeError):
                    row[f'{site}_yield'] = 'N/A'
                row[f'{site}_ratio'] = f"{sp_val}/{st_val}"
            row['all_yield'] = f"{p['yield_pct']:.2f}%"
            row['all_ratio'] = f"{p['pass_count']}/{p['total']}"
            site_matrix.append(row)

        # Sort phases by semiconductor test flow: CP → FT/BF → RT → QA → 2D3D
        PHASE_ORDER = {'CP': 0, 'FT': 1, 'BF': 2, 'RT': 3, 'QA': 4, '2D3D': 5, 'UNKNOWN': 99}

        def phase_sort_key(p):
            name = p['phase']
            for prefix, order in PHASE_ORDER.items():
                if name.startswith(prefix):
                    # Secondary sort by number (CP1 < CP2, QA1 < QA2)
                    m = re.search(r'(\d+)', name)
                    num = int(m.group(1)) if m else 0
                    return (order, num)
            return (99, 0)

        phases.sort(key=phase_sort_key)

        # QA validation
        qa_checks = []
        qa1 = next((p for p in phases if re.match(r'^QA1(_|$)', p['phase'], re.IGNORECASE)), None)
        qa_phases = [p for p in phases if p['phase'].startswith('QA')]
        if qa1 and len(qa_phases) > 1:
            all_qa_pass = sum(p['pass_count'] for p in qa_phases)
            diff = all_qa_pass - qa1['total']
            qa_names = [p['phase'] for p in qa_phases]
            qa_checks.append({
                'check': f"QA系列Pass总数 ({', '.join(qa_names)}) ↔ QA1 Total",
                'expected': str(qa1['total']),
                'actual': str(all_qa_pass),
                'status': '✅ 一致' if diff == 0 else f'⚠️ 差异 {diff} 颗',
            })

        return Response({
            'batch_name': batch_name,
            'kpi': {
                'file_count': len(phases),
                'input_total': input_total,
                'total': total_all,
                'pass': pass_all,
                'fail': fail_all,
                'overall_yield': overall_yield,
                'yield_pct': overall_yield,
            },
            'phases': phases,
            'phase_summary': _build_phase_summary(phases),
            'trend_data': trend,
            'site_pass_data': site_pass_data,
            'bin_distribution': bin_distribution,
            'sorted_sites': sorted_sites,
            'site_matrix': site_matrix,
            'qa_checks': qa_checks,
        })

    @action(detail=False, methods=['post'])
    def generate_report(self, request):
        file_ids = request.data.get('file_ids', [])
        if not file_ids:
            return Response({'error': 'no_files'}, status=400)

        phases = []
        for fid in file_ids:
            df_obj = DataFile.objects.get(pk=fid, owner=request.user)
            df, metadata, fmt = get_cached_parsed_file(int(fid), request.user.pk)
            if df is None:
                continue

            total_rows = df.shape[0]
            bin_stats = calculate_fail_bin_statistics(df, metadata)
            total_pass = sum(1 for bv, s in bin_stats.items() if int(float(bv)) == 1)

            phases.append({
                'filename': df_obj.filename,
                'program_name': df_obj.program_name,
                'format': df_obj.format_type,
                'total': total_rows,
                'pass_count': total_pass,
                'fail_count': total_rows - total_pass,
                'yield_pct': round((total_pass / total_rows * 100), 2) if total_rows > 0 else 0,
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Report"
        hf = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        hfn = Font(bold=True, color="FFFFFF")
        ca = Alignment(horizontal="center", vertical="center")

        headers = ['文件名', '程序', '格式', '总数', 'Pass', 'Fail', '良率']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hf
            cell.font = hfn
            cell.alignment = ca

        for r, p in enumerate(phases, 2):
            for c, v in enumerate([p['filename'], p['program_name'], p['format'], p['total'], p['pass_count'], p['fail_count'], f"{p['yield_pct']}%"], 1):
                ws.cell(row=r, column=c, value=v)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Batch_Report.xlsx',
                           content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
