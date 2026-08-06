"""Unit tests for batch_report batch-level aggregation helpers.

Covers the pure functions in ``apps.batch_report.aggregation`` that roll per-phase
structures up to batch level for the batch report (Bin x Site cross table + UPH),
matching the single-file analysis output shapes consumed by the frontend.
"""
from django.test import TestCase

from apps.batch_report.aggregation import (
    aggregate_bin_site_table,
    aggregate_uph,
)


def _phase_with_bins(bin_info):
    return {'phase': 'CP1', 'bin_info': bin_info}


def _filename_from_cd(cd: str) -> str:
    """提取 Content-Disposition 的 filename*= 或 filename= 文件名。"""
    star = re.search(r"filename\*\s*=\s*(?:UTF-8'')?([^;]+)", cd)
    if star:
        return star.group(1).strip().strip('"')
    plain = re.search(r'filename\s*=\s*"?([^";]+)"?', cd)
    return plain.group(1).strip() if plain else cd


class AggregateBinSiteTableTests(TestCase):
    def setUp(self):
        # 2 phases, 2 sites (A, B), 2 bins (pass '1', fail '2').
        # Phase 1: Bin1 -> A:10 B:20 ; Bin2 -> A:1 B:3
        # Phase 2: Bin1 -> A:5  B:7  ; Bin2 -> A:2 B:0
        self.phases = [
            _phase_with_bins([
                {'name': '1', 'sites': {'A': 10, 'B': 20}},
                {'name': '2', 'sites': {'A': 1, 'B': 3}},
            ]),
            _phase_with_bins([
                {'name': '1', 'sites': {'A': 5, 'B': 7}},
                {'name': '2', 'sites': {'A': 2, 'B': 0}},
            ]),
        ]
        self.sorted_sites = ['A', 'B']

    def test_columns_match_sorted_sites(self):
        _, cols = aggregate_bin_site_table(self.phases, self.sorted_sites)
        self.assertEqual(cols, ['A', 'B'])

    def test_per_cell_sums_and_row_totals(self):
        rows, _ = aggregate_bin_site_table(self.phases, self.sorted_sites)
        by_bin = {r['bin']: r for r in rows}

        # Bin 1 (pass): A = 10+5 = 15, B = 20+7 = 27, all_site = 42
        self.assertEqual(by_bin['Bin 1']['A'], 15)
        self.assertEqual(by_bin['Bin 1']['B'], 27)
        self.assertEqual(by_bin['Bin 1']['all_site'], 42)

        # Bin 2 (fail): A = 1+2 = 3, B = 3+0 = 3, all_site = 6
        self.assertEqual(by_bin['Bin 2']['A'], 3)
        self.assertEqual(by_bin['Bin 2']['B'], 3)
        self.assertEqual(by_bin['Bin 2']['all_site'], 6)

    def test_total_row(self):
        rows, _ = aggregate_bin_site_table(self.phases, self.sorted_sites)
        total = next(r for r in rows if r['bin'] == 'Total')
        # A column total = 15 + 3 = 18, B = 27 + 3 = 30, grand = 48
        self.assertEqual(total['A'], 18)
        self.assertEqual(total['B'], 30)
        self.assertEqual(total['all_site'], 48)

    def test_output_shape_matches_single_file(self):
        """Keys must match compute_bin_site_table so BinSiteCrossTable.vue renders.

        Single-file rows: {'bin': 'Bin N'|'Total', <site>: int, ..., 'all_site': int}.
        """
        rows, cols = aggregate_bin_site_table(self.phases, self.sorted_sites)
        # Total row is last; pass bin (Bin 1) is first.
        self.assertEqual(rows[0]['bin'], 'Bin 1')
        self.assertEqual(rows[-1]['bin'], 'Total')
        for r in rows:
            self.assertIn('bin', r)
            self.assertIn('all_site', r)
            for c in cols:
                self.assertIn(c, r)
                self.assertIsInstance(r[c], int)

    def test_empty_when_no_sites(self):
        rows, cols = aggregate_bin_site_table(self.phases, [])
        self.assertEqual(rows, [])
        self.assertEqual(cols, [])

    def test_empty_when_no_bin_info(self):
        rows, cols = aggregate_bin_site_table([{'phase': 'CP1', 'bin_info': []}], ['A'])
        self.assertEqual(rows, [])
        self.assertEqual(cols, [])


class AggregateUphTests(TestCase):
    def _phase_uph(self, total_tested, total_time_seconds, avg_test_time,
                   by_site=None, source='col', warnings=None):
        return {'uph': {
            'total_tested': total_tested,
            'total_time_seconds': total_time_seconds,
            'avg_test_time': avg_test_time,
            'by_site': by_site or [],
            'source': source,
            'site_count': len(by_site or []),
            'warnings': warnings or [],
        }}

    def test_totals_and_uph(self):
        # Phase 1: 100 units, 50s wall-clock (avg 1.0s serial, 2 sites)
        # Phase 2: 200 units, 200s wall-clock (avg 2.0s serial, 2 sites)
        phases = [
            self._phase_uph(100, 50.0, 1.0,
                            by_site=[{'site': '1', 'tested': 50, 'uph': 3600.0},
                                     {'site': '2', 'tested': 50, 'uph': 3600.0}]),
            self._phase_uph(200, 200.0, 2.0,
                            by_site=[{'site': '1', 'tested': 100, 'uph': 1800.0},
                                     {'site': '2', 'tested': 100, 'uph': 1800.0}]),
        ]
        agg = aggregate_uph(phases)

        self.assertEqual(agg['total_tested'], 300)
        self.assertAlmostEqual(agg['total_time_seconds'], 250.0, places=1)
        # uph = 300 / 250 * 3600 = 4320
        self.assertAlmostEqual(agg['uph'], 4320.0, places=1)
        # avg_test_time = (1.0*100 + 2.0*200) / 300 = 500/300 = 1.667
        self.assertAlmostEqual(agg['avg_test_time'], 1.667, places=3)
        self.assertEqual(agg['source'], 'batch')

    def test_by_site_aggregation(self):
        phases = [
            self._phase_uph(100, 50.0, 1.0,
                            by_site=[{'site': '1', 'tested': 50, 'uph': 3600.0},
                                     {'site': '2', 'tested': 50, 'uph': 3600.0}]),
            self._phase_uph(200, 200.0, 2.0,
                            by_site=[{'site': '1', 'tested': 100, 'uph': 1800.0},
                                     {'site': '2', 'tested': 100, 'uph': 1800.0}]),
        ]
        agg = aggregate_uph(phases)
        by_site = {s['site']: s for s in agg['by_site']}

        # Site 1: tested = 50 + 100 = 150.
        # serial = 50*3600/3600 + 100*3600/1800 = 50 + 200 = 250s
        # uph = 3600 * 150 / 250 = 2160
        self.assertEqual(by_site['1']['tested'], 150)
        self.assertAlmostEqual(by_site['1']['uph'], 2160.0, places=1)
        self.assertEqual(by_site['2']['tested'], 150)
        self.assertAlmostEqual(by_site['2']['uph'], 2160.0, places=1)
        self.assertEqual(agg['site_count'], 2)

    def test_partial_data_warning(self):
        phases = [
            self._phase_uph(100, 50.0, 1.0,
                            by_site=[{'site': '1', 'tested': 100, 'uph': 3600.0}]),
            # Phase missing UPH (e.g. no test time column)
            {'uph': {'total_tested': 0, 'total_time_seconds': 0.0,
                     'avg_test_time': 0.0, 'by_site': [], 'source': 'unavailable',
                     'site_count': 0, 'warnings': ['未找到测试时间列，无法计算 UPH']}},
        ]
        agg = aggregate_uph(phases)
        self.assertEqual(agg['total_tested'], 100)
        # Partial warning present
        self.assertTrue(any('部分汇总' in w for w in agg['warnings']))
        # Per-phase warning merged
        self.assertTrue(any('未找到测试时间列' in w for w in agg['warnings']))

    def test_all_missing_returns_zero(self):
        phases = [
            {'uph': {'total_tested': 0, 'total_time_seconds': 0.0,
                     'avg_test_time': 0.0, 'by_site': [], 'source': 'unavailable',
                     'site_count': 0, 'warnings': []}},
        ]
        agg = aggregate_uph(phases)
        self.assertEqual(agg['total_tested'], 0)
        self.assertEqual(agg['uph'], 0.0)
        self.assertEqual(agg['by_site'], [])
        self.assertEqual(agg['source'], 'batch')

    def test_output_shape(self):
        phases = [
            self._phase_uph(100, 50.0, 1.0,
                            by_site=[{'site': '1', 'tested': 100, 'uph': 3600.0}]),
        ]
        agg = aggregate_uph(phases)
        for key in ('uph', 'avg_test_time', 'total_tested', 'total_time_seconds',
                    'source', 'by_site', 'site_count', 'warnings'):
            self.assertIn(key, agg)
        for s in agg['by_site']:
            self.assertEqual(set(s.keys()), {'site', 'tested', 'uph'})


# ── generate_report API 测试（2026-08-04 迁移 excelize 后新增）──

import io
import os
import re

import openpyxl
from django.contrib.auth import get_user_model
from rest_framework.test import APITestCase

from apps.datafiles.models import DataFile

User = get_user_model()

SAMPLE_GAGE = os.path.join(
    os.path.dirname(__file__), '..', '..', 'Data', 'SampleData', 'Gage', 'gage_m_S1.csv',
)


class GenerateReportApiTests(APITestCase):
    """POST /api/v1/batch-report/generate_report/：excelize 生成批次报表。"""

    def setUp(self):
        self.user = User.objects.create_user(username='batchuser', password='pw')
        self.client.force_authenticate(self.user)
        self.f1 = DataFile.objects.create(
            owner=self.user, filename='gage_m_S1.csv', file_path=SAMPLE_GAGE,
            file_size=os.path.getsize(SAMPLE_GAGE), format_type='CTA8290D',
            file_type='batch', batch_name='BATCH1', status='ready',
        )

    def test_generate_report_excel(self):
        """200 + FileResponse 头 + 7 列表头与数据行可读回。"""
        resp = self.client.post('/api/v1/batch-report/generate_report/',
                                {'file_ids': [self.f1.id]}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', resp['Content-Disposition'])
        buf = b''.join(resp.streaming_content)
        ws = openpyxl.load_workbook(io.BytesIO(buf))['Batch Report']
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        self.assertEqual(headers, ['文件名', '程序', '格式', '总数', 'Pass', 'Fail', '良率'])
        self.assertEqual(ws.max_row, 2, '表头 + 1 个数据文件行')
        self.assertEqual(ws.cell(2, 1).value, 'gage_m_S1.csv')
        self.assertIn('%', ws.cell(2, 7).value, '良率列应为百分比字符串')
        # 表头样式（excelize make_header_style 深色底）
        self.assertEqual(ws.cell(1, 1).fill.start_color.rgb.upper(), 'FF2C3E50')

    def test_generate_report_no_files(self):
        """file_ids 为空 → 400。"""
        resp = self.client.post('/api/v1/batch-report/generate_report/',
                                {'file_ids': []}, format='json')
        self.assertEqual(resp.status_code, 400)


class GenerateReportFilenameTemplateTests(APITestCase):
    """批次报表导出文件名模板。"""

    def setUp(self):
        self.user = User.objects.create_user(username='batchuser', password='pw')
        self.client.force_authenticate(self.user)
        self.files = []
        for i in range(2):
            f = DataFile.objects.create(
                owner=self.user, filename=f'gage_m_S{i + 1}.csv', file_path=SAMPLE_GAGE,
                file_size=os.path.getsize(SAMPLE_GAGE), format_type='CTA8290D',
                file_type='batch', batch_name='BATCH1', status='ready',
            )
            self.files.append(f)

    def test_default_template_contains_datetime(self):
        import re as _re
        resp = self.client.post('/api/v1/batch-report/generate_report/',
                                {'file_ids': [f.id for f in self.files]}, format='json')
        self.assertEqual(resp.status_code, 200)
        m = _re.search(r'Batch_Report_\d{8}_\d{6}\.xlsx', resp['Content-Disposition'])
        self.assertIsNotNone(m, f'默认模板应含日期时间戳: {resp["Content-Disposition"]}')

    def test_custom_template_with_batch_name_and_file_count(self):
        resp = self.client.put('/api/v1/auth/settings/',
                               {'export_filename_templates': {
                                   'batch_report': 'BR_{batch_name}_{file_count}'}},
                               format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        resp = self.client.post('/api/v1/batch-report/generate_report/',
                                {'file_ids': [f.id for f in self.files],
                                 'batch_name': 'BATCH_001'}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(_filename_from_cd(resp['Content-Disposition']), 'BR_BATCH_001_2.xlsx')


class BatchYieldPhaseParsingTests(APITestCase):
    """GET /api/v1/batch-report/batch_yield_data/：UIS 阶段解析 + 分阶段良率 + 短文件名回退。"""

    def setUp(self):
        self.user = User.objects.create_user(username='phaseuser', password='pw')
        self.client.force_authenticate(self.user)
        self.batch = 'UISBATCH'
        for fname in (
            'LOT_UISBATCH_CP1_20260726.csv',
            'BPD80590_C01JC6#AAA1A12606290025_UIS1.0_P262702101_20260715001944.csv',
            'LOT_UISBATCH_FT1_20260726.csv',
        ):
            DataFile.objects.create(
                owner=self.user, filename=fname, file_path=SAMPLE_GAGE,
                file_size=os.path.getsize(SAMPLE_GAGE), format_type='CTA8290D',
                file_type='batch', batch_name=self.batch, status='ready',
            )

    def _fetch(self, batch_name=None):
        return self.client.get(
            f'/api/v1/batch-report/batch_yield_data/?batch_name={batch_name or self.batch}'
        )

    def test_uis_phase_parsed_and_ordered_cp_uis_ft(self):
        resp = self._fetch()
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            [p['phase'] for p in resp.data['phases']],
            ['CP1', 'UIS1.0', 'FT1'],
            '阶段应按 CP → UIS → FT 排序',
        )
        self.assertIn('UIS1.0', [s['phase'] for s in resp.data['phase_summary']])
        # 增量字段：phase 与 phase_summary 均带 stage（前端过滤/树形分组用）
        self.assertEqual([p['stage'] for p in resp.data['phases']], ['CP', 'UIS', 'FT'])
        self.assertEqual(
            {s['phase']: s['stage'] for s in resp.data['phase_summary']},
            {'CP1': 'CP', 'UIS1.0': 'UIS', 'FT1': 'FT'},
        )

    def test_stage_yields_cp_uis_ft_separate(self):
        resp = self._fetch()
        stages = resp.data['stage_yields']
        self.assertEqual([s['stage'] for s in stages], ['CP', 'UIS', 'FT'])
        self.assertEqual([s['file_count'] for s in stages], [1, 1, 1])
        total_all = sum(p['total'] for p in resp.data['phases'])
        self.assertEqual(sum(s['total'] for s in stages), total_all)
        for s in stages:
            self.assertEqual(s['pass_count'] + s['fail_count'], s['total'])
        # UIS 独立成行、不混入整体良率 KPI（input_total 仅 CP+FT）
        self.assertEqual(resp.data['kpi']['input_total'],
                         sum(p['total'] for p in resp.data['phases']
                             if p['phase'].startswith(('CP', 'FT'))))

    def test_unrecognized_phase_falls_back_to_short_filename(self):
        fname = 'BPD80590_C01JC6#AAA1A12606290025_HW_P262702101_20260726200657.csv'
        DataFile.objects.create(
            owner=self.user, filename=fname, file_path=SAMPLE_GAGE,
            file_size=os.path.getsize(SAMPLE_GAGE), format_type='CTA8290D',
            file_type='batch', batch_name='HWBATCH', status='ready',
        )
        resp = self._fetch('HWBATCH')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['phases'][0]['phase'],
                         'BPD80590_C01JC6#AAA1A12606290025_HW_P262702101_20260726200657')
        self.assertEqual(resp.data['stage_yields'][0]['stage'], '其他')

    def test_rt_qa_fold_into_ft_stage(self):
        """RT/QA 与 FT 同属 FT 阶段：分阶段良率只出现 FT 一行聚合三阶段文件。"""
        for fname in (
            'LOT_FTGROUP_RT1_20260726.csv',
            'LOT_FTGROUP_QA1_20260726.csv',
            'LOT_FTGROUP_FT1_20260726.csv',
        ):
            DataFile.objects.create(
                owner=self.user, filename=fname, file_path=SAMPLE_GAGE,
                file_size=os.path.getsize(SAMPLE_GAGE), format_type='CTA8290D',
                file_type='batch', batch_name='FTGROUP', status='ready',
            )
        resp = self._fetch('FTGROUP')
        self.assertEqual(resp.status_code, 200)
        # 三个阶段文件都归入 FT stage
        self.assertEqual([s['stage'] for s in resp.data['stage_yields']], ['FT'])
        self.assertEqual(resp.data['stage_yields'][0]['file_count'], 3)
        # phase 明细级阶段名保持独立（FT1/RT1/QA1 不合并）
        self.assertEqual(
            [p['phase'] for p in resp.data['phases']],
            ['FT1', 'RT1', 'QA1'],
            '阶段明细仍按 FT → RT → QA 排序',
        )
        self.assertEqual([p['stage'] for p in resp.data['phases']], ['FT', 'FT', 'FT'])
